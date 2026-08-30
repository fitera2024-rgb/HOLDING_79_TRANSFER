#!/usr/bin/env python3
"""Run the Issue #13 synthetic pilot acceptance package.

The command deliberately uses only the repository's synthetic workbook.  It
builds the primary result in a private staging directory and publishes the
requested output directory only after all controls, regression probes, and a
deterministic rerun comparison pass.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
import tempfile
import zipfile
from collections import defaultdict
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "src"))

from holding79_transfer import (
    CONTROL_SHEET_NAMES,
    OUTPUT_HEADERS,
    OUTPUT_SHEET_NAME,
    BalanceStatus,
    IntegrationRunConfig,
    IntegrationRunError,
    OutputAdapterConfig,
    ParserDiagnosticCode,
    build_synthetic_osv_workbook,
    parse_grouped_osv,
    run_integration,
    validate_workbook_round_trip,
)

BASE_SHA = "296448216a62249eeeae15a40fd41234d5099c5a"
CONTRACT_VERSION = "0.3-approved"
RUN_INPUT_NAME = "synthetic-pilot-acceptance.xlsx"
EXPECTED_CASES = {
    "DEBIT_79_2_AT": "debit_79_2_AT.yaml",
    "CREDIT_79_2_AT": "credit_79_2_AT.yaml",
    "DEBIT_79_3_AT": "debit_79_3_AT.yaml",
    "CREDIT_79_3_AT": "credit_79_3_AT.yaml",
}
REQUIRED_ROOT_ARTIFACTS = {
    "input_manifest.json",
    "normalized_balances.jsonl",
    "posting_rows.jsonl",
    "summary.json",
    "run_control.xlsx",
    "export_manifest.json",
}
DECIMAL_RE = re.compile(r"^-?(?:0|[1-9]\d*)(?:\.\d+)?$")
FINANCIAL_ID_RE = re.compile(r"^FR-[0-9a-f]{64}$")


def _assert(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def _read_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    _assert(isinstance(value, dict), f"{path.name} must contain a JSON object")
    return value


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    records = [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines()]
    _assert(all(isinstance(record, dict) for record in records), f"{path.name} has a non-object row")
    return records


def _enum_value(value: Any) -> Any:
    return value.value if hasattr(value, "value") else value


def _golden_row(row: Any) -> dict[str, Any]:
    fields = (
        "document_organization",
        "debit_account",
        "debit_department",
        "debit_supplier_rvp",
        "credit_account",
        "credit_department",
        "credit_supplier_rvp",
        "amount",
    )
    result = {field: _enum_value(getattr(row, field)) for field in fields}
    result["amount"] = format(result["amount"], "f")
    return result


def _golden_expected(case_id: str) -> list[dict[str, Any]]:
    path = REPO_ROOT / "tests" / "golden" / EXPECTED_CASES[case_id]
    case = yaml.safe_load(path.read_text(encoding="utf-8"))
    _assert(case["status"] == "APPROVED", f"{case_id} is not APPROVED")
    return case["expected"]


def _source_bytes() -> bytes:
    workbook = build_synthetic_osv_workbook()
    stream = BytesIO()
    try:
        workbook.save(stream)
    finally:
        workbook.close()
    return stream.getvalue()


def _run_config() -> IntegrationRunConfig:
    return IntegrationRunConfig(
        period_end=date(2024, 12, 31),
        input_name=RUN_INPUT_NAME,
    )


def _verify_build() -> str:
    head = subprocess.run(
        ["git", "rev-parse", "HEAD"],
        cwd=REPO_ROOT,
        capture_output=True,
        text=True,
        check=False,
    )
    _assert(head.returncode == 0, "cannot identify the current build HEAD")
    head_sha = head.stdout.strip()
    _assert(bool(re.fullmatch(r"[0-9a-f]{40}", head_sha)), "current build HEAD is not a Git SHA")
    base_object = subprocess.run(
        ["git", "cat-file", "-e", f"{BASE_SHA}^{{commit}}"],
        cwd=REPO_ROOT,
        capture_output=True,
        check=False,
    )
    if base_object.returncode == 0:
        base_check = subprocess.run(
            ["git", "merge-base", "--is-ancestor", BASE_SHA, head_sha],
            cwd=REPO_ROOT,
            capture_output=True,
            text=True,
            check=False,
        )
        _assert(base_check.returncode == 0, f"accepted base SHA is not an ancestor of HEAD: {BASE_SHA}")
    else:
        # Product CI intentionally uses a shallow checkout.  In that case the
        # accepted commit object is unavailable, so require the PR workflow's
        # base ref instead of mutating the checkout with a fetch.
        _assert(
            os.environ.get("GITHUB_BASE_REF") == "main",
            f"accepted base SHA is unavailable and PR base ref is not main: {BASE_SHA}",
        )
    return head_sha


def _canonical_export_rows(path: Path) -> list[tuple[str, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        _assert(workbook.sheetnames == [OUTPUT_SHEET_NAME], f"{path.name} has helper sheets")
        worksheet = workbook[OUTPUT_SHEET_NAME]
        _assert(worksheet.max_column == len(OUTPUT_HEADERS), f"{path.name} is not 27 columns")
        headers = tuple(cell.value for cell in worksheet[1])
        _assert(headers == OUTPUT_HEADERS, f"{path.name} has the wrong header contract")
        rows: list[tuple[str, ...]] = []
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            canonical: list[str] = []
            for header, value in zip(OUTPUT_HEADERS, values, strict=True):
                if value is None or value == "":
                    canonical.append("")
                elif header == "СуммаВВалютеУчета":
                    canonical.append(format(Decimal(str(value)), "f"))
                else:
                    canonical.append(str(value))
            rows.append(tuple(canonical))
        return rows
    finally:
        workbook.close()


def _validate_artifacts_and_controls(root: Path, result: Any) -> dict[str, Any]:
    files = {path.name for path in root.iterdir() if path.is_file()}
    _assert(files == REQUIRED_ROOT_ARTIFACTS, f"root artifact set mismatch: {files!r}")
    _assert((root / "export").is_dir(), "export/ is missing")

    manifest = _read_json(root / "input_manifest.json")
    _assert(manifest["artifact_version"] == "H79_TRANSFER_RUN_V1", "wrong artifact version")
    _assert(manifest["contract_version"] == CONTRACT_VERSION, "wrong Product Contract version")
    _assert(manifest["input_name"] == RUN_INPUT_NAME, "wrong synthetic input identification")
    fingerprint = manifest["normalized_input_sha256"]
    _assert(isinstance(fingerprint, str) and re.fullmatch(r"[0-9a-f]{64}", fingerprint), "missing source SHA-256")
    expected_sheets = set(EXPECTED_CASES) | {"BLOCKED_SOURCE_ROW"}
    actual_sheets = {record["sheet_name"] for record in manifest["source_sheets"]}
    _assert(actual_sheets == expected_sheets, "synthetic source sheet identification changed")
    _assert(len(manifest["parser_diagnostics"]) == 1, "blocked source diagnostic is missing")
    blocked_diagnostic = manifest["parser_diagnostics"][0]
    _assert(blocked_diagnostic["code"] == "MISSING_SUPPLIER_RVP", "wrong blocked-source reason")
    blocked_ref = blocked_diagnostic["source_excel_row_ref"]

    normalized = _read_jsonl(root / "normalized_balances.jsonl")
    posting_records = _read_jsonl(root / "posting_rows.jsonl")
    _assert(len(normalized) == 4, "expected four actionable normalized balances")
    _assert({record["source_account"] for record in normalized} == {"79.2", "79.3"}, "source accounts changed")
    _assert(all(record["status"] == "ACTIONABLE" for record in normalized), "golden source is not actionable")
    _assert(all(isinstance(record["amount"], str) for record in posting_records), "posting amount is not Decimal text")
    _assert(all(DECIMAL_RE.fullmatch(record["amount"]) for record in posting_records), "non-canonical Decimal text")
    _assert(len(posting_records) == 8, "expected eight PostingRows")
    _assert(all(FINANCIAL_ID_RE.fullmatch(record["financial_record_id"]) for record in posting_records), "invalid financial id")
    _assert({record["source_account"] for record in posting_records} == {"79.2", "79.3"}, "PostingRow source accounts changed")

    rows_by_ref: dict[str, list[Any]] = defaultdict(list)
    for row in result.posting_rows:
        rows_by_ref[row.source_excel_row_ref or ""].append(row)
    _assert(set(rows_by_ref) == {f"{case}!R9" for case in EXPECTED_CASES}, "posting source references changed")
    _assert(all(len(rows) == 2 for rows in rows_by_ref.values()), "actionable source is not exactly 1:2")
    _assert(not any(row.source_excel_row_ref == blocked_ref for row in result.posting_rows), "blocked source produced output")

    for case_id in EXPECTED_CASES:
        source_ref = f"{case_id}!R9"
        actual = sorted((_golden_row(row) for row in rows_by_ref[source_ref]), key=lambda row: row["document_organization"])
        _assert(actual == _golden_expected(case_id), f"approved golden changed: {case_id}")

    summary = _read_json(root / "summary.json")
    _assert(summary["status"] == "SUCCESS", "summary did not report SUCCESS")
    _assert(summary["contract_version"] == CONTRACT_VERSION, "summary contract version mismatch")
    _assert(summary["counts"] == {
        "actionable_source_rows": 4,
        "blocked_source_rows": 1,
        "export_rows": 8,
        "export_workbooks": 2,
        "no_action_source_rows": 0,
        "normalized_balances": 4,
        "parser_diagnostics": 1,
        "posting_rows": 8,
        "source_rows": 5,
    }, "summary counts changed")
    _assert(summary["totals"] == {"difference": "0", "gk": "337089.60", "source_org": "337089.60"}, "source/GK totals changed")
    _assert(summary["controls"]["all_passed"] is True, "summary controls are not all PASS")

    control_workbook = load_workbook(root / "run_control.xlsx", read_only=True, data_only=False)
    try:
        _assert(control_workbook.sheetnames == list(CONTROL_SHEET_NAMES), "run control sheet set changed")
        effect_sheet = control_workbook["Контроль_до_после"]
        effect_headers = tuple(cell.value for cell in effect_sheet[1])
        effect_index = {name: index for index, name in enumerate(effect_headers)}
        effect_rows = list(effect_sheet.iter_rows(min_row=2, values_only=True))
        actionable_effects = [
            row for row in effect_rows if row[effect_index["status"]] == BalanceStatus.ACTIONABLE.value
        ]
        _assert(len(actionable_effects) == 4, "source-effect controls do not cover four goldens")
        _assert(all(
            row[effect_index["ending_debit_after"]] == "0"
            and row[effect_index["ending_credit_after"]] == "0"
            for row in actionable_effects
        ), "source-side 79.x after-effect is not zero")
        _assert(all(row[effect_index["source_posting_count"]] == 1 for row in actionable_effects), "source posting count changed")
        _assert(all(row[effect_index["source_effect_zero"]] == "True" for row in actionable_effects), "source-effect control failed")
    finally:
        control_workbook.close()

    export_manifest = _read_json(root / "export_manifest.json")
    _assert(export_manifest["sheet_name"] == OUTPUT_SHEET_NAME, "wrong export sheet name")
    _assert(export_manifest["headers"] == list(OUTPUT_HEADERS), "wrong 27-column export contract")
    _assert(export_manifest["financial_row_count"] == 8, "wrong export row count")
    export_paths = {record["path"] for record in export_manifest["workbooks"]}
    actual_paths = {path.relative_to(root).as_posix() for path in (root / "export").glob("*.xlsx")}
    _assert(actual_paths == export_paths == {"export/2024-12-31__АТ.xlsx", "export/2024-12-31__ГК.xlsx"}, "deterministic export filenames changed")

    adapter = OutputAdapterConfig(run_id=manifest["run_id"])
    export_rows: dict[str, list[tuple[str, ...]]] = {}
    for record in export_manifest["workbooks"]:
        path = root / Path(*record["path"].split("/"))
        document_rows = tuple(
            row for row in result.posting_rows
            if row.document_organization == record["document_organization"]
        )
        validate_workbook_round_trip(path, document_rows, adapter)
        export_rows[record["path"]] = _canonical_export_rows(path)
        _assert(len(export_rows[record["path"]]) == record["financial_row_count"], "export row count does not round-trip")

    return {
        "manifest": manifest,
        "summary": summary,
        "export_manifest": export_manifest,
        "export_rows": export_rows,
        "posting_records": posting_records,
        "blocked_ref": blocked_ref,
    }


def _assert_deterministic(first_root: Path, second_root: Path, first: dict[str, Any], second: dict[str, Any]) -> None:
    for name in ("normalized_balances.jsonl", "posting_rows.jsonl"):
        _assert(
            (first_root / name).read_bytes() == (second_root / name).read_bytes(),
            f"{name} differs on deterministic rerun",
        )
    _assert(first["manifest"]["run_id"] == second["manifest"]["run_id"], "run_id differs on rerun")
    _assert(first["manifest"]["normalized_input_sha256"] == second["manifest"]["normalized_input_sha256"], "source SHA-256 differs on rerun")
    _assert(first["posting_records"] == second["posting_records"], "PostingRows/rule IDs/financial IDs differ on rerun")
    _assert(first["export_manifest"]["workbooks"] == second["export_manifest"]["workbooks"], "deterministic filenames or export manifest differs")
    _assert(first["export_rows"] == second["export_rows"], "export row order or meaning differs on rerun")
    _assert(first["summary"]["totals"] == second["summary"]["totals"], "financial totals differ on rerun")


def _malformed_source_probes(source: bytes) -> None:
    for malformed in (b"not xlsx", b"PK\x03\x04", b""):
        result = parse_grouped_osv(malformed, period_end=date(2024, 12, 31))
        _assert(result.status is BalanceStatus.BLOCKED, "ordinary malformed XLSX was not blocked")
        _assert(result.diagnostics[0].code is ParserDiagnosticCode.INVALID_SOURCE, "ordinary malformed XLSX reason changed")

    malformed_source = BytesIO()
    with zipfile.ZipFile(BytesIO(source)) as original, zipfile.ZipFile(malformed_source, "w") as output:
        for info in original.infolist():
            content = original.read(info.filename)
            if info.filename == "xl/workbook.xml":
                _assert(b'sheetId="1"' in content, "synthetic workbook metadata fixture changed")
                content = content.replace(b'sheetId="1"', b'sheetId="not-an-integer"', 1)
            output.writestr(info, content)
    result = parse_grouped_osv(malformed_source.getvalue(), period_end=date(2024, 12, 31))
    _assert(result.status is BalanceStatus.BLOCKED, "invalid worksheet sheetId type was not blocked")
    _assert(result.diagnostics[0].code is ParserDiagnosticCode.INVALID_SOURCE, "invalid worksheet metadata reason changed")
    _assert(result.message == "source workbook is not a valid XLSX file", "raw workbook exception escaped parser")


def _failed_run_atomicity_probe(repo_root: Path, parent: Path) -> None:
    with tempfile.TemporaryDirectory(prefix=".pilot-failure-", dir=parent) as temp_name:
        output = Path(temp_name) / "failed-run"
        code = """
import sys
from datetime import date
sys.path.insert(0, sys.argv[1])
from holding79_transfer import IntegrationRunConfig, TransferConfig, build_synthetic_osv_workbook, run_integration
workbook = build_synthetic_osv_workbook()
try:
    run_integration(
        workbook,
        sys.argv[2],
        config=IntegrationRunConfig(
            period_end=date(2024, 12, 31),
            transfer_config=TransferConfig(rules_version="H79_TRANSFER_V2"),
        ),
    )
finally:
    workbook.close()
"""
        completed = subprocess.run(
            [sys.executable, "-c", code, str(repo_root / "src"), str(output)],
            capture_output=True,
            text=True,
            check=False,
        )
        combined_output = completed.stdout + completed.stderr
        _assert(completed.returncode != 0, "deliberate mandatory failure returned zero")
        _assert("mandatory control failed" in combined_output, "mandatory failure was not explicit")
        _assert("SUCCESS" not in combined_output, "failed process reported successful acceptance")
        _assert(not output.exists(), "failed run left partial published financial output")


def run(output_dir: Path) -> None:
    _assert(output_dir.name not in {"", ".", ".."}, "output directory must be dedicated")
    _assert(not output_dir.exists(), f"output directory already exists: {output_dir}")
    head_sha = _verify_build()
    output_dir.parent.mkdir(parents=True, exist_ok=True)
    source = _source_bytes()

    with tempfile.TemporaryDirectory(prefix=".pilot-acceptance-", dir=output_dir.parent) as temp_name:
        staging_root = Path(temp_name)
        first_root = staging_root / "first"
        first_result = run_integration(source, first_root, config=_run_config())
        first_evidence = _validate_artifacts_and_controls(first_root, first_result)

        second_root = staging_root / "second"
        second_result = run_integration(source, second_root, config=_run_config())
        second_evidence = _validate_artifacts_and_controls(second_root, second_result)
        _assert_deterministic(first_root, second_root, first_evidence, second_evidence)

        _malformed_source_probes(source)
        _failed_run_atomicity_probe(REPO_ROOT, staging_root)

        first_root.replace(output_dir)

    manifest = _read_json(output_dir / "input_manifest.json")
    print(f"BASE_SHA={BASE_SHA}")
    print(f"HEAD_SHA={head_sha}")
    print(f"PRODUCT_CONTRACT={CONTRACT_VERSION}")
    print(f"source={manifest['input_name']}; normalized_input_sha256={manifest['normalized_input_sha256']}")
    print("four_goldens=PASS")
    print("actionable_1_to_2=PASS")
    print("source_79_zero=PASS")
    print("source_org_equals_gk=PASS")
    print("blocked_zero_output=PASS")
    print("exact_27_columns=PASS")
    print("xlsx_round_trip=PASS")
    print("malformed_xlsx_invalid_source=PASS")
    print("failed_run_atomicity=PASS")
    print("deterministic_rerun=PASS")
    print(f"ARTIFACTS={len(list(output_dir.rglob('*')))} files/directories under {output_dir.name}")
    print("ACCEPTANCE=PASS")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output-dir",
        type=Path,
        required=True,
        help="dedicated, initially absent directory for the successful synthetic run",
    )
    args = parser.parse_args(argv)
    try:
        run(args.output_dir)
    except (AssertionError, IntegrationRunError, OSError, ValueError, zipfile.BadZipFile) as exc:
        print(f"ACCEPTANCE=FAIL: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
