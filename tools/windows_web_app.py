from __future__ import annotations

import json
import mimetypes
import re
import tempfile
import threading
import uuid
import webbrowser
from calendar import monthrange
from collections import defaultdict
from datetime import date
from email import policy
from email.parser import BytesParser
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import quote, urlparse
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook

from holding79_transfer import OUTPUT_SHEET_NAME, PostingRow, run_integration

APP_TITLE = "HOLDING 79 Transfer"
BUILD_MAIN_SHA = "b25da6ceb7a348187482d298f0ae38917dc7ec0e"
MAX_UPLOAD = 100 * 1024 * 1024
MONTH_RE = re.compile(r"^(\d{4})-(\d{2})$")
MONTHS = ("января", "февраля", "марта", "апреля", "мая", "июня", "июля", "августа", "сентября", "октября", "ноября", "декабря")

PAGE = r'''<!doctype html><html lang="ru"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>ФИТЭРА · HOLDING 79 Transfer</title><style>
:root{--fitera:#ff5634;--fitera-dark:#d83d22;--ink:#161616;--ink2:#343434;--muted:#747474;--soft:#f6f4f1;--card:#fff;--line:#e8e4df;--good:#26734d;--bad:#b42318;--warn:#9a6700;--shadow:0 20px 55px rgba(20,20,20,.10)}*{box-sizing:border-box}html{background:#0e0e0e}body{margin:0;background:linear-gradient(180deg,#101010 0,#101010 250px,var(--soft) 250px,var(--soft) 100%);font-family:Inter,"Segoe UI",Arial,sans-serif;color:var(--ink);min-height:100vh}.wrap{max-width:1100px;margin:auto;padding:0 24px 54px}.mast{height:88px;display:flex;justify-content:space-between;align-items:center;color:#fff;border-bottom:1px solid rgba(255,255,255,.12)}.wordmark{display:flex;align-items:center;gap:13px}.mark{width:38px;height:38px;border:2px solid var(--fitera);display:grid;place-items:center;font-weight:900;font-size:18px;letter-spacing:-1px;color:#fff;position:relative}.mark:after{content:"";position:absolute;width:10px;height:10px;background:var(--fitera);right:-5px;bottom:-5px}.fitera{font-size:21px;font-weight:900;letter-spacing:.19em}.era{font-size:11px;color:#a9a9a9;margin-top:3px;letter-spacing:.03em}.mast-meta{font-size:12px;color:#c4c4c4;text-align:right}.hero{padding:38px 0 30px;color:#fff;display:flex;justify-content:space-between;gap:30px;align-items:flex-end}.hero h1{margin:0;font-size:39px;letter-spacing:-.035em;line-height:1.05;max-width:670px}.hero h1 span{color:var(--fitera)}.hero p{margin:13px 0 0;color:#b8b8b8;line-height:1.55;font-size:14px;max-width:660px}.safe{display:flex;align-items:center;gap:8px;border:1px solid rgba(255,255,255,.15);border-radius:999px;padding:9px 13px;font-size:11px;color:#d5d5d5;white-space:nowrap}.safe i{width:8px;height:8px;background:#49b47c;border-radius:50%;box-shadow:0 0 0 4px rgba(73,180,124,.12)}.layout{display:grid;grid-template-columns:minmax(0,1.6fr) minmax(260px,.7fr);gap:18px;align-items:start}.card{background:var(--card);border:1px solid var(--line);box-shadow:var(--shadow)}.main-card{border-radius:6px;overflow:hidden}.head{padding:26px 28px 0}.eyebrow{font-size:10px;color:var(--fitera);font-weight:900;letter-spacing:.17em;text-transform:uppercase;margin-bottom:7px}.head h2{font-size:20px;margin:0;letter-spacing:-.02em}.head p{font-size:13px;color:var(--muted);line-height:1.55;margin:8px 0 0}.form{padding:23px 28px 28px;display:grid;grid-template-columns:1fr 220px;gap:18px}.field label{display:block;font-size:12px;font-weight:800;margin-bottom:8px;color:var(--ink2)}input{width:100%;min-height:50px;border:1px solid #d8d2cb;border-radius:3px;background:#fff;color:var(--ink);font:inherit;outline:none;transition:.15s}input[type=file]{padding:11px 12px}input[type=month]{padding:0 13px}input:focus{border-color:var(--fitera);box-shadow:0 0 0 3px rgba(255,86,52,.10)}.hint{font-size:11px;color:#999;margin-top:7px}.actions{grid-column:1/-1;border-top:1px solid var(--line);padding-top:20px;display:flex;align-items:center;justify-content:space-between;gap:15px}.status{font-size:12px;color:var(--muted)}button,a.btn{font:inherit;border:0;text-decoration:none;cursor:pointer}.primary{background:var(--fitera);color:#fff;padding:14px 21px;font-weight:850;min-width:230px;border-radius:2px;letter-spacing:.01em;box-shadow:0 9px 24px rgba(255,86,52,.20)}.primary:hover{background:var(--fitera-dark)}.primary:disabled{opacity:.55;cursor:default}.secondary{background:transparent;color:#555;padding:10px 0;font-size:12px;border-bottom:1px solid #bdb7b0}.progress{display:none;grid-column:1/-1;height:4px;overflow:hidden;background:#efebe7}.progress.on{display:block}.progress span{display:block;width:35%;height:100%;background:var(--fitera);animation:p 1.05s infinite ease-in-out}@keyframes p{from{transform:translateX(-110%)}to{transform:translateX(315%)}}.msg{display:none;grid-column:1/-1;padding:13px 15px;font-size:12px;line-height:1.5}.msg.err{display:block;background:#fff2ef;color:var(--bad);border-left:3px solid var(--bad)}.side-card{border-radius:6px;padding:25px}.side-number{font-size:48px;font-weight:900;letter-spacing:-.06em;color:var(--fitera);line-height:1}.side-title{font-weight:900;font-size:14px;margin:10px 0 7px}.side-text{font-size:12px;line-height:1.6;color:var(--muted)}.principles{margin-top:18px;border-top:1px solid var(--line);padding-top:16px}.principle{display:flex;gap:10px;margin:11px 0;font-size:11px;color:#656565;line-height:1.45}.principle:before{content:"";width:6px;height:6px;background:var(--fitera);margin-top:5px;flex:none}.result{display:none;margin-top:18px}.result.on{display:block}.result .card{border-radius:6px}.inside{padding:23px 28px 28px}.metrics{display:grid;grid-template-columns:repeat(4,1fr);gap:1px;background:var(--line);border:1px solid var(--line)}.metric{padding:16px;background:#fff}.metric b{display:block;font-size:26px;margin-bottom:2px;letter-spacing:-.04em}.metric span{font-size:10px;color:var(--muted);text-transform:uppercase;letter-spacing:.06em}.files{margin-top:18px;border:1px solid var(--line)}.row{display:flex;justify-content:space-between;gap:15px;align-items:center;padding:13px 15px;border-top:1px solid var(--line)}.row:first-child{border-top:0}.name{min-width:0}.name b{display:block;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;font-size:12px}.name small{display:block;color:#999;margin-top:4px;font-size:11px}.badge{display:inline-block;margin-left:8px;padding:3px 6px;background:#fff0c7;color:#8a5c00;font-size:9px;font-weight:900;letter-spacing:.05em}.download{color:var(--fitera-dark);background:#fff3ef;padding:8px 10px;font-weight:800;font-size:11px;white-space:nowrap}.all{display:flex;justify-content:flex-end;margin-top:16px}.foot{display:flex;justify-content:space-between;align-items:center;margin-top:20px;padding:0 2px;color:#8e8a84;font-size:10px}.fitera-sign{font-weight:800;letter-spacing:.12em;color:#4d4a47}@media(max-width:820px){body{background:linear-gradient(180deg,#101010 0,#101010 300px,var(--soft) 300px)}.wrap{padding:0 14px 40px}.mast{height:76px}.mast-meta{display:none}.hero{padding:28px 0 24px;flex-direction:column;align-items:flex-start}.hero h1{font-size:32px}.layout{grid-template-columns:1fr}.side-card{display:none}.form{grid-template-columns:1fr;padding:20px}.actions{grid-column:1;flex-direction:column;align-items:stretch}.primary{width:100%}.metrics{grid-template-columns:repeat(2,1fr)}}
</style></head><body><div class="wrap"><header class="mast"><div class="wordmark"><div class="mark">F</div><div><div class="fitera">ФИТЭРА</div><div class="era">Эра финансов и информационных технологий</div></div></div><div class="mast-meta">Финансовая автоматизация · 1С · аналитика</div></header><section class="hero"><div><h1>Перенос остатков <span>79.2 / 79.3</span> на 79.1</h1><p>Локальный инструмент ФИТЭРА для подготовки файлов проводок по ОСВ. Расчёты выполняются на компьютере пользователя и не записываются в 1С.</p></div><div class="safe"><i></i> Локально · без записи в 1С</div></section>
<div class="layout"><section class="card main-card"><div class="head"><div class="eyebrow">HOLDING 79 TRANSFER</div><h2>Сформировать файлы проводок</h2><p>Загрузите ОСВ и выберите месяц. Результат будет подготовлен отдельно по организациям.</p></div><form class="form" id="f"><div class="field"><label>Файл ОСВ</label><input name="source" type="file" accept=".xlsx" required><div class="hint">Формат Excel .xlsx</div></div><div class="field"><label>Период расчёта</label><input name="period" type="month" value="{{MONTH}}" required><div class="hint">Выберите месяц</div></div><div class="actions"><div class="status" id="status">Готово к запуску</div><button class="primary" id="go">Сформировать файлы</button></div><div class="progress" id="prog"><span></span></div><div class="msg" id="msg"></div></form></section><aside class="card side-card"><div class="side-number">79</div><div class="side-title">Контролируемый перенос</div><div class="side-text">Программа формирует проводки по утверждённой логике и сохраняет результат в стандартных файлах загрузки.</div><div class="principles"><div class="principle">Организация определяет отдельный выходной файл.</div><div class="principle">Файлы с неполной нижней аналитикой помечаются «СПОРНО».</div><div class="principle">Суммы, счета и стороны Дт/Кт не изменяются интерфейсом.</div></div></aside></div>
<section class="result" id="res"><div class="card"><div class="head"><div class="eyebrow">РЕЗУЛЬТАТ</div><h2>Файлы готовы</h2><p id="per"></p></div><div class="inside"><div class="metrics"><div class="metric"><b id="src">0</b><span>остатков</span></div><div class="metric"><b id="post">0</b><span>проводок</span></div><div class="metric"><b id="books">0</b><span>файлов</span></div><div class="metric"><b id="sp">0</b><span>спорно</span></div></div><div class="files" id="files"></div><div class="all"><a class="btn primary" id="zip">Скачать всё ZIP</a></div></div></div></section>
<div class="foot"><span><span class="fitera-sign">ФИТЭРА</span> · тестовая сборка · main {{SHA}}</span><button class="secondary" id="close">Закрыть программу</button></div></div><script>
const f=document.getElementById('f'),go=document.getElementById('go'),prog=document.getElementById('prog'),msg=document.getElementById('msg'),res=document.getElementById('res'),status=document.getElementById('status');function esc(s){const d=document.createElement('div');d.textContent=s??'';return d.innerHTML}f.addEventListener('submit',async e=>{e.preventDefault();msg.className='msg';msg.textContent='';res.classList.remove('on');go.disabled=true;prog.classList.add('on');status.textContent='Обработка ОСВ…';try{const r=await fetch('/api/run',{method:'POST',body:new FormData(f)}),d=await r.json();if(!r.ok||!d.ok)throw new Error(d.error||'Не удалось сформировать файлы.');document.getElementById('src').textContent=d.source_rows;document.getElementById('post').textContent=d.posting_rows;document.getElementById('books').textContent=d.workbooks.length;document.getElementById('sp').textContent=d.disputed_workbooks;document.getElementById('per').textContent='Период: '+d.period_label;document.getElementById('zip').href=d.zip_url;document.getElementById('files').innerHTML=d.workbooks.map(x=>`<div class="row"><div class="name"><b>${esc(x.name)}${x.disputed?'<span class="badge">СПОРНО</span>':''}</b><small>${x.rows} проводок</small></div><a class="btn download" href="${x.url}">Скачать</a></div>`).join('');res.classList.add('on');status.textContent='Файлы сформированы';res.scrollIntoView({behavior:'smooth'})}catch(x){msg.textContent=x.message||String(x);msg.className='msg err';status.textContent='Файлы не сформированы'}finally{prog.classList.remove('on');go.disabled=false}});document.getElementById('close').onclick=async()=>{if(!confirm('Закрыть HOLDING 79 Transfer?'))return;try{await fetch('/api/shutdown',{method:'POST'})}catch(_){ }document.body.innerHTML='<div style="font:16px Segoe UI;padding:40px">Программа закрыта. Эту вкладку можно закрыть.</div>'};
</script></body></html>'''


def period_end(text: str) -> tuple[date, str]:
    match = MONTH_RE.fullmatch(text.strip())
    if not match:
        raise ValueError("Выберите месяц расчёта.")
    year, month = int(match.group(1)), int(match.group(2))
    if not 1 <= month <= 12:
        raise ValueError("Некорректный месяц периода.")
    end = date(year, month, monthrange(year, month)[1])
    return end, f"{month:02d}.{year}"


def multipart(body: bytes, content_type: str) -> tuple[dict[str, str], dict[str, tuple[str, bytes]]]:
    prefix = f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode()
    message = BytesParser(policy=policy.default).parsebytes(prefix + body)
    fields: dict[str, str] = {}
    files: dict[str, tuple[str, bytes]] = {}
    if not message.is_multipart():
        return fields, files
    for part in message.iter_parts():
        if part.get_content_disposition() != "form-data":
            continue
        name = part.get_param("name", header="content-disposition")
        if not name:
            continue
        payload = part.get_payload(decode=True) or b""
        filename = part.get_filename()
        if filename:
            files[name] = (Path(filename).name, payload)
        else:
            fields[name] = payload.decode(part.get_content_charset() or "utf-8", errors="replace")
    return fields, files


def clean_content(row: PostingRow) -> str:
    side = "дебетового" if row.side and row.side.value == "DEBIT" else "кредитового"
    account = row.source_account.value if row.source_account else "79.x"
    period = f"{MONTHS[row.period_end.month - 1]} {row.period_end.year} года" if row.period_end else "выбранного периода"
    parts = [f"Перенос конечного {side} остатка по счету {account} на 79.1 за {period}.", f"Организация источника: {row.source_organization}."]
    if row.source_department:
        parts.append(f"ЦФО: {row.source_department}.")
    if row.source_supplier_rvp:
        parts.append(f"РВП: {row.source_supplier_rvp}.")
    if row.document_organization != row.source_organization:
        parts.append(f"Организация проводки: {row.document_organization}.")
    return " ".join(parts)


def verify_only_content_changed(original: Path, cleaned: Path, content_col: int) -> None:
    left, right = load_workbook(original, data_only=False), load_workbook(cleaned, data_only=False)
    try:
        if left.sheetnames != right.sheetnames:
            raise RuntimeError("Изменился состав листов выходного файла.")
        for name in left.sheetnames:
            a, b = left[name], right[name]
            if a.max_row != b.max_row or a.max_column != b.max_column:
                raise RuntimeError("Изменилась размерность выходного файла.")
            for r in range(1, a.max_row + 1):
                for c in range(1, a.max_column + 1):
                    if name == OUTPUT_SHEET_NAME and r >= 2 and c == content_col:
                        continue
                    if a.cell(r, c).value != b.cell(r, c).value:
                        raise RuntimeError("Изменились данные проводки вне колонки Содержание.")
    finally:
        left.close(); right.close()


def prepare_downloads(result, folder: Path):
    folder.mkdir(parents=True, exist_ok=False)
    grouped: dict[tuple[object, str], list[PostingRow]] = defaultdict(list)
    for row in result.posting_rows:
        grouped[(row.period_end, row.document_organization)].append(row)
    prepared = []
    for exported in result.exported_workbooks:
        rows = grouped[(exported.document_date, exported.document_organization)]
        if len(rows) != exported.row_count:
            raise RuntimeError("Количество проводок не совпало при подготовке файла.")
        book = load_workbook(exported.path)
        try:
            if OUTPUT_SHEET_NAME not in book.sheetnames:
                raise RuntimeError("Нет листа Загрузка_A_AA.")
            sheet = book[OUTPUT_SHEET_NAME]
            headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
            try:
                content_col = headers.index("Содержание") + 1
            except ValueError as exc:
                raise RuntimeError("Нет колонки Содержание.") from exc
            if sheet.max_row - 1 != len(rows):
                raise RuntimeError("Количество строк файла не совпало с проводками.")
            for excel_row, posting in enumerate(rows, 2):
                sheet.cell(excel_row, content_col).value = clean_content(posting)
            target = folder / exported.path.name
            book.save(target)
        finally:
            book.close()
        verify_only_content_changed(exported.path, target, content_col)
        prepared.append((target, exported))
    return prepared


class Server(ThreadingHTTPServer):
    daemon_threads = True
    def __init__(self, address, root: Path):
        super().__init__(address, Handler); self.root = root; self.downloads = {}
    def add_download(self, path: Path) -> str:
        token = uuid.uuid4().hex; self.downloads[token] = path; return f"/download/{token}"


class Handler(BaseHTTPRequestHandler):
    server: Server
    def log_message(self, *_):
        return
    def send_bytes(self, data: bytes, content_type: str, status=200):
        self.send_response(status); self.send_header("Content-Type", content_type); self.send_header("Content-Length", str(len(data))); self.send_header("Cache-Control", "no-store"); self.end_headers(); self.wfile.write(data)
    def send_json(self, data: dict, status=200):
        self.send_bytes(json.dumps(data, ensure_ascii=False).encode(), "application/json; charset=utf-8", status)
    def do_GET(self):
        path = urlparse(self.path).path
        if path == "/":
            page = PAGE.replace("{{MONTH}}", date.today().strftime("%Y-%m")).replace("{{SHA}}", BUILD_MAIN_SHA[:12] + "…")
            return self.send_bytes(page.encode(), "text/html; charset=utf-8")
        if path.startswith("/download/"):
            item = self.server.downloads.get(path.rsplit("/", 1)[-1])
            if not item or not item.is_file():
                return self.send_json({"ok": False, "error": "Файл больше недоступен."}, HTTPStatus.NOT_FOUND)
            data = item.read_bytes(); ctype = mimetypes.guess_type(item.name)[0] or "application/octet-stream"
            self.send_response(200); self.send_header("Content-Type", ctype); self.send_header("Content-Length", str(len(data))); self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(item.name)}"); self.send_header("Cache-Control", "no-store"); self.end_headers(); self.wfile.write(data); return
        self.send_json({"ok": False, "error": "Страница не найдена."}, HTTPStatus.NOT_FOUND)
    def do_POST(self):
        path = urlparse(self.path).path
        if path == "/api/shutdown":
            self.send_json({"ok": True}); threading.Thread(target=self.server.shutdown, daemon=True).start(); return
        if path != "/api/run":
            return self.send_json({"ok": False, "error": "Команда не найдена."}, HTTPStatus.NOT_FOUND)
        try:
            length = int(self.headers.get("Content-Length", "0"))
            if length <= 0 or length > MAX_UPLOAD:
                raise ValueError("Файл пустой или больше 100 МБ.")
            ctype = self.headers.get("Content-Type", "")
            if "multipart/form-data" not in ctype:
                raise ValueError("Не удалось прочитать файл.")
            fields, files = multipart(self.rfile.read(length), ctype)
            if "source" not in files:
                raise ValueError("Выберите файл ОСВ .xlsx.")
            source_name, source_bytes = files["source"]
            if Path(source_name).suffix.lower() != ".xlsx":
                raise ValueError("Нужен файл Excel .xlsx.")
            end, label = period_end(fields.get("period", ""))
            root = self.server.root / uuid.uuid4().hex; root.mkdir()
            source = root / "source.xlsx"; source.write_bytes(source_bytes)
            result = run_integration(source, root / "result", period_end=end, input_name=source_name)
            prepared = prepare_downloads(result, root / "download")
            books = []; disputed = 0
            for clean, exported in prepared:
                is_sporno = "_СПОРНО" in clean.stem; disputed += int(is_sporno)
                books.append({"name": clean.name, "rows": exported.row_count, "disputed": is_sporno, "url": self.server.add_download(clean)})
            zip_path = root / "HOLDING79_TRANSFER_RESULT.zip"
            with ZipFile(zip_path, "w", ZIP_DEFLATED) as z:
                for clean, _ in prepared:
                    z.write(clean, clean.name)
                control = root / "result" / "run_control.xlsx"
                if control.is_file(): z.write(control, "Контроль_расчета.xlsx")
            self.send_json({"ok": True, "period_label": label, "source_rows": len(result.normalized_balances), "posting_rows": len(result.posting_rows), "disputed_workbooks": disputed, "workbooks": books, "zip_url": self.server.add_download(zip_path)})
        except Exception as exc:
            self.send_json({"ok": False, "error": str(exc).strip() or type(exc).__name__}, HTTPStatus.BAD_REQUEST)


def main():
    with tempfile.TemporaryDirectory(prefix="holding79-web-") as temp:
        server = Server(("127.0.0.1", 0), Path(temp)); host, port = server.server_address
        threading.Timer(.5, lambda: webbrowser.open(f"http://{host}:{port}/", new=1)).start()
        server.serve_forever(.25); server.server_close()


if __name__ == "__main__":
    main()