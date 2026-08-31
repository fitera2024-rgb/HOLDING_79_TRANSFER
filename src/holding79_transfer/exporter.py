"""Deterministic XLSX export for canonical :class:`PostingRow` values.

The exporter intentionally has no accounting logic of its own.  It delegates
the 27-column mapping to :mod:`holding79_transfer.output`, groups already
validated rows by their document organization and date, and validates every
workbook after reopening it.
"""

from __future__ import annotations

import re
import tempfile
from collections import Counter, defaultdict
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, TypeAlias

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .models import BalanceStatus, BlockReason, PostingRow
from .output import (
    OUTPUT_HEADERS,
    OUTPUT_SHEET_NAME,
    BlockedOutputMapping,
    OutputAdapterConfig,
    map_posting_row,
)

_INVALID_FILENAME_CHARACTERS = re.compile(r'[<>:"/\\|?*\x00-\x1f\x7f]')
_WINDOWS_RESERVED_NAMES = {
    "con",
    "prn",
    "aux",
    "nul",
    *(f"com{number}" for number in range(1, 10)),
    *(f"lpt{number}" for number in range(1, 10)),
}
_AMOUNT_COLUMN_INDEX = OUTPUT_HEADERS.index("СуммаВВалютеУчета")
_DEFAULT_AMOUNT_NUMBER_FORMAT = "0.############################"
_MAX_FILENAME_COMPONENT_LENGTH = 180


class XlsxExportError(RuntimeError):
    """Base error for fail-closed XLSX export failures."""


class ExportBlockedError(BlockedOutputMapping, XlsxExportError):
    """Raised when a PostingRow cannot be exported safely."""


class FilenameCollisionError(XlsxExportError):
    """Raised when distinct organization/date groups resolve to one filename."""


class ExistingOutputError(XlsxExportError):
    """Raised when export would silently replace an existing workbook."""


class DuplicatePostingRowError(XlsxExportError):
    """Raised when the input contains duplicate financial output rows."""


class RoundTripMismatchError(XlsxExportError):
    """Raised when a reopened workbook differs from the expected PostingRows."""


@dataclass(frozen=True)
class XlsxExportConfig:
    """Configuration for a batch of deterministic workbook exports.

    ``adapter_config`` is the accepted Issue #1 output adapter configuration.
    ``overwrite`` only controls a pre-existing *same* target path; collisions
    between distinct groups are always rejected.
    """

    output_dir: Path | str
    adapter_config: OutputAdapterConfig = field(default_factory=OutputAdapterConfig)
    overwrite: bool = False
    amount_number_format: str = _DEFAULT_AMOUNT_NUMBER_FORMAT

    def __post_init__(self) -> None:
        output_dir = Path(self.output_dir)
        if output_dir.exists() and not output_dir.is_dir():
            raise ValueError(f"output_dir is not a directory: {output_dir}")
        object.__setattr__(self, "output_dir", output_dir)
        if not isinstance(self.adapter_config, OutputAdapterConfig):
            raise TypeError("adapter_config must be an OutputAdapterConfig")
        if not isinstance(self.overwrite, bool):
            raise TypeError("overwrite must be a bool")
        if not isinstance(self.amount_number_format, str) or not self.amount_number_format:
            raise ValueError("amount_number_format must be a non-empty string")


# Short aliases make the public exporter API discoverable without introducing
# a second configuration contract.
ExportConfig = XlsxExportConfig
ExporterConfig = XlsxExportConfig


@dataclass(frozen=True)
class ExportedWorkbook:
    """Metadata for one workbook that passed round-trip validation."""

    path: Path
    document_organization: str
    document_date: date
    row_count: int


@dataclass(frozen=True)
class XlsxExportResult:
    """Successful export result, ordered by date and organization."""

    workbooks: tuple[ExportedWorkbook, ...] = ()

    @property
    def paths(self) -> tuple[Path, ...]:
        return tuple(workbook.path for workbook in self.workbooks)

    @property
    def files(self) -> tuple[Path, ...]:
        """Compatibility alias for callers that refer to output files."""

        return self.paths

    @property
    def status(self) -> BalanceStatus:
        return BalanceStatus.ACTIONABLE if self.workbooks else BalanceStatus.NO_ACTION

    @property
    def is_success(self) -> bool:
        return True

    def __len__(self) -> int:
        return len(self.workbooks)

    def __iter__(self):
        return iter(self.paths)

    def __getitem__(self, index):
        return self.paths[index]


@dataclass(frozen=True)
class _MappedPosting:
    source: PostingRow
    values: tuple[Any, ...]
    group: tuple[date, str]


PostingRows: TypeAlias = Iterable[PostingRow]


def sanitize_filename_component(value: str) -> str:
    """Return a deterministic Windows-safe filename component.

    The function does not attempt to make collisions unique.  The exporter
    must reject such collisions because adding an arbitrary suffix would make
    the filename-to-group mapping less auditable.
    """

    if not isinstance(value, str):
        raise TypeError("filename component must be a string")
    component = _INVALID_FILENAME_CHARACTERS.sub("_", value.strip())
    component = component.rstrip(" .")
    if not component:
        component = "_"
    if component.casefold() in _WINDOWS_RESERVED_NAMES:
        component = f"_{component}"
    return component[:_MAX_FILENAME_COMPONENT_LENGTH]


def deterministic_filename(
    document_date: date,
    document_organization: str,
    *,
    disputed: bool = False,
) -> str:
    """Build the stable safe filename for one date/organization group."""

    if not isinstance(document_date, date):
        raise TypeError("document_date must be a date")
    if not isinstance(document_organization, str) or not document_organization.strip():
        raise ValueError("document_organization must be a non-empty string")
    if not isinstance(disputed, bool):
        raise TypeError("disputed must be a bool")
    suffix = "_СПОРНО" if disputed else ""
    return (
        f"{document_date.isoformat()}__"
        f"{sanitize_filename_component(document_organization)}{suffix}.xlsx"
    )


def posting_has_blank_lower_analytics(row: PostingRow) -> bool:
    """Return whether an output posting has an absent department or supplier."""

    if not isinstance(row, PostingRow):
        raise TypeError("row must be a PostingRow")
    return any(
        not value
        for value in (
            row.debit_department,
            row.credit_department,
            row.debit_supplier_rvp,
            row.credit_supplier_rvp,
        )
    )


def _decimal_identity(value: Decimal) -> str:
    return format(value.normalize(), "f")


def _sort_key(mapped: _MappedPosting) -> tuple[str, ...]:
    row = mapped.source
    side = row.side.value if row.side else ""
    source_account = row.source_account.value if row.source_account else ""
    account_values = tuple(
        account.value if hasattr(account, "value") else str(account)
        for account in (row.debit_account, row.credit_account)
    )
    return (
        row.period_end.isoformat() if row.period_end else "",
        row.document_organization,
        row.financial_record_id,
        row.source_excel_row_ref or "",
        side,
        source_account,
        *account_values,
        row.debit_department,
        row.credit_department,
        row.debit_supplier_rvp,
        row.credit_supplier_rvp,
        _decimal_identity(row.amount),
    )


def _canonical_amount(value: Any) -> Decimal:
    """Read an XLSX amount without using binary-float arithmetic.

    openpyxl exposes ordinary numeric XLSX cells as ``float`` on read.  The
    value is converted through its decimal text representation immediately;
    all comparison and business values remain Decimal.
    """

    if isinstance(value, float):
        value = repr(value)
    if isinstance(value, Decimal):
        amount = value
    elif isinstance(value, (int, str)):
        try:
            amount = Decimal(str(value).strip())
        except InvalidOperation as exc:
            raise RoundTripMismatchError(f"invalid XLSX amount: {value!r}") from exc
    else:
        raise RoundTripMismatchError(f"invalid XLSX amount type: {type(value).__name__}")
    if not amount.is_finite():
        raise RoundTripMismatchError("XLSX amount must be finite")
    return amount


def _canonical_cell(header: str, value: Any) -> Any:
    if header == "СуммаВВалютеУчета":
        return _canonical_amount(value)
    if value is None or value == "":
        return ""
    return str(value)


def _canonical_row(values: Sequence[Any]) -> tuple[Any, ...]:
    if len(values) != len(OUTPUT_HEADERS):
        raise RoundTripMismatchError(
            f"expected {len(OUTPUT_HEADERS)} cells, got {len(values)}"
        )
    return tuple(
        _canonical_cell(header, value) for header, value in zip(OUTPUT_HEADERS, values, strict=True)
    )


def _expected_canonical_rows(mapped_rows: Sequence[_MappedPosting]) -> list[tuple[Any, ...]]:
    return [_canonical_row(mapped.values) for mapped in mapped_rows]


def _prepare_rows(
    rows: PostingRows, adapter_config: OutputAdapterConfig
) -> dict[tuple[date, str], list[_MappedPosting]]:
    grouped: dict[tuple[date, str], list[_MappedPosting]] = defaultdict(list)
    for row in rows:
        if not isinstance(row, PostingRow):
            raise TypeError("rows must contain PostingRow values")
        try:
            mapped = map_posting_row(row, adapter_config)
        except BlockedOutputMapping as exc:
            raise ExportBlockedError(exc.reason, exc.message) from exc
        if row.period_end is None:
            # This is defensive because validate_posting_row already checks it.
            raise ExportBlockedError(
                BlockReason.BLOCKED_INVALID_POSTING,
                "PostingRow period_end is required for workbook grouping",
            )
        group = (row.period_end, row.document_organization)
        grouped[group].append(
            _MappedPosting(row, tuple(mapped[header] for header in OUTPUT_HEADERS), group)
        )

    for group, mapped_rows in grouped.items():
        mapped_rows.sort(key=_sort_key)
        canonical = _expected_canonical_rows(mapped_rows)
        counts = Counter(canonical)
        duplicates = [key for key, count in counts.items() if count > 1]
        if duplicates:
            raise DuplicatePostingRowError(
                "duplicate PostingRow output in group "
                f"{group[0].isoformat()} / {group[1]}"
            )
    return dict(grouped)


def _check_filename_collisions(
    groups: Mapping[tuple[date, str], Sequence[_MappedPosting]],
) -> dict[tuple[date, str], str]:
    names: dict[tuple[date, str], str] = {}
    component_owners: dict[str, tuple[date, str]] = {}
    filename_owners: dict[str, tuple[date, str]] = {}
    for group in sorted(groups):
        disputed = any(
            posting_has_blank_lower_analytics(mapped.source)
            for mapped in groups[group]
        )
        filename = deterministic_filename(*group, disputed=disputed)
        component_key = deterministic_filename(*group).casefold()
        previous = component_owners.get(component_key)
        if previous is not None and previous != group:
            raise FilenameCollisionError(
                f"filename collision: {filename!r} represents both "
                f"{previous!r} and {group!r}"
            )
        filename_key = filename.casefold()
        previous = filename_owners.get(filename_key)
        if previous is not None and previous != group:
            raise FilenameCollisionError(
                f"filename collision: {filename!r} represents both "
                f"{previous!r} and {group!r}"
            )
        component_owners[component_key] = group
        filename_owners[filename_key] = group
        names[group] = filename
    return names


def _write_workbook(
    path: Path,
    mapped_rows: Sequence[_MappedPosting],
    amount_number_format: str,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = OUTPUT_SHEET_NAME
    worksheet.append(list(OUTPUT_HEADERS))
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for mapped in mapped_rows:
        worksheet.append(list(mapped.values))

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:AA{max(1, len(mapped_rows) + 1)}"
    worksheet.sheet_view.showGridLines = False
    for cell in worksheet.iter_cols(
        min_col=_AMOUNT_COLUMN_INDEX + 1,
        max_col=_AMOUNT_COLUMN_INDEX + 1,
        min_row=2,
        max_row=max(1, len(mapped_rows) + 1),
    ):
        for amount_cell in cell:
            amount_cell.number_format = amount_number_format
    workbook.save(path)
    workbook.close()


def _round_trip_validate(
    path: Path,
    mapped_rows: Sequence[_MappedPosting],
) -> None:
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:  # pragma: no cover - openpyxl-specific corruption
        raise RoundTripMismatchError(f"cannot reopen workbook {path}: {exc}") from exc

    try:
        if workbook.sheetnames != [OUTPUT_SHEET_NAME]:
            raise RoundTripMismatchError(
                f"{path.name}: expected only sheet {OUTPUT_SHEET_NAME!r}, "
                f"got {workbook.sheetnames!r}"
            )
        worksheet = workbook[OUTPUT_SHEET_NAME]
        if worksheet.max_column != len(OUTPUT_HEADERS):
            raise RoundTripMismatchError(
                f"{path.name}: expected {len(OUTPUT_HEADERS)} columns, "
                f"got {worksheet.max_column}"
            )
        header = tuple(
            worksheet.cell(row=1, column=column).value
            for column in range(1, worksheet.max_column + 1)
        )
        if header != OUTPUT_HEADERS:
            raise RoundTripMismatchError(
                f"{path.name}: header mismatch; expected {OUTPUT_HEADERS!r}, got {header!r}"
            )

        expected = _expected_canonical_rows(mapped_rows)
        actual = [
            _canonical_row(values)
            for values in worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=len(OUTPUT_HEADERS),
                values_only=True,
            )
        ]
        if len(actual) != len(expected):
            raise RoundTripMismatchError(
                f"{path.name}: expected {len(expected)} financial rows, got {len(actual)}"
            )
        expected_counts = Counter(expected)
        actual_counts = Counter(actual)
        if actual_counts != expected_counts:
            unexpected = list((actual_counts - expected_counts).elements())
            missing = list((expected_counts - actual_counts).elements())
            raise RoundTripMismatchError(
                f"{path.name}: duplicate/unexpected output detected; "
                f"missing={len(missing)}, unexpected={len(unexpected)}"
            )
        if actual != expected:
            for index, (expected_row, actual_row) in enumerate(zip(expected, actual), start=2):
                if expected_row != actual_row:
                    raise RoundTripMismatchError(
                        f"{path.name}: row order/content mismatch at worksheet row {index}"
                    )
            raise RoundTripMismatchError(f"{path.name}: row count/order mismatch")
    finally:
        workbook.close()


def validate_workbook_round_trip(
    path: Path | str,
    rows: Iterable[PostingRow],
    config: OutputAdapterConfig | None = None,
) -> None:
    """Reopen and validate one workbook against its expected PostingRows.

    This public control is useful to callers that need to validate a workbook
    after an external handoff or to diagnose a failed export.
    """

    adapter_config = config or OutputAdapterConfig()
    grouped = _prepare_rows(rows, adapter_config)
    workbook_path = Path(path)
    if len(grouped) != 1:
        raise RoundTripMismatchError(
            "validate_workbook_round_trip expects rows from exactly one group"
        )
    _round_trip_validate(workbook_path, next(iter(grouped.values())))


def _resolve_export_config(
    output_dir: Path | str | None,
    config: OutputAdapterConfig | XlsxExportConfig | None,
    overwrite: bool | None,
) -> XlsxExportConfig:
    if isinstance(config, XlsxExportConfig):
        if output_dir is not None and Path(output_dir) != config.output_dir:
            raise ValueError("output_dir conflicts with XlsxExportConfig.output_dir")
        return XlsxExportConfig(
            output_dir=config.output_dir,
            adapter_config=config.adapter_config,
            overwrite=config.overwrite if overwrite is None else overwrite,
            amount_number_format=config.amount_number_format,
        )
    if config is not None and not isinstance(config, OutputAdapterConfig):
        raise TypeError("config must be OutputAdapterConfig or XlsxExportConfig")
    if output_dir is None:
        raise ValueError("output_dir is required")
    return XlsxExportConfig(
        output_dir=output_dir,
        adapter_config=config or OutputAdapterConfig(),
        overwrite=False if overwrite is None else overwrite,
    )


def export_posting_rows(
    rows: PostingRows,
    output_dir: Path | str | XlsxExportConfig | None = None,
    config: OutputAdapterConfig | XlsxExportConfig | None = None,
    *,
    overwrite: bool | None = None,
) -> XlsxExportResult:
    """Export PostingRows to one validated workbook per date and organization.

    All mapping, grouping, filename checks, writes, and round-trip checks are
    deterministic.  Final files are moved into place only after every staged
    workbook in the batch has passed validation.
    """

    if isinstance(output_dir, XlsxExportConfig):
        if config is not None:
            raise ValueError("config cannot be combined with XlsxExportConfig output_dir")
        config = output_dir
        output_dir = None
    export_config = _resolve_export_config(output_dir, config, overwrite)
    grouped = _prepare_rows(rows, export_config.adapter_config)
    filenames = _check_filename_collisions(grouped)
    output_root = export_config.output_dir
    if output_root.exists() and not output_root.is_dir():
        raise ExistingOutputError(f"output_dir is not a directory: {output_root}")
    existing = [
        output_root / filename
        for filename in filenames.values()
        if (output_root / filename).exists()
    ]
    if existing and not export_config.overwrite:
        raise ExistingOutputError(
            "refusing to replace existing output workbook(s): "
            + ", ".join(str(path) for path in existing)
        )

    output_root.mkdir(parents=True, exist_ok=True)
    if not grouped:
        return XlsxExportResult()

    staged: list[tuple[Path, Path, tuple[date, str], int]] = []
    with tempfile.TemporaryDirectory(prefix=".xlsx-export-", dir=output_root) as staging_dir:
        staging_root = Path(staging_dir)
        for group in sorted(grouped):
            mapped_rows = grouped[group]
            filename = filenames[group]
            staged_path = staging_root / filename
            _write_workbook(
                staged_path,
                mapped_rows,
                export_config.amount_number_format,
            )
            _round_trip_validate(staged_path, mapped_rows)
            staged.append((staged_path, output_root / filename, group, len(mapped_rows)))

        result: list[ExportedWorkbook] = []
        for staged_path, final_path, group, row_count in staged:
            staged_path.replace(final_path)
            result.append(
                ExportedWorkbook(
                    path=final_path,
                    document_organization=group[1],
                    document_date=group[0],
                    row_count=row_count,
                )
            )
    return XlsxExportResult(tuple(result))


def export_posting_rows_to_xlsx(
    rows: PostingRows,
    output_dir: Path | str | XlsxExportConfig | None = None,
    config: OutputAdapterConfig | XlsxExportConfig | None = None,
    *,
    overwrite: bool | None = None,
) -> XlsxExportResult:
    """Descriptive alias for :func:`export_posting_rows`."""

    return export_posting_rows(rows, output_dir, config, overwrite=overwrite)


class XlsxExporter:
    """Small stateful facade around :func:`export_posting_rows`."""

    def __init__(
        self,
        output_dir: Path | str,
        adapter_config: OutputAdapterConfig | None = None,
        *,
        overwrite: bool = False,
        amount_number_format: str = _DEFAULT_AMOUNT_NUMBER_FORMAT,
    ) -> None:
        self.config = XlsxExportConfig(
            output_dir=output_dir,
            adapter_config=adapter_config or OutputAdapterConfig(),
            overwrite=overwrite,
            amount_number_format=amount_number_format,
        )

    def export(self, rows: PostingRows) -> XlsxExportResult:
        return export_posting_rows(rows, config=self.config)


# Common verb aliases for integration callers.
export = export_posting_rows
write_xlsx = export_posting_rows


__all__ = [
    "DuplicatePostingRowError",
    "ExistingOutputError",
    "ExportBlockedError",
    "ExportConfig",
    "ExportedWorkbook",
    "ExporterConfig",
    "FilenameCollisionError",
    "RoundTripMismatchError",
    "XlsxExportConfig",
    "XlsxExportError",
    "XlsxExportResult",
    "XlsxExporter",
    "deterministic_filename",
    "export",
    "export_posting_rows",
    "export_posting_rows_to_xlsx",
    "posting_has_blank_lower_analytics",
    "sanitize_filename_component",
    "validate_workbook_round_trip",
    "write_xlsx",
]
