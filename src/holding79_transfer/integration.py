"""Deterministic local orchestration for one 79.x transfer run.

The integration layer owns run artifacts and controls only.  Accounting
semantics remain in :mod:`holding79_transfer.parser`, ``transfer`` and
``exporter``; this module composes those accepted components and fails closed
before publishing a run directory when a mandatory control does not pass.
"""

from __future__ import annotations

import hashlib
import json
import re
import tempfile
from collections import Counter, defaultdict
from collections.abc import Iterable, Mapping
from dataclasses import dataclass, field, replace
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import Any
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from .config import TransferConfig
from .exporter import (
    ExportedWorkbook,
    XlsxExportError,
    deterministic_filename,
    export_posting_rows,
    posting_has_blank_lower_analytics,
    validate_workbook_round_trip,
)
from .models import (
    CONTRACT_VERSION,
    RULES_VERSION,
    BalanceStatus,
    BlockReason,
    NormalizedBalance,
    PostingRow,
    financial_record_id,
)
from .output import OUTPUT_HEADERS, OUTPUT_SHEET_NAME, OutputAdapterConfig
from .parser import GroupedOsvParser, ParserDiagnostic, ParseResult
from .transfer import SourceEffectControl, TransferBatchResult, TransferEngine, TransferResult

RUN_ARTIFACT_NAMES: tuple[str, ...] = (
    "input_manifest.json",
    "normalized_balances.jsonl",
    "posting_rows.jsonl",
    "summary.json",
    "run_control.xlsx",
    "export_manifest.json",
)

CONTROL_SHEET_NAMES: tuple[str, ...] = (
    "Итоги",
    "Параметры_запуска",
    "Остатки_79",
    "Готовые_проводки",
    "Блокировки",
    "Контроль_до_после",
    "Исходные_строки",
    "Проверка_экспорта",
)

_SAFE_ID_RE = re.compile(r"[A-Za-z0-9][A-Za-z0-9_.-]{0,63}\Z")
_FINANCIAL_RECORD_ID_RE = re.compile(r"FR-[0-9a-f]{64}\Z")
_ZERO = Decimal(0)

_BALANCE_ARTIFACT_KEYS: tuple[str, ...] = (
    "period_end",
    "organization",
    "source_account",
    "department",
    "supplier_rvp",
    "ending_debit",
    "ending_credit",
    "source_excel_row_ref",
    "status",
    "block_reason",
)
_POSTING_ARTIFACT_KEYS: tuple[str, ...] = (
    "period_end",
    "document_organization",
    "debit_account",
    "debit_department",
    "debit_supplier_rvp",
    "credit_account",
    "credit_department",
    "credit_supplier_rvp",
    "amount",
    "source_organization",
    "source_account",
    "source_department",
    "source_supplier_rvp",
    "source_excel_row_ref",
    "financial_record_id",
    "side",
    "rules_version",
)
_DIAGNOSTIC_ARTIFACT_KEYS: tuple[str, ...] = (
    "sheet_name",
    "excel_row",
    "excel_column",
    "source_excel_row_ref",
    "code",
    "reason",
    "message",
    "status",
)


class IntegrationRunError(RuntimeError):
    """Raised when a local run cannot be published as a valid run."""


@dataclass(frozen=True)
class IntegrationRunConfig:
    """Explicit configuration for one deterministic local run."""

    period_end: date | datetime | str | None = None
    transfer_config: TransferConfig = field(default_factory=TransferConfig)
    adapter_config: OutputAdapterConfig = field(default_factory=OutputAdapterConfig)
    input_name: str = "osv.xlsx"
    run_id: str | None = None
    sheet_names: tuple[str, ...] | None = None

    def __post_init__(self) -> None:
        if not isinstance(self.transfer_config, TransferConfig):
            raise TypeError("transfer_config must be a TransferConfig")
        if not isinstance(self.adapter_config, OutputAdapterConfig):
            raise TypeError("adapter_config must be an OutputAdapterConfig")
        if not isinstance(self.input_name, str) or not self.input_name.strip():
            raise ValueError("input_name must be a non-empty string")
        input_name = self.input_name.strip()
        if input_name in {".", ".."} or "/" in input_name or "\\" in input_name:
            raise ValueError("input_name must be a file name, not a path")
        object.__setattr__(self, "input_name", input_name)

        if self.run_id is not None:
            if not isinstance(self.run_id, str) or not _SAFE_ID_RE.fullmatch(self.run_id.strip()):
                raise ValueError("run_id must be a safe non-empty identifier")
            object.__setattr__(self, "run_id", self.run_id.strip())

        if self.sheet_names is not None:
            names = tuple(name.strip() if isinstance(name, str) else name for name in self.sheet_names)
            if not names or any(not isinstance(name, str) or not name.strip() for name in names):
                raise ValueError("sheet_names must contain non-empty names")
            if len(set(names)) != len(names):
                raise ValueError("sheet_names must not contain duplicates")
            object.__setattr__(self, "sheet_names", names)


@dataclass(frozen=True)
class IntegrationRunResult:
    """Published run metadata and the canonical financial values."""

    output_dir: Path
    run_id: str
    status: str
    normalized_balances: tuple[NormalizedBalance, ...]
    posting_rows: tuple[PostingRow, ...]
    diagnostics: tuple[ParserDiagnostic, ...]
    exported_workbooks: tuple[ExportedWorkbook, ...]

    @property
    def is_success(self) -> bool:
        return self.status == "SUCCESS"

    @property
    def artifacts(self) -> tuple[Path, ...]:
        export_paths = tuple(workbook.path for workbook in self.exported_workbooks)
        return tuple(self.output_dir / name for name in RUN_ARTIFACT_NAMES) + export_paths


def _decimal_text(value: Decimal) -> str:
    if value == _ZERO:
        return "0"
    return format(value, "f")


def _canonical_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _write_json(path: Path, value: Any) -> None:
    path.write_text(
        json.dumps(value, ensure_ascii=False, sort_keys=True, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )


def _write_jsonl(path: Path, records: Iterable[Mapping[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="\n") as stream:
        for record in records:
            stream.write(_canonical_json(dict(record)))
            stream.write("\n")


def _balance_record(balance: NormalizedBalance) -> dict[str, Any]:
    return {
        "period_end": balance.period_end.isoformat() if balance.period_end else None,
        "organization": balance.organization,
        "source_account": balance.source_account.value if balance.source_account else None,
        "department": balance.department,
        "supplier_rvp": balance.supplier_rvp,
        "ending_debit": _decimal_text(balance.ending_debit),
        "ending_credit": _decimal_text(balance.ending_credit),
        "source_excel_row_ref": balance.source_excel_row_ref,
        "status": balance.status.value,
        "block_reason": balance.block_reason.value if balance.block_reason else None,
    }


def _posting_record(row: PostingRow) -> dict[str, Any]:
    return {
        "period_end": row.period_end.isoformat() if row.period_end else None,
        "document_organization": row.document_organization,
        "debit_account": row.debit_account.value,
        "debit_department": row.debit_department,
        "debit_supplier_rvp": row.debit_supplier_rvp,
        "credit_account": row.credit_account.value,
        "credit_department": row.credit_department,
        "credit_supplier_rvp": row.credit_supplier_rvp,
        "amount": _decimal_text(row.amount),
        "source_organization": row.source_organization,
        "source_account": row.source_account.value if row.source_account else None,
        "source_department": row.source_department,
        "source_supplier_rvp": row.source_supplier_rvp,
        "source_excel_row_ref": row.source_excel_row_ref,
        "financial_record_id": row.financial_record_id,
        "side": row.side.value if row.side else None,
        "rules_version": row.rules_version,
    }


def _balance_sort_key(balance: NormalizedBalance) -> tuple[str, ...]:
    return (
        balance.period_end.isoformat() if balance.period_end else "",
        balance.organization,
        balance.source_account.value if balance.source_account else "",
        balance.department,
        balance.supplier_rvp,
        balance.source_excel_row_ref or "",
        _decimal_text(balance.ending_debit),
        _decimal_text(balance.ending_credit),
    )


def _posting_sort_key(row: PostingRow) -> tuple[str, ...]:
    return (
        row.period_end.isoformat() if row.period_end else "",
        row.document_organization,
        row.financial_record_id,
        row.source_excel_row_ref or "",
        row.side.value if row.side else "",
        row.debit_account.value,
        row.credit_account.value,
        row.debit_department,
        row.credit_department,
        row.debit_supplier_rvp,
        row.credit_supplier_rvp,
        _decimal_text(row.amount),
    )


def _posting_identity(row: PostingRow) -> tuple[str, ...]:
    record = _posting_record(row)
    return tuple(str(record[key]) for key in record)


def _diagnostic_record(diagnostic: ParserDiagnostic) -> dict[str, Any]:
    return {
        "sheet_name": diagnostic.sheet_name,
        "excel_row": diagnostic.excel_row,
        "excel_column": diagnostic.excel_column,
        "source_excel_row_ref": (
            f"{diagnostic.sheet_name}!R{diagnostic.excel_row}"
            if diagnostic.sheet_name and diagnostic.excel_row is not None
            else None
        ),
        "code": diagnostic.code.value,
        "reason": diagnostic.reason.value,
        "message": diagnostic.message,
        "status": diagnostic.status.value,
    }


def _source_ref(diagnostic: ParserDiagnostic) -> str | None:
    if diagnostic.sheet_name and diagnostic.excel_row is not None:
        return f"{diagnostic.sheet_name}!R{diagnostic.excel_row}"
    return None


def _diagnostic_sort_key(diagnostic: ParserDiagnostic) -> tuple[str, int, int, str, str]:
    return (
        diagnostic.sheet_name or "",
        diagnostic.excel_row or 0,
        diagnostic.excel_column or 0,
        diagnostic.code.value,
        diagnostic.message,
    )


def _load_input_workbook(source: Any) -> tuple[Workbook, bool]:
    if isinstance(source, Workbook):
        return source, False
    if isinstance(source, Worksheet):
        return source.parent, False
    try:
        if isinstance(source, (str, Path)):
            return load_workbook(source, data_only=True), True
        if isinstance(source, (bytes, bytearray)):
            return load_workbook(BytesIO(bytes(source)), data_only=True), True
        if hasattr(source, "read"):
            stream = source
            if hasattr(stream, "seek"):
                stream.seek(0)
            return load_workbook(BytesIO(stream.read()), data_only=True), True
    except Exception as exc:  # pragma: no cover - openpyxl version-specific failures
        raise IntegrationRunError("input source is not a readable XLSX workbook") from exc
    raise TypeError(f"unsupported XLSX source type: {type(source).__name__}")


def _sheet_names(workbook: Workbook, configured: tuple[str, ...] | None) -> tuple[str, ...]:
    if configured is not None:
        missing = [name for name in configured if name not in workbook.sheetnames]
        if missing:
            raise IntegrationRunError(f"configured worksheet not found: {', '.join(missing)}")
        return configured
    if len(workbook.sheetnames) == 1:
        return (workbook.sheetnames[0],)
    preferred = tuple(
        name
        for name in workbook.sheetnames
        if "осв" in name.casefold() or "osv" in name.casefold()
    )
    return preferred or tuple(workbook.sheetnames)


def _parse_sheets(
    workbook: Workbook,
    config: IntegrationRunConfig,
) -> tuple[tuple[NormalizedBalance, ...], tuple[ParserDiagnostic, ...], list[dict[str, Any]]]:
    balances: list[NormalizedBalance] = []
    diagnostics: list[ParserDiagnostic] = []
    sheet_records: list[dict[str, Any]] = []
    for name in _sheet_names(workbook, config.sheet_names):
        result: ParseResult = GroupedOsvParser(
            config.period_end,
            sheet_name=name,
        ).parse(workbook)
        balances.extend(result.balances)
        diagnostics.extend(result.diagnostics)
        sheet_records.append(
            {
                "sheet_name": name,
                "status": result.status.value,
                "normalized_balance_count": len(result.balances),
                "diagnostic_count": len(result.diagnostics),
            }
        )
    ordered_balances = tuple(sorted(balances, key=_balance_sort_key))
    ordered_diagnostics = tuple(sorted(diagnostics, key=_diagnostic_sort_key))
    if any(diagnostic.excel_row is None for diagnostic in ordered_diagnostics):
        codes = ", ".join(diagnostic.code.value for diagnostic in ordered_diagnostics)
        raise IntegrationRunError(f"mandatory parser/source control failed: {codes}")
    if not ordered_balances and not ordered_diagnostics:
        raise IntegrationRunError("mandatory parser/source control failed: no normalized balances or diagnostics")
    return ordered_balances, ordered_diagnostics, sheet_records


def _config_identity(config: IntegrationRunConfig) -> dict[str, Any]:
    adapter = config.adapter_config
    return {
        "period_end": str(config.period_end) if config.period_end is not None else None,
        "transfer": {
            "manager_organization": config.transfer_config.manager_organization,
            "manager_financial_department": config.transfer_config.manager_financial_department,
            "rules_version": config.transfer_config.rules_version,
        },
        "adapter": {
            "operation_type": adapter.operation_type,
            "currency": adapter.currency,
            "debit_activity": adapter.debit_activity,
            "credit_activity": adapter.credit_activity,
            "contract_version": adapter.contract_version,
            "rules_version": adapter.rules_version,
        },
        "input_name": config.input_name,
        "sheet_names": list(config.sheet_names) if config.sheet_names else None,
    }


def _run_id(
    config: IntegrationRunConfig,
    balances: tuple[NormalizedBalance, ...],
    diagnostics: tuple[ParserDiagnostic, ...],
) -> tuple[str, str]:
    fingerprint_payload = {
        "contract_version": CONTRACT_VERSION,
        "config": _config_identity(config),
        "normalized_balances": [_balance_record(balance) for balance in balances],
        "diagnostics": [_diagnostic_record(diagnostic) for diagnostic in diagnostics],
    }
    fingerprint = hashlib.sha256(_canonical_json(fingerprint_payload).encode("utf-8")).hexdigest()
    configured_ids = {
        value
        for value in (config.run_id, config.adapter_config.run_id)
        if value
    }
    if len(configured_ids) > 1:
        raise ValueError("run_id and adapter_config.run_id must match when both are supplied")
    explicit = next(iter(configured_ids), None)
    if explicit is not None and not _SAFE_ID_RE.fullmatch(explicit):
        raise ValueError("adapter_config.run_id must be a safe non-empty identifier")
    return explicit or f"run-{fingerprint[:16]}", fingerprint


def _control_failure(message: str) -> IntegrationRunError:
    return IntegrationRunError(f"mandatory control failed: {message}")


def _canonical_transfer_batch(
    balances: tuple[NormalizedBalance, ...],
    transfer_config: TransferConfig,
) -> TransferBatchResult:
    """Build the approved batch independently of the caller-provided batch."""

    transfers: list[TransferResult] = []
    rows: list[PostingRow] = []
    seen_source_refs: set[str] = set()
    blocked_message: str | None = None

    for balance in balances:
        # A new engine and a new source-effect control are required for every
        # source balance.  The cached controls in the accepted batch are not
        # evidence for this independent canonical result.
        transfer = TransferEngine(transfer_config).generate(balance)
        if transfer.is_actionable:
            source_ref = balance.source_excel_row_ref
            if source_ref in seen_source_refs:
                transfer = TransferResult(
                    status=BalanceStatus.BLOCKED,
                    reason=BlockReason.BLOCKED_INVALID_POSTING,
                    message=f"source row consumed more than once: {source_ref}",
                )
            else:
                seen_source_refs.add(source_ref or "")
                rows.extend(transfer.rows)
        if transfer.status is BalanceStatus.BLOCKED and blocked_message is None:
            blocked_message = transfer.message
        transfers.append(transfer)

    if any(transfer.status is BalanceStatus.BLOCKED for transfer in transfers):
        return TransferBatchResult(
            status=BalanceStatus.BLOCKED,
            transfers=tuple(transfers),
            reason=BlockReason.BLOCKED_INVALID_POSTING,
            message=blocked_message or "one or more transfers are blocked",
        )
    if any(transfer.status is BalanceStatus.ACTIONABLE for transfer in transfers):
        return TransferBatchResult(
            status=BalanceStatus.ACTIONABLE,
            rows=tuple(rows),
            transfers=tuple(transfers),
        )
    return TransferBatchResult(
        status=BalanceStatus.NO_ACTION,
        transfers=tuple(transfers),
    )


def _posting_rows_match(
    actual: Iterable[Any],
    expected: tuple[PostingRow, ...],
) -> bool:
    try:
        actual_rows = tuple(actual)
    except TypeError:
        return False
    if not all(isinstance(row, PostingRow) for row in actual_rows):
        return False
    return tuple(_posting_identity(row) for row in actual_rows) == tuple(
        _posting_identity(row) for row in expected
    )


def _transfer_matches(actual: Any, expected: TransferResult) -> bool:
    if not isinstance(actual, TransferResult):
        return False
    if actual.source_effect is not None and not isinstance(
        actual.source_effect, SourceEffectControl
    ):
        return False
    try:
        rows_match = _posting_rows_match(actual.rows, expected.rows)
    except (AttributeError, TypeError):
        return False
    return (
        actual.status is expected.status
        and rows_match
        and actual.source_effect == expected.source_effect
        and actual.reason is expected.reason
        and actual.message == expected.message
    )


def _batch_matches_canonical(
    actual: Any,
    expected: TransferBatchResult,
) -> bool:
    if not isinstance(actual, TransferBatchResult):
        return False
    if actual.status is not expected.status:
        return False
    if not _posting_rows_match(actual.rows, expected.rows):
        return False
    try:
        actual_transfers = tuple(actual.transfers)
    except TypeError:
        return False
    return len(actual_transfers) == len(expected.transfers) and all(
        _transfer_matches(actual_transfer, expected_transfer)
        for actual_transfer, expected_transfer in zip(
            actual_transfers, expected.transfers, strict=True
        )
    ) and actual.reason is expected.reason and actual.message == expected.message


def _financial_controls(
    balances: tuple[NormalizedBalance, ...],
    batch: TransferBatchResult,
    diagnostics: tuple[ParserDiagnostic, ...],
    transfer_config: TransferConfig,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    canonical_batch = _canonical_transfer_batch(balances, transfer_config)
    if not _batch_matches_canonical(batch, canonical_batch):
        raise _control_failure("transfer batch does not match independent canonical output")
    if batch.status is BalanceStatus.BLOCKED:
        raise _control_failure(batch.message or "transfer batch is blocked")

    transfer_by_ref: dict[str, TransferResult] = {}
    expected_rows: list[PostingRow] = []
    control_rows: list[dict[str, Any]] = []
    source_total = _ZERO
    gk_total = _ZERO

    for balance, transfer in zip(balances, batch.transfers, strict=True):
        source_ref = balance.source_excel_row_ref or ""
        if transfer.status is BalanceStatus.NO_ACTION:
            control_rows.append(
                {
                    "source_excel_row_ref": source_ref,
                    "period_end": balance.period_end.isoformat() if balance.period_end else "",
                    "organization": balance.organization,
                    "source_account": balance.source_account.value if balance.source_account else "",
                    "department": balance.department,
                    "supplier_rvp": balance.supplier_rvp,
                    "ending_debit_before": _decimal_text(balance.ending_debit),
                    "ending_credit_before": _decimal_text(balance.ending_credit),
                    "ending_debit_after": _decimal_text(balance.ending_debit),
                    "ending_credit_after": _decimal_text(balance.ending_credit),
                    "source_posting_count": 0,
                    "source_amount": "0",
                    "gk_amount": "0",
                    "source_effect_zero": "NO_ACTION",
                    "amounts_match": "NO_ACTION",
                    "direction_correct": "NO_ACTION",
                    "source_consumed_once": "NO_ACTION",
                    "status": BalanceStatus.NO_ACTION.value,
                }
            )
            continue
        if transfer.status is not BalanceStatus.ACTIONABLE or len(transfer.rows) != 2:
            raise _control_failure(f"expected two postings for {source_ref}")
        if transfer.source_effect is None or not transfer.source_effect.passed:
            raise _control_failure(f"source-effect control failed for {source_ref}")

        transfer_by_ref[source_ref] = transfer
        source_row, gk_row = transfer.rows
        if (
            source_row.document_organization != balance.organization
            or gk_row.document_organization != transfer_config.manager_organization
        ):
            raise _control_failure(f"source/GK document organizations do not reconcile for {source_ref}")
        expected_rows.extend(transfer.rows)
        source_total += source_row.amount
        gk_total += gk_row.amount
        effect = transfer.source_effect
        control_rows.append(
            {
                "source_excel_row_ref": source_ref,
                "period_end": balance.period_end.isoformat() if balance.period_end else "",
                "organization": balance.organization,
                "source_account": balance.source_account.value if balance.source_account else "",
                "department": balance.department,
                "supplier_rvp": balance.supplier_rvp,
                "ending_debit_before": _decimal_text(effect.ending_debit_before),
                "ending_credit_before": _decimal_text(effect.ending_credit_before),
                "ending_debit_after": _decimal_text(effect.ending_debit_after),
                "ending_credit_after": _decimal_text(effect.ending_credit_after),
                "source_posting_count": effect.source_posting_count,
                "source_amount": _decimal_text(effect.source_amount or _ZERO),
                "gk_amount": _decimal_text(effect.gk_amount or _ZERO),
                "source_effect_zero": str(effect.source_effect_zero),
                "amounts_match": str(effect.amounts_match),
                "direction_correct": str(effect.direction_correct),
                "source_consumed_once": str(effect.source_consumed_once),
                "status": BalanceStatus.ACTIONABLE.value,
            }
        )

    actual_rows = tuple(batch.rows)
    financial_id_refs: dict[str, set[str]] = defaultdict(set)
    for row in actual_rows:
        financial_id_refs[row.financial_record_id].add(row.source_excel_row_ref or "")
    colliding_ids = sorted(
        financial_id
        for financial_id, source_refs in financial_id_refs.items()
        if len(source_refs) > 1
    )
    if colliding_ids:
        collisions = ", ".join(
            f"{financial_id}: {sorted(financial_id_refs[financial_id])}"
            for financial_id in colliding_ids
        )
        raise _control_failure(
            "financial_record_id maps to multiple source_excel_row_ref values: "
            + collisions
        )

    expected_counts = Counter(_posting_identity(row) for row in expected_rows)
    actual_counts = Counter(_posting_identity(row) for row in actual_rows)
    if actual_counts != expected_counts:
        raise _control_failure(
            f"posting reconciliation mismatch: missing={sum((expected_counts - actual_counts).values())}, "
            f"extra={sum((actual_counts - expected_counts).values())}"
        )

    actual_refs = {row.source_excel_row_ref for row in actual_rows}
    blocked_refs = {_source_ref(diagnostic) for diagnostic in diagnostics}
    blocked_refs.discard(None)
    if actual_refs.intersection(blocked_refs):
        raise _control_failure("blocked source row produced a financial PostingRow")

    for balance in balances:
        if balance.status is not BalanceStatus.ACTIONABLE:
            continue
        source_ref = balance.source_excel_row_ref or ""
        transfer = transfer_by_ref.get(source_ref)
        matching = [row for row in actual_rows if row.source_excel_row_ref == source_ref]
        if transfer is None or len(matching) != 2:
            raise _control_failure(f"source row does not reconcile 1:1 to two postings: {source_ref}")
        if sum((row.amount for row in matching), _ZERO) != balance.amount * 2:
            raise _control_failure(f"posting amount mismatch for {source_ref}")

    if source_total != gk_total:
        raise _control_failure(
            f"source-org/GK total mismatch: {source_total} != {gk_total}"
        )
    blocked_output_passed = all(
        row.source_excel_row_ref not in blocked_refs for row in actual_rows
    )
    if not blocked_output_passed:
        raise _control_failure("blocked parser/source row has financial output")

    checks = [
        {
            "control": "financial_record_id_one_source_ref",
            "status": "PASS",
            "value": len(financial_id_refs),
        },
        {"control": "source_row_reconciliation", "status": "PASS", "value": len(expected_rows)},
        {
            "control": "source_effect_zero",
            "status": "PASS",
            "value": sum(
                1
                for control in control_rows
                if control["status"] == BalanceStatus.ACTIONABLE.value
                and control["ending_debit_after"] == "0"
                and control["ending_credit_after"] == "0"
            ),
        },
        {
            "control": "source_org_gk_totals_match",
            "status": "PASS",
            "value": _decimal_text(source_total),
        },
        {
            "control": "blocked_rows_no_financial_output",
            "status": "PASS",
            "value": len(blocked_refs),
        },
    ]
    metrics = {
        "source_org_total": _decimal_text(source_total),
        "gk_total": _decimal_text(gk_total),
        "posting_row_count": len(actual_rows),
        "actionable_source_row_count": sum(
            balance.status is BalanceStatus.ACTIONABLE for balance in balances
        ),
        "no_action_source_row_count": sum(
            balance.status is BalanceStatus.NO_ACTION for balance in balances
        ),
        "blocked_source_row_count": len(blocked_refs),
    }
    return control_rows, checks, metrics


def _write_control_sheet(
    workbook: Workbook,
    name: str,
    headers: tuple[str, ...],
    rows: Iterable[Iterable[Any]],
) -> None:
    worksheet = workbook.create_sheet(name)
    worksheet.append(list(headers))
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        worksheet.append(["" if value is None else value for value in row])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{chr(64 + min(len(headers), 26))}{max(1, worksheet.max_row)}"
    worksheet.sheet_view.showGridLines = False


def _write_run_control(
    path: Path,
    *,
    run_id: str,
    config: IntegrationRunConfig,
    adapter_config: OutputAdapterConfig,
    balances: tuple[NormalizedBalance, ...],
    posting_rows: tuple[PostingRow, ...],
    diagnostics: tuple[ParserDiagnostic, ...],
    control_rows: list[dict[str, Any]],
    checks: list[dict[str, Any]],
    export_rows: list[dict[str, Any]],
    metrics: dict[str, Any],
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    summary_rows = [
        ("run_id", "PASS", run_id, "non-empty deterministic run id"),
        ("contract_version", "PASS", CONTRACT_VERSION, CONTRACT_VERSION),
        ("rules_version", "PASS", config.transfer_config.rules_version, RULES_VERSION),
        ("source_rows", "PASS", len(balances) + len({_source_ref(d) for d in diagnostics}), "source rows retained"),
        ("posting_rows", "PASS", metrics["posting_row_count"], "two per actionable source row"),
        ("source_org_total", "PASS", metrics["source_org_total"], "Decimal total"),
        ("gk_total", "PASS", metrics["gk_total"], "Decimal total"),
        ("blocked_source_rows", "PASS", metrics["blocked_source_row_count"], "retained in diagnostics"),
        ("export_round_trip", "PASS", len(export_rows), "all produced workbooks reopened"),
    ]
    summary_rows.extend(
        (
            check["control"],
            check["status"],
            check["value"],
            "PASS",
        )
        for check in checks
    )
    _write_control_sheet(
        workbook,
        "Итоги",
        ("Контроль", "Статус", "Значение", "Ожидание"),
        summary_rows,
    )

    parameters = (
        ("run_id", run_id),
        ("contract_version", CONTRACT_VERSION),
        ("period_end", str(config.period_end) if config.period_end is not None else "discovered"),
        ("manager_organization", config.transfer_config.manager_organization),
        ("manager_financial_department", config.transfer_config.manager_financial_department),
        ("rules_version", config.transfer_config.rules_version),
        ("operation_type", adapter_config.operation_type),
        ("currency", adapter_config.currency),
        ("debit_activity", adapter_config.debit_activity),
        ("credit_activity", adapter_config.credit_activity),
        ("input_name", config.input_name),
    )
    _write_control_sheet(workbook, "Параметры_запуска", ("Параметр", "Значение"), parameters)

    _write_control_sheet(
        workbook,
        "Остатки_79",
        (
            "source_excel_row_ref",
            "period_end",
            "organization",
            "source_account",
            "department",
            "supplier_rvp",
            "ending_debit",
            "ending_credit",
            "status",
            "block_reason",
        ),
        (
            (
                record["source_excel_row_ref"],
                record["period_end"],
                record["organization"],
                record["source_account"],
                record["department"],
                record["supplier_rvp"],
                record["ending_debit"],
                record["ending_credit"],
                record["status"],
                record["block_reason"],
            )
            for record in sorted((_balance_record(balance) for balance in balances), key=lambda item: item["source_excel_row_ref"] or "")
        ),
    )

    posting_headers = (
        "period_end",
        "document_organization",
        "debit_account",
        "debit_department",
        "debit_supplier_rvp",
        "credit_account",
        "credit_department",
        "credit_supplier_rvp",
        "amount",
        "source_organization",
        "source_account",
        "source_department",
        "source_supplier_rvp",
        "source_excel_row_ref",
        "financial_record_id",
        "side",
        "rules_version",
    )
    _write_control_sheet(
        workbook,
        "Готовые_проводки",
        posting_headers,
        (
            tuple(
                "" if record[header] is None else record[header]
                for header in posting_headers
            )
            for record in (_posting_record(row) for row in posting_rows)
        ),
    )

    diagnostic_rows = [
        (
            _source_ref(diagnostic) or "",
            diagnostic.sheet_name or "",
            diagnostic.excel_row or "",
            "parser",
            diagnostic.code.value,
            diagnostic.reason.value,
            diagnostic.message,
            0,
            0,
            "BLOCKED",
        )
        for diagnostic in diagnostics
    ]
    _write_control_sheet(
        workbook,
        "Блокировки",
        (
            "source_excel_row_ref",
            "sheet_name",
            "excel_row",
            "stage",
            "code",
            "reason",
            "message",
            "financial_posting_count",
            "financial_export_count",
            "status",
        ),
        diagnostic_rows,
    )

    effect_headers = (
        "source_excel_row_ref",
        "period_end",
        "organization",
        "source_account",
        "department",
        "supplier_rvp",
        "ending_debit_before",
        "ending_credit_before",
        "ending_debit_after",
        "ending_credit_after",
        "source_posting_count",
        "source_amount",
        "gk_amount",
        "source_effect_zero",
        "amounts_match",
        "direction_correct",
        "source_consumed_once",
        "status",
    )
    _write_control_sheet(
        workbook,
        "Контроль_до_после",
        effect_headers,
        (tuple(row.get(header, "") for header in effect_headers) for row in control_rows),
    )

    diagnostic_by_ref: dict[str | None, list[ParserDiagnostic]] = defaultdict(list)
    for diagnostic in diagnostics:
        diagnostic_by_ref[_source_ref(diagnostic)].append(diagnostic)
    source_rows = []
    for balance in balances:
        source_rows.append(
            (
                balance.source_excel_row_ref or "",
                "normalized",
                balance.status.value,
                balance.organization,
                balance.source_account.value if balance.source_account else "",
                balance.department,
                balance.supplier_rvp,
                "",
                "",
            )
        )
    for source_ref, source_diagnostics in sorted(diagnostic_by_ref.items(), key=lambda item: item[0] or ""):
        for diagnostic in source_diagnostics:
            source_rows.append(
                (
                    source_ref or "",
                    "parser",
                    BalanceStatus.BLOCKED.value,
                    "",
                    "",
                    "",
                    "",
                    diagnostic.code.value,
                    diagnostic.message,
                )
            )
    _write_control_sheet(
        workbook,
        "Исходные_строки",
        (
            "source_excel_row_ref",
            "stage",
            "status",
            "organization",
            "source_account",
            "department",
            "supplier_rvp",
            "block_code",
            "message",
        ),
        source_rows,
    )

    _write_control_sheet(
        workbook,
        "Проверка_экспорта",
        (
            "path",
            "document_organization",
            "document_date",
            "sheet_name",
            "column_count",
            "financial_row_count",
            "round_trip",
            "status",
        ),
        (
            (
                record["path"],
                record["document_organization"],
                record["document_date"],
                record["sheet_name"],
                record["column_count"],
                record["financial_row_count"],
                record["round_trip"],
                record["status"],
            )
            for record in export_rows
        ),
    )

    workbook.save(path)
    workbook.close()


def _export_manifest(
    stage_root: Path,
    export_result: Any,
    rows: tuple[PostingRow, ...],
    adapter_config: OutputAdapterConfig,
) -> tuple[list[dict[str, Any]], int]:
    export_records: list[dict[str, Any]] = []
    for workbook in export_result.workbooks:
        group_rows = tuple(
            row
            for row in rows
            if row.period_end == workbook.document_date
            and row.document_organization == workbook.document_organization
        )
        validate_workbook_round_trip(workbook.path, group_rows, adapter_config)
        relative_path = workbook.path.relative_to(stage_root).as_posix()
        export_records.append(
            {
                "path": relative_path,
                "document_organization": workbook.document_organization,
                "document_date": workbook.document_date.isoformat(),
                "sheet_name": OUTPUT_SHEET_NAME,
                "column_count": len(OUTPUT_HEADERS),
                "financial_row_count": workbook.row_count,
                "round_trip": True,
                "status": "PASS",
            }
        )
    export_records.sort(key=lambda record: record["path"])
    export_row_count = sum(record["financial_row_count"] for record in export_records)
    if export_row_count != len(rows):
        raise _control_failure("export manifest does not cover every financial PostingRow")
    return export_records, export_row_count


def _artifact_failure(name: str, message: str) -> None:
    raise IntegrationRunError(f"artifact validation failed: {name}: {message}")


def _read_json_artifact(path: Path) -> dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        _artifact_failure(path.name, f"invalid JSON: {exc}")
    if not isinstance(value, dict):
        _artifact_failure(path.name, "top-level value must be an object")
    return value


def _read_jsonl_artifact(path: Path) -> list[dict[str, Any]]:
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except OSError as exc:
        _artifact_failure(path.name, f"cannot read JSONL: {exc}")
    records: list[dict[str, Any]] = []
    for line_number, line in enumerate(lines, start=1):
        if not line.strip():
            _artifact_failure(path.name, f"blank JSONL line {line_number}")
        try:
            value = json.loads(line)
        except json.JSONDecodeError as exc:
            _artifact_failure(path.name, f"invalid JSON on line {line_number}: {exc}")
        if not isinstance(value, dict):
            _artifact_failure(path.name, f"line {line_number} must be an object")
        records.append(value)
    return records


def _require_artifact_keys(
    record: Mapping[str, Any], expected: tuple[str, ...], label: str
) -> None:
    if set(record) != set(expected):
        _artifact_failure(
            label,
            f"schema keys differ: expected {list(expected)!r}, got {sorted(record)!r}",
        )


def _artifact_text(
    record: Mapping[str, Any], key: str, label: str, *, allow_empty: bool = False
) -> str:
    value = record.get(key)
    if not isinstance(value, str) or (not allow_empty and not value):
        _artifact_failure(label, f"{key} must be a {'non-empty ' if not allow_empty else ''}string")
    return value


def _artifact_bool(record: Mapping[str, Any], key: str, label: str) -> bool:
    value = record.get(key)
    if not isinstance(value, bool):
        _artifact_failure(label, f"{key} must be a boolean")
    return value


def _artifact_int(record: Mapping[str, Any], key: str, label: str) -> int:
    value = record.get(key)
    if isinstance(value, bool) or not isinstance(value, int):
        _artifact_failure(label, f"{key} must be an integer")
    return value


def _artifact_date(value: Any, label: str, *, allow_none: bool = False) -> date | None:
    if value is None and allow_none:
        return None
    if not isinstance(value, str) or not value:
        _artifact_failure(label, "date must be an ISO string")
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        _artifact_failure(label, f"invalid ISO date: {value!r}: {exc}")


def _artifact_decimal(value: Any, label: str) -> Decimal:
    if not isinstance(value, str):
        _artifact_failure(label, "monetary value must be a string")
    try:
        result = Decimal(value)
    except InvalidOperation as exc:
        _artifact_failure(label, f"invalid Decimal value: {value!r}: {exc}")
    if not result.is_finite() or _decimal_text(result) != value:
        _artifact_failure(label, f"non-canonical or non-finite Decimal value: {value!r}")
    return result


def _control_rows(
    workbook: Any,
    name: str,
    headers: tuple[str, ...],
    label: str,
) -> list[tuple[Any, ...]]:
    if name not in workbook.sheetnames:
        _artifact_failure(label, f"missing control sheet {name!r}")
    worksheet = workbook[name]
    actual_headers = tuple(
        worksheet.cell(row=1, column=column).value
        for column in range(1, worksheet.max_column + 1)
    )
    if actual_headers != headers:
        _artifact_failure(
            label,
            f"{name} headers differ: expected {headers!r}, got {actual_headers!r}",
        )
    if worksheet.max_column != len(headers):
        _artifact_failure(label, f"{name} has wrong column count")
    if worksheet.max_row < 2:
        return []
    return [
        tuple(
            "" if value is None else value
            for value in row
        )
        for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=len(headers),
            values_only=True,
        )
    ]


def _posting_model_from_record(record: Mapping[str, Any], label: str) -> PostingRow:
    try:
        row = PostingRow.model_validate(record)
    except (TypeError, ValueError) as exc:  # pragma: no cover - pydantic message varies
        _artifact_failure(label, f"invalid PostingRow schema: {exc}")
    if row.status is not BalanceStatus.ACTIONABLE:
        _artifact_failure(label, "PostingRow is not actionable")
    if not _FINANCIAL_RECORD_ID_RE.fullmatch(row.financial_record_id):
        _artifact_failure(label, "financial_record_id is not a canonical deterministic id")
    return row


def _validate_artifacts(root: Path) -> None:
    """Validate the complete staged artifact graph before publication.

    The JSON/JSONL files are the machine-readable audit trail.  The control
    workbook and export workbooks are checked against that trail so a staged
    artifact cannot be edited into a superficially successful run.
    """

    for name in RUN_ARTIFACT_NAMES:
        path = root / name
        if not path.is_file():
            raise IntegrationRunError(f"mandatory artifact is missing: {name}")
    export_root = root / "export"
    if not export_root.is_dir():
        raise IntegrationRunError("mandatory artifact is missing: export/")

    input_manifest = _read_json_artifact(root / "input_manifest.json")
    _require_artifact_keys(
        input_manifest,
        (
            "artifact_version",
            "run_id",
            "contract_version",
            "input_name",
            "normalized_input_sha256",
            "period_end",
            "source_sheets",
            "parser_diagnostics",
            "config",
        ),
        "input_manifest.json",
    )
    artifact_version = _artifact_text(input_manifest, "artifact_version", "input_manifest.json")
    run_id = _artifact_text(input_manifest, "run_id", "input_manifest.json")
    contract_version = _artifact_text(input_manifest, "contract_version", "input_manifest.json")
    input_name = _artifact_text(input_manifest, "input_name", "input_manifest.json")
    fingerprint = _artifact_text(
        input_manifest, "normalized_input_sha256", "input_manifest.json"
    )
    if artifact_version != "H79_TRANSFER_RUN_V1":
        _artifact_failure("input_manifest.json", "unsupported artifact_version")
    if not _SAFE_ID_RE.fullmatch(run_id):
        _artifact_failure("input_manifest.json", "run_id is not safe and deterministic")
    if contract_version != CONTRACT_VERSION:
        _artifact_failure("input_manifest.json", "contract_version is not approved")
    if not re.fullmatch(r"[0-9a-f]{64}", fingerprint):
        _artifact_failure("input_manifest.json", "normalized_input_sha256 is not a SHA-256")

    source_sheets = input_manifest["source_sheets"]
    if not isinstance(source_sheets, list) or not source_sheets:
        _artifact_failure("input_manifest.json", "source_sheets must be a non-empty list")
    sheet_names: list[str] = []
    for index, sheet_record in enumerate(source_sheets):
        label = f"input_manifest.json source_sheets[{index}]"
        if not isinstance(sheet_record, dict):
            _artifact_failure(label, "sheet record must be an object")
        _require_artifact_keys(
            sheet_record,
            ("sheet_name", "status", "normalized_balance_count", "diagnostic_count"),
            label,
        )
        sheet_name = _artifact_text(sheet_record, "sheet_name", label)
        if sheet_name in sheet_names:
            _artifact_failure(label, f"duplicate sheet_name: {sheet_name}")
        sheet_names.append(sheet_name)
        status = _artifact_text(sheet_record, "status", label)
        if status not in {
            BalanceStatus.ACTIONABLE.value,
            BalanceStatus.NO_ACTION.value,
            BalanceStatus.BLOCKED.value,
        }:
            _artifact_failure(label, f"unsupported sheet status: {status}")
        if _artifact_int(sheet_record, "normalized_balance_count", label) < 0:
            _artifact_failure(label, "normalized_balance_count must not be negative")
        if _artifact_int(sheet_record, "diagnostic_count", label) < 0:
            _artifact_failure(label, "diagnostic_count must not be negative")

    config_record = input_manifest["config"]
    if not isinstance(config_record, dict):
        _artifact_failure("input_manifest.json", "config must be an object")
    _require_artifact_keys(
        config_record,
        ("period_end", "transfer", "adapter", "input_name", "sheet_names"),
        "input_manifest.json config",
    )
    if config_record["input_name"] != input_name:
        _artifact_failure("input_manifest.json", "config input_name does not match manifest")
    transfer_config = config_record["transfer"]
    adapter_config = config_record["adapter"]
    if not isinstance(transfer_config, dict) or not isinstance(adapter_config, dict):
        _artifact_failure("input_manifest.json config", "transfer and adapter must be objects")
    _require_artifact_keys(
        transfer_config,
        ("manager_organization", "manager_financial_department", "rules_version"),
        "input_manifest.json config.transfer",
    )
    _require_artifact_keys(
        adapter_config,
        (
            "operation_type",
            "currency",
            "debit_activity",
            "credit_activity",
            "contract_version",
            "rules_version",
        ),
        "input_manifest.json config.adapter",
    )
    for key in ("manager_organization", "manager_financial_department", "rules_version"):
        _artifact_text(transfer_config, key, "input_manifest.json config.transfer")
    for key in (
        "operation_type",
        "currency",
        "debit_activity",
        "credit_activity",
        "contract_version",
        "rules_version",
    ):
        _artifact_text(adapter_config, key, "input_manifest.json config.adapter", allow_empty=key in {
            "currency",
            "debit_activity",
            "credit_activity",
        })
    if transfer_config["rules_version"] != RULES_VERSION:
        _artifact_failure("input_manifest.json", "transfer rules_version is not approved")
    if (
        adapter_config["rules_version"] != transfer_config["rules_version"]
        or adapter_config["contract_version"] != contract_version
    ):
        _artifact_failure("input_manifest.json", "transfer/output versions do not reconcile")
    configured_sheet_names = config_record["sheet_names"]
    if configured_sheet_names is not None and (
        not isinstance(configured_sheet_names, list)
        or configured_sheet_names != sheet_names
    ):
        _artifact_failure("input_manifest.json", "configured sheet_names do not reconcile")

    normalized_records = _read_jsonl_artifact(root / "normalized_balances.jsonl")
    balances: list[NormalizedBalance] = []
    balances_by_ref: dict[str, NormalizedBalance] = {}
    for index, record in enumerate(normalized_records):
        label = f"normalized_balances.jsonl line {index + 1}"
        _require_artifact_keys(record, _BALANCE_ARTIFACT_KEYS, label)
        period_end = _artifact_date(record["period_end"], label)
        organization = _artifact_text(record, "organization", label)
        source_account = _artifact_text(record, "source_account", label)
        if source_account not in {"79.2", "79.3"}:
            _artifact_failure(label, f"unsupported source_account: {source_account}")
        department = _artifact_text(record, "department", label, allow_empty=True)
        supplier = _artifact_text(record, "supplier_rvp", label, allow_empty=True)
        source_ref = _artifact_text(record, "source_excel_row_ref", label)
        status = _artifact_text(record, "status", label)
        if status not in {BalanceStatus.ACTIONABLE.value, BalanceStatus.NO_ACTION.value}:
            _artifact_failure(label, f"unsupported balance status: {status}")
        if record["block_reason"] is not None:
            _artifact_failure(label, "successful normalized balance cannot have block_reason")
        ending_debit = _artifact_decimal(record["ending_debit"], label)
        ending_credit = _artifact_decimal(record["ending_credit"], label)
        if status == BalanceStatus.NO_ACTION.value and (ending_debit or ending_credit):
            _artifact_failure(label, "NO_ACTION balance has a non-zero ending side")
        if status == BalanceStatus.ACTIONABLE.value and (
            (ending_debit == _ZERO) == (ending_credit == _ZERO)
        ):
            _artifact_failure(label, "ACTIONABLE balance does not have exactly one ending side")
        if source_ref in balances_by_ref:
            _artifact_failure(label, f"duplicate source_excel_row_ref: {source_ref}")
        try:
            balance = NormalizedBalance(
                period_end=period_end,
                organization=organization,
                source_account=source_account,
                department=department,
                supplier_rvp=supplier,
                ending_debit=ending_debit,
                ending_credit=ending_credit,
                source_excel_row_ref=source_ref,
            )
        except (TypeError, ValueError) as exc:
            _artifact_failure(label, f"invalid normalized balance: {exc}")
        if balance.status.value != status:
            _artifact_failure(label, "serialized balance status does not match its values")
        balances.append(balance)
        balances_by_ref[source_ref] = balance

    diagnostics = input_manifest["parser_diagnostics"]
    if not isinstance(diagnostics, list):
        _artifact_failure("input_manifest.json", "parser_diagnostics must be a list")
    diagnostic_records: list[dict[str, Any]] = []
    for index, record in enumerate(diagnostics):
        label = f"input_manifest.json parser_diagnostics[{index}]"
        if not isinstance(record, dict):
            _artifact_failure(label, "diagnostic must be an object")
        _require_artifact_keys(record, _DIAGNOSTIC_ARTIFACT_KEYS, label)
        sheet_name = _artifact_text(record, "sheet_name", label)
        if sheet_name not in sheet_names:
            _artifact_failure(label, f"unknown diagnostic sheet: {sheet_name}")
        excel_row = _artifact_int(record, "excel_row", label)
        if excel_row <= 0:
            _artifact_failure(label, "excel_row must be positive")
        if record["excel_column"] is not None:
            excel_column = _artifact_int(record, "excel_column", label)
            if excel_column <= 0:
                _artifact_failure(label, "excel_column must be positive")
        source_ref = _artifact_text(record, "source_excel_row_ref", label)
        if source_ref != f"{sheet_name}!R{excel_row}":
            _artifact_failure(label, "source_excel_row_ref does not match sheet and row")
        _artifact_text(record, "code", label)
        _artifact_text(record, "reason", label)
        _artifact_text(record, "message", label)
        if _artifact_text(record, "status", label) != BalanceStatus.BLOCKED.value:
            _artifact_failure(label, "parser diagnostic status must be BLOCKED")
        diagnostic_records.append(record)

    source_sheet_counts = {record["sheet_name"]: record for record in source_sheets}
    for sheet_name, record in source_sheet_counts.items():
        sheet_balances = sum(
            1 for balance in balances if balance.source_excel_row_ref.startswith(f"{sheet_name}!R")
        )
        sheet_diagnostics = sum(
            1 for diagnostic in diagnostic_records if diagnostic["sheet_name"] == sheet_name
        )
        if record["normalized_balance_count"] != sheet_balances:
            _artifact_failure("input_manifest.json", f"normalized count mismatch for {sheet_name}")
        if record["diagnostic_count"] != sheet_diagnostics:
            _artifact_failure("input_manifest.json", f"diagnostic count mismatch for {sheet_name}")
        expected_status = (
            BalanceStatus.BLOCKED.value
            if sheet_diagnostics
            else (
                BalanceStatus.ACTIONABLE.value
                if any(
                    balance.status is BalanceStatus.ACTIONABLE
                    and balance.source_excel_row_ref.startswith(f"{sheet_name}!R")
                    for balance in balances
                )
                else BalanceStatus.NO_ACTION.value
            )
        )
        if record["status"] != expected_status:
            _artifact_failure("input_manifest.json", f"status mismatch for {sheet_name}")
    manifest_period = _artifact_date(
        input_manifest["period_end"], "input_manifest.json", allow_none=True
    )
    periods = {balance.period_end for balance in balances}
    if len(periods) > 1 or (periods and manifest_period != next(iter(periods))):
        _artifact_failure("input_manifest.json", "period_end does not reconcile with balances")
    expected_fingerprint_payload = {
        "contract_version": CONTRACT_VERSION,
        "config": config_record,
        "normalized_balances": normalized_records,
        "diagnostics": diagnostic_records,
    }
    expected_fingerprint = hashlib.sha256(
        _canonical_json(expected_fingerprint_payload).encode("utf-8")
    ).hexdigest()
    if fingerprint != expected_fingerprint:
        _artifact_failure("input_manifest.json", "normalized input fingerprint does not reconcile")

    posting_records = _read_jsonl_artifact(root / "posting_rows.jsonl")
    posting_rows: list[PostingRow] = []
    blocked_source_refs = {record["source_excel_row_ref"] for record in diagnostic_records}
    financial_id_refs: dict[str, set[str]] = defaultdict(set)
    rows_by_ref: dict[str, list[PostingRow]] = defaultdict(list)
    for index, record in enumerate(posting_records):
        label = f"posting_rows.jsonl line {index + 1}"
        _require_artifact_keys(record, _POSTING_ARTIFACT_KEYS, label)
        row = _posting_model_from_record(record, label)
        source_ref = row.source_excel_row_ref or ""
        balance = balances_by_ref.get(source_ref)
        if balance is None:
            _artifact_failure(label, f"unknown source_excel_row_ref: {source_ref}")
        if source_ref in blocked_source_refs:
            _artifact_failure(label, "financial row references a blocked source reference")
        if balance.status is not BalanceStatus.ACTIONABLE:
            _artifact_failure(label, "financial row references a non-actionable balance")
        if row.rules_version != transfer_config["rules_version"]:
            _artifact_failure(label, "financial row rules_version does not match configuration")
        if row.financial_record_id != financial_record_id(
            balance, transfer_config["rules_version"]
        ):
            _artifact_failure(label, "financial_record_id does not match normalized business identity")
        if row.amount != balance.amount or row.side is not balance.ending_side:
            _artifact_failure(label, "financial row amount or side does not match balance")
        if any(
            (
                row.period_end != balance.period_end,
                row.source_organization != balance.organization,
                (
                    row.source_account.value if row.source_account else None
                )
                != (
                    balance.source_account.value if balance.source_account else None
                ),
                row.source_department != balance.department,
                row.source_supplier_rvp != balance.supplier_rvp,
            )
        ):
            _artifact_failure(label, "financial row source trace does not match balance")
        financial_id_refs[row.financial_record_id].add(source_ref)
        rows_by_ref[source_ref].append(row)
        posting_rows.append(row)
    for financial_id, source_refs in financial_id_refs.items():
        if len(source_refs) != 1:
            _artifact_failure(
                "posting_rows.jsonl",
                f"financial_record_id maps to multiple source references: {financial_id}",
            )
    for source_ref, rows in rows_by_ref.items():
        if len(rows) != 2:
            _artifact_failure(
                "posting_rows.jsonl",
                f"source reference does not have exactly two PostingRows: {source_ref}",
            )
    posting_group_counts = Counter(
        (row.period_end, row.document_organization) for row in posting_rows
    )

    export_manifest = _read_json_artifact(root / "export_manifest.json")
    _require_artifact_keys(
        export_manifest,
        (
            "run_id",
            "contract_version",
            "sheet_name",
            "headers",
            "financial_row_count",
            "round_trip_validated",
            "workbooks",
        ),
        "export_manifest.json",
    )
    if export_manifest["run_id"] != run_id:
        _artifact_failure("export_manifest.json", "run_id does not match input_manifest")
    if export_manifest["contract_version"] != contract_version:
        _artifact_failure("export_manifest.json", "contract_version does not match input_manifest")
    if export_manifest["sheet_name"] != OUTPUT_SHEET_NAME:
        _artifact_failure("export_manifest.json", "unexpected output sheet name")
    if export_manifest["headers"] != list(OUTPUT_HEADERS):
        _artifact_failure("export_manifest.json", "output headers do not match the 27-column contract")
    if not _artifact_bool(export_manifest, "round_trip_validated", "export_manifest.json"):
        _artifact_failure("export_manifest.json", "round_trip_validated must be true")
    if _artifact_int(export_manifest, "financial_row_count", "export_manifest.json") != len(
        posting_rows
    ):
        _artifact_failure("export_manifest.json", "financial_row_count does not match PostingRows")
    workbooks = export_manifest["workbooks"]
    if not isinstance(workbooks, list):
        _artifact_failure("export_manifest.json", "workbooks must be a list")
    try:
        output_adapter = OutputAdapterConfig(
            operation_type=adapter_config["operation_type"],
            currency=adapter_config["currency"],
            debit_activity=adapter_config["debit_activity"],
            credit_activity=adapter_config["credit_activity"],
            run_id=run_id,
            contract_version=adapter_config["contract_version"],
            rules_version=adapter_config["rules_version"],
        )
    except (TypeError, ValueError) as exc:
        _artifact_failure("input_manifest.json config.adapter", f"invalid adapter config: {exc}")

    manifest_export_paths: list[str] = []
    manifest_group_counts: Counter[tuple[date, str]] = Counter()
    manifest_export_row_count = 0
    for index, record in enumerate(workbooks):
        label = f"export_manifest.json workbooks[{index}]"
        if not isinstance(record, dict):
            _artifact_failure(label, "workbook record must be an object")
        _require_artifact_keys(
            record,
            (
                "path",
                "document_organization",
                "document_date",
                "sheet_name",
                "column_count",
                "financial_row_count",
                "round_trip",
                "status",
            ),
            label,
        )
        relative_path = _artifact_text(record, "path", label)
        pure_path = PurePosixPath(relative_path)
        if (
            pure_path.parts[:1] != ("export",)
            or len(pure_path.parts) != 2
            or pure_path.suffix.casefold() != ".xlsx"
        ):
            _artifact_failure(label, f"unsafe export path: {relative_path}")
        document_organization = _artifact_text(record, "document_organization", label)
        document_date = _artifact_date(record["document_date"], label)
        group = (document_date, document_organization)
        if group in manifest_group_counts:
            _artifact_failure(
                label,
                f"duplicate workbook group: {document_date.isoformat()} / {document_organization}",
            )
        manifest_group_counts[group] += 1
        if record["sheet_name"] != OUTPUT_SHEET_NAME:
            _artifact_failure(label, "unexpected workbook sheet name")
        if _artifact_int(record, "column_count", label) != len(OUTPUT_HEADERS):
            _artifact_failure(label, "workbook column count is not 27")
        row_count = _artifact_int(record, "financial_row_count", label)
        if row_count <= 0:
            _artifact_failure(label, "published workbook must contain financial rows")
        if not _artifact_bool(record, "round_trip", label):
            _artifact_failure(label, "workbook round_trip must be true")
        if _artifact_text(record, "status", label) != "PASS":
            _artifact_failure(label, "workbook status must be PASS")
        group_rows = tuple(
            row
            for row in posting_rows
            if row.period_end == document_date
            and row.document_organization == document_organization
        )
        disputed = any(posting_has_blank_lower_analytics(row) for row in group_rows)
        expected_path = (
            f"export/{deterministic_filename(document_date, document_organization, disputed=disputed)}"
        )
        if relative_path != expected_path:
            _artifact_failure(label, "workbook path is not deterministic for its group")
        if relative_path in manifest_export_paths:
            _artifact_failure(label, f"duplicate workbook path: {relative_path}")
        manifest_export_paths.append(relative_path)
        if row_count != len(group_rows):
            _artifact_failure(label, "workbook row count does not match PostingRows group")
        manifest_export_row_count += row_count
        workbook_path = root / Path(*pure_path.parts)
        if not workbook_path.is_file():
            _artifact_failure(label, f"workbook file is missing: {relative_path}")
        try:
            validate_workbook_round_trip(workbook_path, group_rows, output_adapter)
        except (OSError, TypeError, ValueError, KeyError, XlsxExportError) as exc:
            # Exporter and workbook errors are all fail-closed validation failures.
            _artifact_failure(label, f"workbook round-trip validation failed: {exc}")

    if set(manifest_group_counts) != set(posting_group_counts):
        missing_groups = sorted(set(posting_group_counts) - set(manifest_group_counts))
        extra_groups = sorted(set(manifest_group_counts) - set(posting_group_counts))
        _artifact_failure(
            "export_manifest.json",
            f"workbook groups do not match PostingRows: missing={missing_groups!r}, "
            f"extra={extra_groups!r}",
        )
    if manifest_export_row_count != len(posting_rows):
        _artifact_failure(
            "export_manifest.json",
            "sum of workbook financial_row_count values does not match PostingRows",
        )

    actual_export_entries = list(export_root.iterdir())
    if any(entry.is_dir() for entry in actual_export_entries):
        _artifact_failure("export/", "nested directories are not allowed")
    actual_export_paths = sorted(
        entry.relative_to(root).as_posix()
        for entry in actual_export_entries
        if entry.is_file()
    )
    if actual_export_paths != sorted(manifest_export_paths):
        _artifact_failure(
            "export_manifest.json",
            f"export files do not match manifest: files={actual_export_paths!r}, "
            f"manifest={sorted(manifest_export_paths)!r}",
        )
    if not posting_rows and (workbooks or actual_export_paths):
        _artifact_failure("export_manifest.json", "zero financial rows must have an empty export directory")
    if posting_rows and not workbooks:
        _artifact_failure("export_manifest.json", "financial rows require published workbooks")

    summary = _read_json_artifact(root / "summary.json")
    _require_artifact_keys(
        summary,
        (
            "status",
            "run_id",
            "contract_version",
            "rules_version",
            "counts",
            "totals",
            "controls",
            "artifacts",
        ),
        "summary.json",
    )
    if summary["status"] != "SUCCESS":
        _artifact_failure("summary.json", "status must be SUCCESS before publication")
    if summary["run_id"] != run_id or summary["contract_version"] != contract_version:
        _artifact_failure("summary.json", "run identity does not match input_manifest")
    if summary["rules_version"] != transfer_config["rules_version"]:
        _artifact_failure("summary.json", "rules_version does not match configuration")
    blocked_refs = blocked_source_refs
    actionable_count = sum(balance.status is BalanceStatus.ACTIONABLE for balance in balances)
    no_action_count = sum(balance.status is BalanceStatus.NO_ACTION for balance in balances)
    expected_total = sum(
        (balance.amount for balance in balances if balance.status is BalanceStatus.ACTIONABLE),
        _ZERO,
    )
    expected_counts = {
        "actionable_source_rows": actionable_count,
        "blocked_source_rows": len(blocked_refs),
        "export_rows": len(posting_rows),
        "export_workbooks": len(workbooks),
        "no_action_source_rows": no_action_count,
        "normalized_balances": len(balances),
        "parser_diagnostics": len(diagnostic_records),
        "posting_rows": len(posting_rows),
        "source_rows": len(balances) + len(blocked_refs),
    }
    counts = summary["counts"]
    if counts != expected_counts:
        _artifact_failure("summary.json", f"counts do not reconcile: expected {expected_counts!r}")
    totals = summary["totals"]
    if not isinstance(totals, dict) or set(totals) != {"source_org", "gk", "difference"}:
        _artifact_failure("summary.json", "totals schema is invalid")
    expected_total_text = _decimal_text(expected_total)
    if totals != {
        "source_org": expected_total_text,
        "gk": expected_total_text,
        "difference": "0",
    }:
        _artifact_failure("summary.json", "totals do not reconcile with normalized balances")
    controls = summary["controls"]
    if not isinstance(controls, dict) or set(controls) != {"all_passed", "records"}:
        _artifact_failure("summary.json", "controls schema is invalid")
    if controls["all_passed"] is not True or not isinstance(controls["records"], list):
        _artifact_failure("summary.json", "controls do not report an explicit all-pass result")
    expected_checks = [
        {
            "control": "financial_record_id_one_source_ref",
            "status": "PASS",
            "value": len(financial_id_refs),
        },
        {"control": "source_row_reconciliation", "status": "PASS", "value": len(posting_rows)},
        {
            "control": "source_effect_zero",
            "status": "PASS",
            "value": actionable_count,
        },
        {
            "control": "source_org_gk_totals_match",
            "status": "PASS",
            "value": expected_total_text,
        },
        {
            "control": "blocked_rows_no_financial_output",
            "status": "PASS",
            "value": len(blocked_refs),
        },
        {"control": "export_round_trip", "status": "PASS", "value": len(posting_rows)},
    ]
    if controls["records"] != expected_checks:
        _artifact_failure("summary.json", "control results do not reconcile")
    expected_artifacts = list(RUN_ARTIFACT_NAMES) + sorted(manifest_export_paths)
    if summary["artifacts"] != expected_artifacts:
        _artifact_failure("summary.json", "artifact list does not match staged files")

    control_path = root / "run_control.xlsx"
    try:
        control_workbook = load_workbook(control_path, read_only=True, data_only=False)
    except (OSError, ValueError, KeyError, BadZipFile) as exc:  # openpyxl-specific corruption
        _artifact_failure("run_control.xlsx", f"cannot reopen control workbook: {exc}")
    try:
        if control_workbook.sheetnames != list(CONTROL_SHEET_NAMES):
            _artifact_failure("run_control.xlsx", "control sheet names do not match the contract")

        summary_headers = ("Контроль", "Статус", "Значение", "Ожидание")
        parameter_headers = ("Параметр", "Значение")
        balance_headers = (
            "source_excel_row_ref",
            "period_end",
            "organization",
            "source_account",
            "department",
            "supplier_rvp",
            "ending_debit",
            "ending_credit",
            "status",
            "block_reason",
        )
        posting_headers = _POSTING_ARTIFACT_KEYS
        blocked_headers = (
            "source_excel_row_ref",
            "sheet_name",
            "excel_row",
            "stage",
            "code",
            "reason",
            "message",
            "financial_posting_count",
            "financial_export_count",
            "status",
        )
        effect_headers = (
            "source_excel_row_ref",
            "period_end",
            "organization",
            "source_account",
            "department",
            "supplier_rvp",
            "ending_debit_before",
            "ending_credit_before",
            "ending_debit_after",
            "ending_credit_after",
            "source_posting_count",
            "source_amount",
            "gk_amount",
            "source_effect_zero",
            "amounts_match",
            "direction_correct",
            "source_consumed_once",
            "status",
        )
        source_headers = (
            "source_excel_row_ref",
            "stage",
            "status",
            "organization",
            "source_account",
            "department",
            "supplier_rvp",
            "block_code",
            "message",
        )
        export_headers = (
            "path",
            "document_organization",
            "document_date",
            "sheet_name",
            "column_count",
            "financial_row_count",
            "round_trip",
            "status",
        )

        actual_summary_rows = _control_rows(
            control_workbook, "Итоги", summary_headers, "run_control.xlsx"
        )
        expected_summary_rows = [
            ("run_id", "PASS", run_id, "non-empty deterministic run id"),
            ("contract_version", "PASS", contract_version, CONTRACT_VERSION),
            ("rules_version", "PASS", transfer_config["rules_version"], RULES_VERSION),
            ("source_rows", "PASS", expected_counts["source_rows"], "source rows retained"),
            ("posting_rows", "PASS", len(posting_rows), "two per actionable source row"),
            ("source_org_total", "PASS", expected_total_text, "Decimal total"),
            ("gk_total", "PASS", expected_total_text, "Decimal total"),
            ("blocked_source_rows", "PASS", len(blocked_refs), "retained in diagnostics"),
            ("export_round_trip", "PASS", len(workbooks), "all produced workbooks reopened"),
        ] + [
            (check["control"], check["status"], check["value"], "PASS")
            for check in expected_checks
        ]
        if actual_summary_rows != expected_summary_rows:
            _artifact_failure("run_control.xlsx", "Итоги does not match summary/control results")

        actual_parameters = _control_rows(
            control_workbook, "Параметры_запуска", parameter_headers, "run_control.xlsx"
        )
        expected_parameters = [
            ("run_id", run_id),
            ("contract_version", contract_version),
            (
                "period_end",
                config_record["period_end"] if config_record["period_end"] is not None else "discovered",
            ),
            ("manager_organization", transfer_config["manager_organization"]),
            ("manager_financial_department", transfer_config["manager_financial_department"]),
            ("rules_version", transfer_config["rules_version"]),
            ("operation_type", adapter_config["operation_type"]),
            ("currency", adapter_config["currency"]),
            ("debit_activity", adapter_config["debit_activity"]),
            ("credit_activity", adapter_config["credit_activity"]),
            ("input_name", input_name),
        ]
        if actual_parameters != expected_parameters:
            _artifact_failure("run_control.xlsx", "Параметры_запуска does not match input_manifest")

        actual_balances = _control_rows(
            control_workbook, "Остатки_79", balance_headers, "run_control.xlsx"
        )
        expected_balances = [
            tuple(
                record[key] if record[key] is not None else ""
                for key in balance_headers
            )
            for record in sorted(normalized_records, key=lambda item: item["source_excel_row_ref"])
        ]
        if actual_balances != expected_balances:
            _artifact_failure("run_control.xlsx", "Остатки_79 does not match normalized_balances.jsonl")

        actual_postings = _control_rows(
            control_workbook, "Готовые_проводки", posting_headers, "run_control.xlsx"
        )
        expected_postings = [
            tuple(record[key] if record[key] is not None else "" for key in posting_headers)
            for record in posting_records
        ]
        if actual_postings != expected_postings:
            _artifact_failure("run_control.xlsx", "Готовые_проводки does not match posting_rows.jsonl")

        actual_blocked = _control_rows(
            control_workbook, "Блокировки", blocked_headers, "run_control.xlsx"
        )
        expected_blocked = [
            (
                record["source_excel_row_ref"],
                record["sheet_name"],
                record["excel_row"],
                "parser",
                record["code"],
                record["reason"],
                record["message"],
                0,
                0,
                "BLOCKED",
            )
            for record in diagnostic_records
        ]
        if actual_blocked != expected_blocked:
            _artifact_failure("run_control.xlsx", "Блокировки does not match parser diagnostics")

        actual_effects = _control_rows(
            control_workbook, "Контроль_до_после", effect_headers, "run_control.xlsx"
        )
        expected_effects: list[tuple[Any, ...]] = []
        for balance in balances:
            is_actionable = balance.status is BalanceStatus.ACTIONABLE
            amount = _decimal_text(balance.amount)
            expected_effects.append(
                (
                    balance.source_excel_row_ref,
                    balance.period_end.isoformat() if balance.period_end else "",
                    balance.organization,
                    balance.source_account.value if balance.source_account else "",
                    balance.department,
                    balance.supplier_rvp,
                    _decimal_text(balance.ending_debit),
                    _decimal_text(balance.ending_credit),
                    "0" if is_actionable else _decimal_text(balance.ending_debit),
                    "0" if is_actionable else _decimal_text(balance.ending_credit),
                    1 if is_actionable else 0,
                    amount if is_actionable else "0",
                    amount if is_actionable else "0",
                    "True" if is_actionable else "NO_ACTION",
                    "True" if is_actionable else "NO_ACTION",
                    "True" if is_actionable else "NO_ACTION",
                    "True" if is_actionable else "NO_ACTION",
                    BalanceStatus.ACTIONABLE.value if is_actionable else BalanceStatus.NO_ACTION.value,
                )
            )
        if actual_effects != expected_effects:
            _artifact_failure("run_control.xlsx", "Контроль_до_после does not match financial controls")

        actual_source_rows = _control_rows(
            control_workbook, "Исходные_строки", source_headers, "run_control.xlsx"
        )
        expected_source_rows = [
            (
                record["source_excel_row_ref"],
                "normalized",
                record["status"],
                record["organization"],
                record["source_account"],
                record["department"],
                record["supplier_rvp"],
                "",
                "",
            )
            for record in normalized_records
        ] + [
            (
                record["source_excel_row_ref"],
                "parser",
                BalanceStatus.BLOCKED.value,
                "",
                "",
                "",
                "",
                record["code"],
                record["message"],
            )
            for record in sorted(
                diagnostic_records, key=lambda item: item["source_excel_row_ref"]
            )
        ]
        if actual_source_rows != expected_source_rows:
            _artifact_failure("run_control.xlsx", "Исходные_строки does not match source artifacts")

        actual_export_rows = _control_rows(
            control_workbook, "Проверка_экспорта", export_headers, "run_control.xlsx"
        )
        expected_export_rows = [
            (
                record["path"],
                record["document_organization"],
                record["document_date"],
                record["sheet_name"],
                record["column_count"],
                record["financial_row_count"],
                record["round_trip"],
                record["status"],
            )
            for record in workbooks
        ]
        if actual_export_rows != expected_export_rows:
            _artifact_failure("run_control.xlsx", "Проверка_экспорта does not match export_manifest")
    finally:
        control_workbook.close()


def run_integration(
    source: Any,
    output_dir: Path | str,
    *,
    config: IntegrationRunConfig | None = None,
    period_end: date | datetime | str | None = None,
    transfer_config: TransferConfig | None = None,
    adapter_config: OutputAdapterConfig | None = None,
    input_name: str | None = None,
) -> IntegrationRunResult:
    """Run parser -> transfer -> controls -> export as one local transaction.

    The destination must not already exist.  All files are built in a sibling
    staging directory and the complete directory is published only after the
    mandatory controls and workbook round trips pass.
    """

    if config is not None and any(
        value is not None for value in (period_end, transfer_config, adapter_config, input_name)
    ):
        raise ValueError("config cannot be combined with individual run options")
    run_config = config or IntegrationRunConfig(
        period_end=period_end,
        transfer_config=transfer_config or TransferConfig(),
        adapter_config=adapter_config or OutputAdapterConfig(),
        input_name=input_name or "osv.xlsx",
    )
    run_root = Path(output_dir)
    if run_root.exists():
        raise IntegrationRunError(f"run output already exists: {run_root}")
    if run_root.name in {"", ".", ".."}:
        raise ValueError("output_dir must name a dedicated run directory")
    parent = run_root.parent
    parent.mkdir(parents=True, exist_ok=True)

    workbook, close_workbook = _load_input_workbook(source)
    try:
        balances, diagnostics, sheet_records = _parse_sheets(workbook, run_config)
        run_id, fingerprint = _run_id(run_config, balances, diagnostics)
        if (
            run_config.transfer_config.rules_version != RULES_VERSION
            or run_config.adapter_config.rules_version != run_config.transfer_config.rules_version
            or run_config.adapter_config.contract_version != CONTRACT_VERSION
        ):
            raise _control_failure("transfer/output versions are not approved and identical")
        effective_adapter = replace(run_config.adapter_config, run_id=run_id)
        engine = TransferEngine(run_config.transfer_config)
        batch = engine.generate_batch(balances)
        control_rows, checks, metrics = _financial_controls(
            balances,
            batch,
            diagnostics,
            run_config.transfer_config,
        )
        if any(check["status"] != "PASS" for check in checks):
            raise _control_failure("one or more financial controls did not pass")

        with tempfile.TemporaryDirectory(prefix=".holding79-run-", dir=parent) as staging_name:
            stage_root = Path(staging_name)
            (stage_root / "export").mkdir()
            _write_json(
                stage_root / "input_manifest.json",
                {
                    "artifact_version": "H79_TRANSFER_RUN_V1",
                    "run_id": run_id,
                    "contract_version": CONTRACT_VERSION,
                    "input_name": run_config.input_name,
                    "normalized_input_sha256": fingerprint,
                    "period_end": (
                        balances[0].period_end.isoformat() if balances and balances[0].period_end else None
                    ),
                    "source_sheets": sheet_records,
                    "parser_diagnostics": [_diagnostic_record(diagnostic) for diagnostic in diagnostics],
                    "config": _config_identity(run_config),
                },
            )
            _write_jsonl(
                stage_root / "normalized_balances.jsonl",
                (_balance_record(balance) for balance in balances),
            )
            ordered_rows = tuple(sorted(batch.rows, key=_posting_sort_key))
            _write_jsonl(
                stage_root / "posting_rows.jsonl",
                (_posting_record(row) for row in ordered_rows),
            )

            export_result = export_posting_rows(
                ordered_rows,
                stage_root / "export",
                effective_adapter,
            )
            export_records, export_row_count = _export_manifest(
                stage_root,
                export_result,
                ordered_rows,
                effective_adapter,
            )
            if export_row_count != metrics["posting_row_count"]:
                raise _control_failure("export row count does not match financial PostingRows")
            checks.append({"control": "export_round_trip", "status": "PASS", "value": export_row_count})

            _write_run_control(
                stage_root / "run_control.xlsx",
                run_id=run_id,
                config=run_config,
                adapter_config=effective_adapter,
                balances=balances,
                posting_rows=ordered_rows,
                diagnostics=diagnostics,
                control_rows=control_rows,
                checks=checks,
                export_rows=export_records,
                metrics=metrics,
            )

            artifact_paths = list(RUN_ARTIFACT_NAMES)
            artifact_paths.extend(record["path"] for record in export_records)
            summary = {
                "status": "SUCCESS",
                "run_id": run_id,
                "contract_version": CONTRACT_VERSION,
                "rules_version": run_config.transfer_config.rules_version,
                "counts": {
                    "source_rows": len(balances) + len({_source_ref(d) for d in diagnostics}),
                    "normalized_balances": len(balances),
                    "actionable_source_rows": metrics["actionable_source_row_count"],
                    "no_action_source_rows": metrics["no_action_source_row_count"],
                    "blocked_source_rows": metrics["blocked_source_row_count"],
                    "posting_rows": metrics["posting_row_count"],
                    "export_rows": export_row_count,
                    "export_workbooks": len(export_records),
                    "parser_diagnostics": len(diagnostics),
                },
                "totals": {
                    "source_org": metrics["source_org_total"],
                    "gk": metrics["gk_total"],
                    "difference": _decimal_text(
                        Decimal(metrics["source_org_total"]) - Decimal(metrics["gk_total"])
                    ),
                },
                "controls": {
                    "all_passed": True,
                    "records": checks,
                },
                "artifacts": artifact_paths,
            }
            _write_json(stage_root / "export_manifest.json", {
                "run_id": run_id,
                "contract_version": CONTRACT_VERSION,
                "sheet_name": OUTPUT_SHEET_NAME,
                "headers": list(OUTPUT_HEADERS),
                "financial_row_count": export_row_count,
                "round_trip_validated": True,
                "workbooks": export_records,
            })
            _write_json(stage_root / "summary.json", summary)
            _validate_artifacts(stage_root)
            stage_root.replace(run_root)

        published_workbooks = tuple(
            ExportedWorkbook(
                path=run_root / record["path"],
                document_organization=record["document_organization"],
                document_date=date.fromisoformat(record["document_date"]),
                row_count=record["financial_row_count"],
            )
            for record in export_records
        )
        return IntegrationRunResult(
            output_dir=run_root,
            run_id=run_id,
            status="SUCCESS",
            normalized_balances=balances,
            posting_rows=ordered_rows,
            diagnostics=diagnostics,
            exported_workbooks=published_workbooks,
        )
    finally:
        if close_workbook:
            workbook.close()


def build_synthetic_osv_workbook() -> Workbook:
    """Create the in-memory OSV covering the goldens and a missing-organization block."""

    cases = (
        ("DEBIT_79_2_AT", "79.2", "84272.40", "0"),
        ("CREDIT_79_2_AT", "79.2", "0", "84272.40"),
        ("DEBIT_79_3_AT", "79.3", "84272.40", "0"),
        ("CREDIT_79_3_AT", "79.3", "0", "84272.40"),
    )
    workbook = Workbook()
    for index, (name, account, debit, credit) in enumerate(cases):
        worksheet = workbook.active if index == 0 else workbook.create_sheet()
        worksheet.title = name
        _populate_synthetic_sheet(
            worksheet,
            account=account,
            organization="АТ",
            department="Б_АТ Коммерческий отдел",
            supplier="Производитель",
            debit=debit,
            credit=credit,
        )
    blocked = workbook.create_sheet("BLOCKED_SOURCE_ROW")
    _populate_synthetic_sheet(
        blocked,
        account="79.2",
        organization=None,
        department=None,
        supplier=None,
        debit="100.00",
        credit="0",
    )
    return workbook


def _populate_synthetic_sheet(
    worksheet: Worksheet,
    *,
    account: str,
    organization: str | None,
    department: str | None,
    supplier: str | None,
    debit: str,
    credit: str,
) -> None:
    worksheet.cell(1, 1).value = "Синтетическая ОСВ"
    worksheet.cell(2, 1).value = "Период: 01.12.2024 - 31.12.2024"
    worksheet.merge_cells("H4:I4")
    worksheet["H4"] = "Сальдо на конец периода"
    worksheet["H5"] = "Дебет"
    worksheet["I5"] = "Кредит"
    rows = (
        (6, account, 0),
        (7, "" if organization is None else f"Организация: {organization}", 1),
        (8, "" if department is None else f"ЦФО: {department}", 2),
        (9, "Поставщик РВП" if supplier is None else f"Поставщик РВП: {supplier}", 3),
    )
    for row, value, indent in rows:
        worksheet.cell(row, 1).value = value
        worksheet.cell(row, 1).alignment = Alignment(indent=indent)
    worksheet.cell(9, 8).value = debit
    worksheet.cell(9, 9).value = credit


def run_synthetic_integration(output_dir: Path | str) -> IntegrationRunResult:
    """Execute the permanent local synthetic Issue #5 scenario."""

    workbook = build_synthetic_osv_workbook()
    try:
        return run_integration(
            workbook,
            output_dir,
            config=IntegrationRunConfig(
                period_end=date(2024, 12, 31),
                input_name="synthetic-integration.xlsx",
            ),
        )
    finally:
        workbook.close()


# Descriptive aliases for callers that use the run terminology from the issue.
run_end_to_end = run_integration
run_local_synthetic = run_synthetic_integration


__all__ = [
    "CONTROL_SHEET_NAMES",
    "RUN_ARTIFACT_NAMES",
    "IntegrationRunConfig",
    "IntegrationRunError",
    "IntegrationRunResult",
    "build_synthetic_osv_workbook",
    "run_end_to_end",
    "run_integration",
    "run_local_synthetic",
    "run_synthetic_integration",
]
