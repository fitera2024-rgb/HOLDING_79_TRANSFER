"""Fail-closed parser for grouped 1C ОСВ XLSX workbooks.

The report is a grouped, hierarchical export rather than a flat table.  This
module intentionally keeps the hierarchy recovery separate from the transfer
engine: it emits only the canonical :class:`NormalizedBalance` model and
returns explicit diagnostics when the source cannot be interpreted safely.
"""

from __future__ import annotations

import io
import re
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from pathlib import Path
from typing import Any, BinaryIO
from xml.etree.ElementTree import ParseError as XmlParseError

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from .models import (
    BalanceStatus,
    BlockReason,
    NormalizedBalance,
    SourceAccount,
    normalize_decimal,
)

_END_BALANCE_HEADER = "сальдо на конец периода"
_MEASURE_HEADER_MARKERS = (
    "сальдо",
    "оборот",
    "остаток",
    "opening balance",
    "turnover",
    "ending balance",
)
_SIDE_LABELS = {
    "дебет": "debit",
    "дт": "debit",
    "debit": "debit",
    "dr": "debit",
    "кредит": "credit",
    "кт": "credit",
    "credit": "credit",
    "cr": "credit",
}
_ACCOUNT_RE = re.compile(r"\d+(?:\.(?:\d+|\*))+")
_ACCOUNT_LIKE_RE = re.compile(r"\d+[.,]\d+(?:[.,]\d+)*[A-Za-zА-Яа-я0-9._,*-]*")
_DATE_RE = re.compile(
    r"(?<!\d)(?P<day>\d{1,2})[./-](?P<month>\d{1,2})[./-](?P<year>\d{4})(?!\d)"
)
_ISO_DATE_RE = re.compile(r"(?<!\d)(?P<year>\d{4})-(?P<month>\d{1,2})-(?P<day>\d{1,2})(?!\d)")
_SOURCE_LOAD_ERRORS = (
    zipfile.BadZipFile,
    zipfile.LargeZipFile,
    InvalidFileException,
    OSError,
    EOFError,
    KeyError,
    TypeError,
    ValueError,
    XmlParseError,
)


class ParserDiagnosticCode(str, Enum):
    """Stable, parser-specific reason codes for blocked source rows."""

    MISSING_SHEET = "MISSING_SHEET"
    AMBIGUOUS_SHEET = "AMBIGUOUS_SHEET"
    INVALID_SOURCE = "INVALID_SOURCE"
    MISSING_PERIOD_END = "MISSING_PERIOD_END"
    AMBIGUOUS_PERIOD_END = "AMBIGUOUS_PERIOD_END"
    MISSING_ENDING_BALANCE_HEADERS = "MISSING_ENDING_BALANCE_HEADERS"
    AMBIGUOUS_ENDING_BALANCE_HEADERS = "AMBIGUOUS_ENDING_BALANCE_HEADERS"
    MALFORMED_ENDING_BALANCE = "MALFORMED_ENDING_BALANCE"
    INVALID_ENDING_BALANCE = "INVALID_ENDING_BALANCE"
    MISSING_ACCOUNT_CONTEXT = "MISSING_ACCOUNT_CONTEXT"
    AMBIGUOUS_ACCOUNT_CONTEXT = "AMBIGUOUS_ACCOUNT_CONTEXT"
    UNSUPPORTED_ACCOUNT = "UNSUPPORTED_ACCOUNT"
    MISSING_ORGANIZATION = "MISSING_ORGANIZATION"
    AMBIGUOUS_ORGANIZATION = "AMBIGUOUS_ORGANIZATION"
    MISSING_DEPARTMENT = "MISSING_DEPARTMENT"
    AMBIGUOUS_DEPARTMENT = "AMBIGUOUS_DEPARTMENT"
    MISSING_SUPPLIER_RVP = "MISSING_SUPPLIER_RVP"
    AMBIGUOUS_SUPPLIER_RVP = "AMBIGUOUS_SUPPLIER_RVP"
    AMBIGUOUS_HIERARCHY = "AMBIGUOUS_HIERARCHY"
    DUPLICATE_SOURCE_ROW = "DUPLICATE_SOURCE_ROW"

    # The aliases make the BLOCKED nature obvious to callers without changing
    # the compact serialized code used in diagnostics.
    BLOCKED_MISSING_ORGANIZATION = MISSING_ORGANIZATION
    BLOCKED_MISSING_DEPARTMENT = MISSING_DEPARTMENT
    BLOCKED_MISSING_SUPPLIER_RVP = MISSING_SUPPLIER_RVP
    BLOCKED_AMBIGUOUS_HIERARCHY = AMBIGUOUS_HIERARCHY
    BLOCKED_INVALID_ENDING_BALANCE = INVALID_ENDING_BALANCE


@dataclass(frozen=True)
class ParserDiagnostic:
    """One explicit reason why an XLSX source cannot be consumed safely."""

    code: ParserDiagnosticCode
    message: str
    sheet_name: str | None = None
    excel_row: int | None = None
    excel_column: int | None = None
    status: BalanceStatus = BalanceStatus.BLOCKED

    @property
    def row(self) -> int | None:
        """Short compatibility alias for consumers displaying row diagnostics."""

        return self.excel_row

    @property
    def reason(self) -> BlockReason:
        """Map parser failures to the existing canonical domain reasons."""

        if self.code in {
            ParserDiagnosticCode.INVALID_ENDING_BALANCE,
            ParserDiagnosticCode.MALFORMED_ENDING_BALANCE,
            ParserDiagnosticCode.MISSING_ENDING_BALANCE_HEADERS,
            ParserDiagnosticCode.AMBIGUOUS_ENDING_BALANCE_HEADERS,
        }:
            return BlockReason.BLOCKED_INVALID_ENDING_BALANCE
        if self.code in {
            ParserDiagnosticCode.MISSING_ORGANIZATION,
            ParserDiagnosticCode.AMBIGUOUS_ORGANIZATION,
            ParserDiagnosticCode.MISSING_DEPARTMENT,
            ParserDiagnosticCode.AMBIGUOUS_DEPARTMENT,
            ParserDiagnosticCode.MISSING_SUPPLIER_RVP,
            ParserDiagnosticCode.AMBIGUOUS_SUPPLIER_RVP,
            ParserDiagnosticCode.MISSING_ACCOUNT_CONTEXT,
            ParserDiagnosticCode.AMBIGUOUS_ACCOUNT_CONTEXT,
            ParserDiagnosticCode.UNSUPPORTED_ACCOUNT,
            ParserDiagnosticCode.AMBIGUOUS_HIERARCHY,
            ParserDiagnosticCode.MISSING_PERIOD_END,
            ParserDiagnosticCode.AMBIGUOUS_PERIOD_END,
            ParserDiagnosticCode.MISSING_SHEET,
            ParserDiagnosticCode.AMBIGUOUS_SHEET,
        }:
            return BlockReason.BLOCKED_MISSING_SOURCE_IDENTITY
        return BlockReason.BLOCKED_INVALID_POSTING


@dataclass(frozen=True)
class EndingBalanceColumns:
    """Columns discovered from the semantic ending-balance header."""

    debit_column: int
    credit_column: int
    header_rows: tuple[int, ...]

    @property
    def debit(self) -> int:
        return self.debit_column

    @property
    def credit(self) -> int:
        return self.credit_column


@dataclass(frozen=True)
class ParseResult:
    """Fail-closed parser output."""

    status: BalanceStatus
    balances: tuple[NormalizedBalance, ...] = ()
    diagnostics: tuple[ParserDiagnostic, ...] = ()

    @property
    def normalized_balances(self) -> tuple[NormalizedBalance, ...]:
        return self.balances

    @property
    def rows(self) -> tuple[NormalizedBalance, ...]:
        return self.balances

    @property
    def is_actionable(self) -> bool:
        return self.status is BalanceStatus.ACTIONABLE

    @property
    def blocked(self) -> bool:
        return self.status is BalanceStatus.BLOCKED

    @property
    def reason(self) -> BlockReason | None:
        return self.diagnostics[0].reason if self.diagnostics else None

    @property
    def message(self) -> str | None:
        return self.diagnostics[0].message if self.diagnostics else None


@dataclass
class _HierarchyState:
    account: SourceAccount | None = None
    organization: str | None = None
    department: str | None = None
    supplier_rvp: str | None = None
    account_depth: int = 0
    role_depths: dict[str, int] | None = None

    def __post_init__(self) -> None:
        if self.role_depths is None:
            self.role_depths = {}

    def reset_for_account(self, account: SourceAccount | None, account_depth: int | None = None) -> None:
        self.account = account
        self.organization = None
        self.department = None
        self.supplier_rvp = None
        self.account_depth = 0 if account_depth is None else account_depth
        self.role_depths = {}

    def set_identity(self, role: str, value: str) -> None:
        self.invalidate_for_boundary(role)
        if role == "organization":
            self.organization = value
        elif role == "department":
            self.department = value
        elif role == "supplier":
            self.supplier_rvp = value
        else:  # pragma: no cover - internal classifier guard
            raise ValueError(f"unsupported hierarchy role: {role}")

    def invalidate_for_boundary(self, role: str) -> None:
        """Drop the identity at a role boundary and every descendant."""

        if role == "organization":
            self.organization = None
            self.department = None
            self.supplier_rvp = None
        elif role == "department":
            self.department = None
            self.supplier_rvp = None
        elif role == "supplier":
            self.supplier_rvp = None
        else:  # pragma: no cover - internal classifier guard
            raise ValueError(f"unsupported hierarchy role: {role}")


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        return str(value).strip()
    normalized = unicodedata.normalize("NFKC", value)
    normalized = normalized.replace("\u00a0", " ")
    return " ".join(normalized.strip().casefold().split())


def _display_text(value: Any) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        return str(value).strip()
    return unicodedata.normalize("NFKC", value).replace("\u00a0", " ").strip()


def _leading_indent(value: Any) -> int | None:
    if not isinstance(value, str):
        return None
    prefix = value[: len(value) - len(value.lstrip())]
    return len(prefix) if prefix else None


def _side_label(value: Any) -> str | None:
    token = _normalize_text(value).strip(" :;,-—–")
    return _SIDE_LABELS.get(token)


def _side_in_text(value: Any) -> str | None:
    text = _normalize_text(value)
    if _END_BALANCE_HEADER not in text:
        return None
    found = [side for token, side in _SIDE_LABELS.items() if re.search(rf"\b{re.escape(token)}\b", text)]
    return found[0] if len(set(found)) == 1 else None


def _account_token(value: Any) -> str | None:
    """Return an exact account token, never a prefix/fuzzy match."""

    text = _display_text(value)
    if not text:
        return None
    match = re.fullmatch(r"(?:счет|account)\s*(?:№|:)?\s*(.+)", text, re.IGNORECASE)
    candidate = match.group(1).strip() if match else text
    if candidate == "79":
        return candidate
    if _ACCOUNT_RE.fullmatch(candidate):
        return candidate
    return None


def _account_boundary_token(value: Any) -> str | None:
    """Return supported or unsupported account-like boundary text.

    Account boundaries must be discovered independently of the supported
    account allow-list.  Otherwise an unsupported account can leave the
    previous 79.x context active and leak rows into the financial output.
    """

    token = _account_token(value)
    if token is not None:
        return token
    text = _display_text(value)
    match = re.fullmatch(r"(?:счет|account)\s*(?:№|:)?\s*(.+)", text, re.IGNORECASE)
    candidate = match.group(1).strip() if match else text
    if _ACCOUNT_LIKE_RE.fullmatch(candidate):
        return candidate
    return None


def _supported_account(value: str | None) -> SourceAccount | None:
    if value in {account.value for account in SourceAccount}:
        return SourceAccount(value)
    return None


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not isinstance(value, str):
        return None
    text = _display_text(value)
    matches = list(_ISO_DATE_RE.finditer(text))
    if matches:
        candidate = matches[-1]
        try:
            return date(
                int(candidate.group("year")),
                int(candidate.group("month")),
                int(candidate.group("day")),
            )
        except ValueError:
            return None
    matches = list(_DATE_RE.finditer(text))
    if not matches:
        return None
    candidate = matches[-1]
    try:
        return date(
            int(candidate.group("year")),
            int(candidate.group("month")),
            int(candidate.group("day")),
        )
    except ValueError:
        return None


def _date_candidates(value: Any) -> list[date]:
    if isinstance(value, datetime):
        return [value.date()]
    if isinstance(value, date):
        return [value]
    if not isinstance(value, str):
        return []
    result: list[date] = []
    for pattern in (_ISO_DATE_RE, _DATE_RE):
        for match in pattern.finditer(_display_text(value)):
            try:
                if pattern is _ISO_DATE_RE:
                    result.append(
                        date(
                            int(match.group("year")),
                            int(match.group("month")),
                            int(match.group("day")),
                        )
                    )
                else:
                    result.append(
                        date(
                            int(match.group("year")),
                            int(match.group("month")),
                            int(match.group("day")),
                        )
                    )
            except ValueError:
                continue
    return result


def _source_row_ref(sheet_name: str, row: int) -> str:
    """Stable trace reference independent of the workbook's machine path."""

    return f"{sheet_name}!R{row}"


def _detect_ending_columns(ws: Worksheet) -> EndingBalanceColumns | ParserDiagnostic:
    parents: list[tuple[int, int, int, Any, bool]] = []
    for row in range(1, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            value = ws.cell(row, column).value
            text = _normalize_text(value)
            if _END_BALANCE_HEADER not in text:
                continue
            span = None
            for merged in ws.merged_cells.ranges:
                if (
                    merged.min_row <= row <= merged.max_row
                    and merged.min_col <= column <= merged.max_col
                ):
                    span = merged
                    break
            if span is None:
                parents.append((row, column, column, value, False))
            elif (row, column) == (span.min_row, span.min_col):
                parents.append((row, span.min_col, span.max_col, value, True))

    candidates: list[tuple[int, int, tuple[int, ...]]] = []
    incomplete_group = False
    for parent_row, min_column, max_column, parent_value, merged in parents:
        occurrences: list[tuple[str, int, int]] = []
        direct_side = _side_in_text(parent_value)
        if direct_side is not None:
            occurrences.append((direct_side, parent_row, min_column))

        search_end = min(ws.max_row, parent_row + 4)
        if merged:
            search_columns = range(min_column, max_column + 1)
            for row in range(parent_row, search_end + 1):
                for column in search_columns:
                    side = _side_label(_display_text(ws.cell(row, column).value))
                    if side is not None:
                        occurrences.append((side, row, column))
        else:
            # An unmerged semantic parent has no span to delimit its group.
            # The only defensible child positions are the parent's column or
            # the one-column-right shifted pair.  A broader search would let
            # arbitrary adjacent labels elsewhere in the window complete a
            # malformed group.
            search_columns = range(min_column, min(ws.max_column, min_column + 2) + 1)
            for row in range(parent_row, search_end + 1):
                row_occurrences = [
                    (side, row, column)
                    for column in search_columns
                    if (side := _side_label(_display_text(ws.cell(row, column).value)))
                    is not None
                ]
                occurrences.extend(row_occurrences)

        debit_occurrences = [item for item in occurrences if item[0] == "debit"]
        credit_occurrences = [item for item in occurrences if item[0] == "credit"]
        group_candidates: list[tuple[int, int, tuple[int, ...]]] = []
        for _, debit_row, debit_column in debit_occurrences:
            for _, credit_row, credit_column in credit_occurrences:
                if debit_column == credit_column:
                    continue
                if not merged and (
                    debit_row != credit_row or abs(debit_column - credit_column) != 1
                ):
                    continue
                group_candidates.append(
                    (
                        debit_column,
                        credit_column,
                        tuple(sorted({parent_row, debit_row, credit_row})),
                    )
                )

        if not group_candidates:
            incomplete_group = True
        else:
            candidates.extend(group_candidates)

    if not candidates or incomplete_group:
        return ParserDiagnostic(
            ParserDiagnosticCode.MISSING_ENDING_BALANCE_HEADERS,
            "missing semantic 'Сальдо на конец периода' Debit/Credit columns",
            ws.title,
        )
    if len(candidates) != 1:
        return ParserDiagnostic(
            ParserDiagnosticCode.AMBIGUOUS_ENDING_BALANCE_HEADERS,
            "ambiguous semantic ending-balance Debit/Credit columns",
            ws.title,
        )
    debit, credit, rows = candidates[0]
    return EndingBalanceColumns(debit, credit, rows)


def _measure_columns_from_headers(
    ws: Worksheet, header_start: int, data_start: int
) -> set[int]:
    """Find columns structurally used for numeric balance/turnover measures.

    Grouped ОСВ exports commonly put several numeric sections beside the
    hierarchy column.  Their values can look exactly like account codes after
    Excel has converted them to numbers, so the header role is useful negative
    evidence when recovering the hierarchy column.
    """

    measure_columns: set[int] = set()
    for row in range(header_start, data_start):
        for column in range(1, ws.max_column + 1):
            value = ws.cell(row, column).value
            text = _normalize_text(value)
            is_measure_header = (
                _side_label(value) is not None
                or any(marker in text for marker in _MEASURE_HEADER_MARKERS)
            )
            if not is_measure_header:
                continue
            measure_columns.add(column)
            for merged in ws.merged_cells.ranges:
                if (
                    merged.min_row <= row <= merged.max_row
                    and merged.min_col <= column <= merged.max_col
                ):
                    measure_columns.update(range(merged.min_col, merged.max_col + 1))
                    break
    return measure_columns


def _is_numeric_like(value: Any) -> bool:
    """Return whether a cell is a numeric measure representation."""

    if value is None:
        return True
    try:
        _excel_decimal(value)
    except (TypeError, ValueError):
        return False
    return True


def _has_grouping_structure(
    ws: Worksheet,
    column: int,
    data_start: int,
    measure_columns: set[int],
) -> bool:
    """Require hierarchy semantics before accepting an account-token column."""

    explicit_roles = 0
    unlabelled_hierarchy_values = 0
    indented_hierarchy_values = 0
    for row in range(data_start, ws.max_row + 1):
        cell = ws.cell(row, column)
        value = cell.value
        if not _display_text(value):
            continue
        marker = _role_marker(value)
        if marker is not None:
            if marker[0] in {"organization", "department", "supplier"}:
                explicit_roles += 1
            # ОВ/ФВ and total labels are presentation rows, not hierarchy
            # evidence for a column that otherwise contains only amounts.
            continue
        if (
            column in measure_columns
            or _account_boundary_token(value) is not None
            or _is_numeric_like(value)
            or _side_label(value) is not None
        ):
            continue
        unlabelled_hierarchy_values += 1
        if (
            _leading_indent(value) is not None
            or (cell.alignment.indent is not None and cell.alignment.indent > 0)
        ):
            indented_hierarchy_values += 1

    if explicit_roles:
        return True
    return unlabelled_hierarchy_values >= 1 and indented_hierarchy_values > 0


def _find_grouping_column(
    ws: Worksheet, layout: EndingBalanceColumns
) -> int | ParserDiagnostic:
    data_start = layout.header_rows[-1] + 1
    measure_columns = _measure_columns_from_headers(
        ws, min(layout.header_rows), data_start
    )
    candidates: set[int] = set()
    account_marker_columns: set[int] = set()
    for row in range(data_start, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            if column in {layout.debit_column, layout.credit_column}:
                continue
            token = _account_boundary_token(ws.cell(row, column).value)
            if token is not None:
                previous = _normalize_text(ws.cell(row, column - 1).value) if column > 1 else ""
                candidate = column - 1 if previous in {"счет", "account"} else column
                # A token in a semantically numeric section is an amount, not
                # an account boundary.  Keep an explicit account-label/value
                # pair eligible because its label column is the hierarchy
                # column by construction.
                if candidate == column and column in measure_columns:
                    continue
                account_marker_columns.add(candidate)

    for candidate in account_marker_columns:
        if _has_grouping_structure(ws, candidate, data_start, measure_columns):
            candidates.add(candidate)
    if not candidates:
        return ParserDiagnostic(
            ParserDiagnosticCode.MISSING_ACCOUNT_CONTEXT,
            "no structurally supported account boundary was found in the grouped column",
            ws.title,
        )
    if len(candidates) > 1:
        return ParserDiagnostic(
            ParserDiagnosticCode.AMBIGUOUS_ACCOUNT_CONTEXT,
            "account markers occur in multiple possible grouping columns",
            ws.title,
        )
    return next(iter(candidates))


def _row_depth(ws: Worksheet, row: int, grouping_column: int) -> int | None:
    value = ws.cell(row, grouping_column).value
    leading = _leading_indent(value)
    if leading is not None:
        return leading
    indent = ws.cell(row, grouping_column).alignment.indent
    if indent is not None and indent > 0:
        return int(indent)
    outline_level = ws.row_dimensions[row].outlineLevel
    if outline_level > 0:
        return int(outline_level)
    return None


def _role_marker(value: Any) -> tuple[str, str | None] | None:
    text = _display_text(value)
    normalized = _normalize_text(value)
    if re.fullmatch(r"(?:ов|фв)(?:\s*[:\-].*)?", normalized):
        return "technical", None
    if re.match(r"^(?:итого|всего|total)(?:$|[\s:;-])", normalized):
        return "total", None

    prefixes = (
        ("organization", ("организация", "organization", "орг.")),
        ("department", ("цфо", "подразделение", "department", "департамент")),
        ("supplier", ("поставщик рвп", "поставщикрвп", "supplier rvp")),
    )
    for role, labels in prefixes:
        for label in labels:
            if normalized == label:
                return role, None
            if normalized.startswith((label + ":", label + " -")):
                suffix = text[len(text) - len(text.lstrip()) + len(label) :].strip(" :;-—–")
                return role, suffix
    return None


def _technical_row_classification(
    ws: Worksheet,
    row: int,
    grouping_column: int,
    state: _HierarchyState,
) -> bool | None:
    """Classify an ОВ/ФВ-like row using hierarchy depth.

    ``True`` means technical, ``False`` means a supplier-level row, and
    ``None`` means the structure is insufficient to distinguish the two.
    """

    marker = _role_marker(ws.cell(row, grouping_column).value)
    if marker is None or marker[0] != "technical":
        return False
    depth = _row_depth(ws, row, grouping_column)
    if depth is None or state.role_depths is None:
        return None
    supplier_depth = state.role_depths.get("supplier")
    if supplier_depth is None:
        department_depth = state.role_depths.get("department")
        if department_depth is None:
            return None
        supplier_depth = department_depth + 1
    if depth > supplier_depth:
        return True
    if depth == supplier_depth:
        return False
    return None


def _row_account_token(
    ws: Worksheet, row: int, grouping_column: int, layout: EndingBalanceColumns
) -> str | None:
    token = _account_boundary_token(ws.cell(row, grouping_column).value)
    if token is not None:
        return token
    if _normalize_text(ws.cell(row, grouping_column).value) not in {"счет", "account"}:
        return None
    for column in range(grouping_column + 1, ws.max_column + 1):
        if column in {layout.debit_column, layout.credit_column}:
            continue
        token = _account_boundary_token(ws.cell(row, column).value)
        if token is not None:
            return token
    return None


def _identity_candidates(
    ws: Worksheet,
    row: int,
    grouping_column: int,
    layout: EndingBalanceColumns,
    role: str,
) -> list[str]:
    grouping_value = ws.cell(row, grouping_column).value
    marker = _role_marker(grouping_value)
    if marker and marker[0] == role and marker[1] is not None:
        return [marker[1]]
    if marker and marker[0] != role:
        return []
    candidates: list[str] = []
    for column in range(1, ws.max_column + 1):
        if column in {layout.debit_column, layout.credit_column}:
            continue
        value = ws.cell(row, column).value
        text = _display_text(value)
        if not text:
            continue
        if column == grouping_column and marker is None:
            candidates.append(text)
            continue
        if _role_marker(value) is None and not _account_boundary_token(value):
            candidates.append(text)
    return list(dict.fromkeys(candidates))


def _has_balance_payload(ws: Worksheet, row: int, layout: EndingBalanceColumns) -> bool:
    for column in (layout.debit_column, layout.credit_column):
        value = ws.cell(row, column).value
        if value is None:
            continue
        if isinstance(value, str) and value.strip() in {"", "-", "—", "–"}:
            continue
        try:
            if _excel_decimal(value) != 0:
                return True
        except (TypeError, ValueError):
            # A non-blank, non-numeric value is still payload: it must not be
            # silently ignored just because it cannot be normalized.
            return True
    return False


def _missing_context_code(state: _HierarchyState) -> ParserDiagnosticCode | None:
    if state.account is None:
        return ParserDiagnosticCode.MISSING_ACCOUNT_CONTEXT
    if not state.organization:
        return ParserDiagnosticCode.MISSING_ORGANIZATION
    if not state.department:
        return ParserDiagnosticCode.MISSING_DEPARTMENT
    if not state.supplier_rvp:
        return ParserDiagnosticCode.MISSING_SUPPLIER_RVP
    return None


def _is_presentation_duplicate(
    ws: Worksheet,
    previous_row: int,
    current_row: int,
    grouping_column: int,
    layout: EndingBalanceColumns,
    state: _HierarchyState,
) -> bool:
    """Require technical and total rows as evidence of a repeated presentation.

    Equal business values alone are insufficient: two supplier leaves may be
    legitimate separate source rows.  The current supported presentation has
    an ОВ/ФВ detail and a following total between repeated leaf rows.
    """

    if current_row <= previous_row + 1:
        return False
    supplier_depth = state.role_depths.get("supplier") if state.role_depths else None
    previous_depth = _row_depth(ws, previous_row, grouping_column)
    current_depth = _row_depth(ws, current_row, grouping_column)
    if (
        supplier_depth is None
        or previous_depth is None
        or current_depth is None
        or previous_depth != supplier_depth
        or current_depth != supplier_depth
    ):
        return False

    has_technical = False
    has_total = False
    for row in range(previous_row + 1, current_row):
        if _account_boundary_token(ws.cell(row, grouping_column).value) is not None:
            return False
        marker = _role_marker(ws.cell(row, grouping_column).value)
        if marker is None:
            # Any unlabelled supplier-level row is a real leaf even when its
            # amount is zero.  A non-zero payload at any other row is also
            # evidence that this is not one repeated presentation.
            if _row_depth(ws, row, grouping_column) == supplier_depth and _display_text(
                ws.cell(row, grouping_column).value
            ):
                return False
            if _has_balance_payload(ws, row, layout):
                return False
            continue
        if marker[0] == "technical":
            if _technical_row_classification(ws, row, grouping_column, state) is not True:
                return False
            technical_depth = _row_depth(ws, row, grouping_column)
            if technical_depth is None or technical_depth <= supplier_depth or has_total:
                return False
            has_technical = True
        elif marker[0] == "total" and has_technical:
            total_depth = _row_depth(ws, row, grouping_column)
            if total_depth is None or total_depth != supplier_depth:
                return False
            has_total = True
        else:
            # An organization, department, supplier, or unknown marker is a
            # structural boundary, not evidence for the same presentation.
            return False
    return has_technical and has_total


def _excel_decimal(value: Any) -> Decimal:
    """Normalize Excel's numeric/string forms without binary arithmetic."""

    if isinstance(value, float):
        value = Decimal(str(value))
    elif isinstance(value, str):
        value = value.replace("\u00a0", " ").strip()
        if " " in value:
            value = value.replace(" ", "")
    return normalize_decimal(value)


def _normalized_ending_pair(debit: Decimal, credit: Decimal) -> tuple[Decimal, Decimal]:
    """Preserve explicit sides without inferring accounting direction."""

    if debit < 0 or credit < 0:
        raise ValueError("negative ending balances are unsupported")
    return debit, credit


def _diagnostic(
    code: ParserDiagnosticCode,
    message: str,
    ws: Worksheet,
    row: int | None = None,
    column: int | None = None,
) -> ParserDiagnostic:
    return ParserDiagnostic(code, message, ws.title, row, column)


def _period_from_workbook(ws: Worksheet) -> date | ParserDiagnostic:
    metadata_candidates: list[date] = []
    direct_candidates: list[date] = []
    for row in ws.iter_rows():
        for cell in row:
            dates = _date_candidates(cell.value)
            if not dates:
                continue
            text = _normalize_text(cell.value)
            if isinstance(cell.value, (date, datetime)) or any(
                marker in text for marker in ("период", "по состоянию", "на дату", "на конец")
            ):
                if isinstance(cell.value, str) and len(dates) > 1:
                    metadata_candidates.append(dates[-1])
                else:
                    metadata_candidates.extend(dates)
            else:
                direct_candidates.extend(dates)
    candidates = metadata_candidates or direct_candidates
    unique = sorted(set(candidates))
    if len(unique) == 1:
        return unique[0]
    if not unique:
        return ParserDiagnostic(
            ParserDiagnosticCode.MISSING_PERIOD_END,
            "period_end is not supplied and no report period end was found",
            ws.title,
        )
    return ParserDiagnostic(
        ParserDiagnosticCode.AMBIGUOUS_PERIOD_END,
        "more than one possible period_end was found",
        ws.title,
    )


def _select_worksheet(
    workbook: Workbook, sheet_name: str | None
) -> Worksheet | ParserDiagnostic:
    if sheet_name is not None:
        if sheet_name not in workbook.sheetnames:
            return ParserDiagnostic(
                ParserDiagnosticCode.MISSING_SHEET,
                f"worksheet not found: {sheet_name}",
            )
        return workbook[sheet_name]
    if len(workbook.worksheets) == 1:
        return workbook.worksheets[0]
    preferred = [
        ws
        for ws in workbook.worksheets
        if "осв" in _normalize_text(ws.title) or "osv" in _normalize_text(ws.title)
    ]
    if len(preferred) == 1:
        return preferred[0]
    return ParserDiagnostic(
        ParserDiagnosticCode.AMBIGUOUS_SHEET,
        "multiple worksheets exist and no ОСВ worksheet was selected",
    )


def _load_source(source: Any) -> tuple[Workbook, bool] | Worksheet:
    if isinstance(source, Worksheet):
        return source
    if isinstance(source, Workbook):
        return source, False
    if isinstance(source, (str, Path)):
        return load_workbook(source, data_only=True), True
    if isinstance(source, (bytes, bytearray)):
        return load_workbook(io.BytesIO(bytes(source)), data_only=True), True
    if hasattr(source, "read"):
        stream: BinaryIO = source
        if hasattr(stream, "seek"):
            stream.seek(0)
        return load_workbook(stream, data_only=True), True
    raise TypeError(f"unsupported XLSX source type: {type(source).__name__}")


class GroupedOsvParser:
    """Parse grouped 1C ОСВ input into canonical normalized balances."""

    def __init__(
        self,
        period_end: date | datetime | str | None = None,
        *,
        sheet_name: str | None = None,
    ) -> None:
        self.period_end = _parse_date(period_end) if period_end is not None else None
        if period_end is not None and self.period_end is None:
            raise ValueError(f"invalid period_end: {period_end!r}")
        self.sheet_name = sheet_name

    def parse(self, source: Any) -> ParseResult:
        loaded = None
        close_workbook = False
        try:
            try:
                loaded = _load_source(source)
            except _SOURCE_LOAD_ERRORS:
                diagnostic = ParserDiagnostic(
                    ParserDiagnosticCode.INVALID_SOURCE,
                    "source workbook is not a valid XLSX file",
                )
                return ParseResult(BalanceStatus.BLOCKED, diagnostics=(diagnostic,))

            if isinstance(loaded, Worksheet):
                worksheet = loaded
            else:
                workbook, close_workbook = loaded
                selected = _select_worksheet(workbook, self.sheet_name)
                if isinstance(selected, ParserDiagnostic):
                    return ParseResult(BalanceStatus.BLOCKED, diagnostics=(selected,))
                worksheet = selected
            return self._parse_worksheet(worksheet)
        finally:
            if close_workbook and loaded is not None and not isinstance(loaded, Worksheet):
                loaded[0].close()

    def _parse_worksheet(self, ws: Worksheet) -> ParseResult:
        layout = _detect_ending_columns(ws)
        if isinstance(layout, ParserDiagnostic):
            return ParseResult(BalanceStatus.BLOCKED, diagnostics=(layout,))

        period_end = self.period_end
        if period_end is None:
            discovered = _period_from_workbook(ws)
            if isinstance(discovered, ParserDiagnostic):
                return ParseResult(BalanceStatus.BLOCKED, diagnostics=(discovered,))
            period_end = discovered

        grouping_column = _find_grouping_column(ws, layout)
        if isinstance(grouping_column, ParserDiagnostic):
            return ParseResult(BalanceStatus.BLOCKED, diagnostics=(grouping_column,))

        state = _HierarchyState()
        balances: list[NormalizedBalance] = []
        diagnostics: list[ParserDiagnostic] = []
        seen_source_keys: dict[tuple[Any, ...], list[tuple[int, Decimal, Decimal]]] = {}
        start_row = layout.header_rows[-1] + 1

        for row in range(start_row, ws.max_row + 1):
            grouping_value = ws.cell(row, grouping_column).value
            account = _row_account_token(ws, row, grouping_column, layout)
            if account is not None:
                supported = _supported_account(account)
                state.reset_for_account(
                    supported,
                    _row_depth(ws, row, grouping_column),
                )
                if supported is None:
                    diagnostics.append(
                        _diagnostic(
                            ParserDiagnosticCode.UNSUPPORTED_ACCOUNT,
                            f"unsupported exact account context: {account}",
                            ws,
                            row,
                            grouping_column,
                        )
                    )
                if _has_balance_payload(ws, row, layout):
                    diagnostics.append(
                        _diagnostic(
                            ParserDiagnosticCode.MISSING_ORGANIZATION
                            if supported is not None
                            else ParserDiagnosticCode.MISSING_ACCOUNT_CONTEXT,
                            "financial payload occurs on an account boundary before a complete hierarchy context",
                            ws,
                            row,
                            grouping_column,
                        )
                    )
                continue

            if state.account is None:
                if _has_balance_payload(ws, row, layout):
                    diagnostics.append(
                        _diagnostic(
                            ParserDiagnosticCode.MISSING_ACCOUNT_CONTEXT,
                            "financial leaf row occurs outside a supported account context",
                            ws,
                            row,
                            grouping_column,
                        )
                    )
                continue

            marker = _role_marker(grouping_value)
            if marker and marker[0] == "total":
                if _has_balance_payload(ws, row, layout):
                    code = _missing_context_code(state)
                    if code is not None:
                        diagnostics.append(
                            _diagnostic(
                                code,
                                "financial payload on a total row lacks complete hierarchy context",
                                ws,
                                row,
                                grouping_column,
                            )
                        )
                continue
            if marker and marker[0] == "technical":
                technical = _technical_row_classification(
                    ws,
                    row,
                    grouping_column,
                    state,
                )
                if technical is True:
                    if _has_balance_payload(ws, row, layout):
                        code = _missing_context_code(state)
                        if code is not None:
                            diagnostics.append(
                                _diagnostic(
                                    code,
                                    "financial payload on a technical row lacks complete hierarchy context",
                                    ws,
                                    row,
                                    grouping_column,
                                )
                            )
                    continue
                if technical is None:
                    if _has_balance_payload(ws, row, layout):
                        diagnostics.append(
                            _diagnostic(
                                ParserDiagnosticCode.AMBIGUOUS_HIERARCHY,
                                "ОВ/ФВ-like row cannot be distinguished from a supplier leaf",
                                ws,
                                row,
                                grouping_column,
                            )
                        )
                    continue
            if not _display_text(grouping_value) and not _has_balance_payload(ws, row, layout):
                continue

            role, value_or_diagnostic = self._classify_row(ws, row, grouping_column, layout, state)
            if isinstance(value_or_diagnostic, ParserDiagnostic):
                diagnostics.append(value_or_diagnostic)
                continue
            if role is None:
                continue
            if value_or_diagnostic is None:
                # A bare hierarchy label is a presentation header.  A row
                # with amounts, however, is a malformed financial leaf.
                if _has_balance_payload(ws, row, layout):
                    code = {
                        "organization": ParserDiagnosticCode.MISSING_ORGANIZATION,
                        "department": ParserDiagnosticCode.MISSING_DEPARTMENT,
                        "supplier": ParserDiagnosticCode.MISSING_SUPPLIER_RVP,
                    }[role]
                    diagnostics.append(
                        _diagnostic(
                            code,
                            f"{role} identity is missing on a financial hierarchy row",
                            ws,
                            row,
                            grouping_column,
                        )
                    )
                continue

            state.set_identity(role, value_or_diagnostic)
            if role != "supplier":
                if _has_balance_payload(ws, row, layout):
                    code = _missing_context_code(state)
                    if code is not None:
                        diagnostics.append(
                            _diagnostic(
                                code,
                                f"financial payload on a {role} hierarchy row lacks complete context",
                                ws,
                                row,
                                grouping_column,
                            )
                        )
                continue

            if not state.organization:
                diagnostics.append(
                    _diagnostic(
                        ParserDiagnosticCode.MISSING_ORGANIZATION,
                        "supplier leaf has no organization context",
                        ws,
                        row,
                        grouping_column,
                    )
                )
                continue
            if not state.department:
                diagnostics.append(
                    _diagnostic(
                        ParserDiagnosticCode.MISSING_DEPARTMENT,
                        "supplier leaf has no department/ЦФО context",
                        ws,
                        row,
                        grouping_column,
                    )
                )
                continue
            if not state.supplier_rvp:
                diagnostics.append(
                    _diagnostic(
                        ParserDiagnosticCode.MISSING_SUPPLIER_RVP,
                        "supplier leaf has an empty Поставщик РВП identity",
                        ws,
                        row,
                        grouping_column,
                    )
                )
                continue

            try:
                debit = _excel_decimal(ws.cell(row, layout.debit_column).value)
                credit = _excel_decimal(ws.cell(row, layout.credit_column).value)
                if debit < 0 or credit < 0:
                    diagnostics.append(
                        _diagnostic(
                            ParserDiagnosticCode.INVALID_ENDING_BALANCE,
                            "negative ending balances are unsupported; accounting side was not inferred",
                            ws,
                            row,
                        )
                    )
                    continue
                debit, credit = _normalized_ending_pair(debit, credit)
            except (TypeError, ValueError) as exc:
                diagnostics.append(
                    _diagnostic(
                        ParserDiagnosticCode.MALFORMED_ENDING_BALANCE,
                        f"malformed ending balance: {exc}",
                        ws,
                        row,
                    )
                )
                continue

            source_key = (
                period_end,
                state.organization,
                state.account,
                state.department,
                state.supplier_rvp,
            )
            previous_rows = seen_source_keys.get(source_key, [])
            presentation_previous = next(
                (
                    previous
                    for previous in reversed(previous_rows)
                    if _is_presentation_duplicate(
                        ws,
                        previous[0],
                        row,
                        grouping_column,
                        layout,
                        state,
                    )
                ),
                None,
            )
            if presentation_previous is not None:
                previous_row, previous_debit, previous_credit = presentation_previous
                if (previous_debit, previous_credit) == (debit, credit):
                    continue
                diagnostics.append(
                    _diagnostic(
                        ParserDiagnosticCode.DUPLICATE_SOURCE_ROW,
                        "the same presentation source identity has conflicting rows "
                        f"at rows {previous_row} and {row}",
                        ws,
                        row,
                    )
                )
                continue

            if debit and credit:
                diagnostics.append(
                    _diagnostic(
                        ParserDiagnosticCode.INVALID_ENDING_BALANCE,
                        "ending debit and credit are both non-zero after normalization",
                        ws,
                        row,
                    )
                )
                continue

            try:
                balance = NormalizedBalance(
                    period_end=period_end,
                    organization=state.organization,
                    source_account=state.account,
                    department=state.department,
                    supplier_rvp=state.supplier_rvp,
                    ending_debit=debit,
                    ending_credit=credit,
                    source_excel_row_ref=_source_row_ref(ws.title, row),
                )
            except (TypeError, ValueError) as exc:
                diagnostics.append(
                    _diagnostic(
                        ParserDiagnosticCode.MALFORMED_ENDING_BALANCE,
                        f"cannot construct canonical normalized balance: {exc}",
                        ws,
                        row,
                    )
                )
                continue
            seen_source_keys.setdefault(source_key, []).append((row, debit, credit))
            balances.append(balance)

        if diagnostics:
            return ParseResult(BalanceStatus.BLOCKED, diagnostics=tuple(diagnostics))
        status = (
            BalanceStatus.ACTIONABLE
            if any(balance.status is BalanceStatus.ACTIONABLE for balance in balances)
            else BalanceStatus.NO_ACTION
        )
        return ParseResult(status, balances=tuple(balances))

    def _classify_row(
        self,
        ws: Worksheet,
        row: int,
        grouping_column: int,
        layout: EndingBalanceColumns,
        state: _HierarchyState,
    ) -> tuple[str | None, str | ParserDiagnostic | None]:
        grouping_value = ws.cell(row, grouping_column).value
        marker = _role_marker(grouping_value)
        if marker is not None:
            role, marker_value = marker
            if role == "total":
                return None, None
            if role != "technical":
                state.invalidate_for_boundary(role)
                if role == "department" and not state.organization:
                    return role, _diagnostic(
                        ParserDiagnosticCode.MISSING_ORGANIZATION,
                        "department row occurs before an organization context",
                        ws,
                        row,
                        grouping_column,
                    )
                if role == "supplier":
                    if not state.organization:
                        return role, _diagnostic(
                            ParserDiagnosticCode.MISSING_ORGANIZATION,
                            "supplier row occurs before an organization context",
                            ws,
                            row,
                            grouping_column,
                        )
                    if not state.department:
                        return role, _diagnostic(
                            ParserDiagnosticCode.MISSING_DEPARTMENT,
                            "supplier row occurs before a department context",
                            ws,
                            row,
                            grouping_column,
                        )
                if marker_value is None:
                    candidates = _identity_candidates(ws, row, grouping_column, layout, role)
                    if len(candidates) == 1:
                        marker_value = candidates[0]
                    elif len(candidates) > 1:
                        code = {
                            "organization": ParserDiagnosticCode.AMBIGUOUS_ORGANIZATION,
                            "department": ParserDiagnosticCode.AMBIGUOUS_DEPARTMENT,
                            "supplier": ParserDiagnosticCode.AMBIGUOUS_SUPPLIER_RVP,
                        }[role]
                        return role, _diagnostic(
                            code,
                            f"ambiguous {role} identity candidates: {candidates}",
                            ws,
                            row,
                            grouping_column,
                        )
                depth = _row_depth(ws, row, grouping_column)
                depth_diagnostic = self._validate_role_depth(
                    ws,
                    row,
                    grouping_column,
                    state,
                    role,
                    depth,
                )
                if depth_diagnostic is not None:
                    return role, depth_diagnostic
                self._record_role_depth(state, role, depth)
                return role, marker_value

        value = _display_text(grouping_value)
        depth = _row_depth(ws, row, grouping_column)
        role = self._role_from_depth(state, depth)
        if role is None:
            return None, _diagnostic(
                ParserDiagnosticCode.AMBIGUOUS_HIERARCHY,
                "hierarchy row cannot be assigned to organization, department, or supplier",
                ws,
                row,
                grouping_column,
            )
        depth_diagnostic = self._validate_role_depth(
            ws,
            row,
            grouping_column,
            state,
            role,
            depth,
        )
        if depth_diagnostic is not None:
            return role, depth_diagnostic
        self._record_role_depth(state, role, depth)
        if value:
            return role, value
        candidates = _identity_candidates(ws, row, grouping_column, layout, role)
        if len(candidates) == 1:
            return role, candidates[0]
        if len(candidates) > 1:
            code = {
                "organization": ParserDiagnosticCode.AMBIGUOUS_ORGANIZATION,
                "department": ParserDiagnosticCode.AMBIGUOUS_DEPARTMENT,
                "supplier": ParserDiagnosticCode.AMBIGUOUS_SUPPLIER_RVP,
            }[role]
            return role, _diagnostic(
                code,
                f"ambiguous {role} identity candidates: {candidates}",
                ws,
                row,
                grouping_column,
            )
        return role, None

    @staticmethod
    def _record_role_depth(state: _HierarchyState, role: str, depth: int | None) -> None:
        if depth is None:
            return
        assert state.role_depths is not None
        existing = state.role_depths.get(role)
        if existing is None:
            state.role_depths[role] = depth

    @staticmethod
    def _validate_role_depth(
        ws: Worksheet,
        row: int,
        grouping_column: int,
        state: _HierarchyState,
        role: str,
        depth: int | None,
    ) -> ParserDiagnostic | None:
        if depth is None or state.role_depths is None:
            return None
        known_depth = state.role_depths.get(role)
        if known_depth is not None and known_depth != depth:
            return _diagnostic(
                ParserDiagnosticCode.AMBIGUOUS_HIERARCHY,
                f"{role} row depth {depth} conflicts with established depth {known_depth}",
                ws,
                row,
                grouping_column,
            )
        if role == "organization" and depth <= state.account_depth:
            return _diagnostic(
                ParserDiagnosticCode.AMBIGUOUS_HIERARCHY,
                "organization row does not follow the account hierarchy level",
                ws,
                row,
                grouping_column,
            )
        parent_role = {
            "department": "organization",
            "supplier": "department",
        }.get(role)
        if parent_role is not None:
            parent_depth = state.role_depths.get(parent_role)
            if parent_depth is not None and depth <= parent_depth:
                return _diagnostic(
                    ParserDiagnosticCode.AMBIGUOUS_HIERARCHY,
                    f"{role} row does not follow the {parent_role} hierarchy level",
                    ws,
                    row,
                    grouping_column,
                )
        return None

    @staticmethod
    def _role_from_depth(state: _HierarchyState, depth: int | None) -> str | None:
        if not state.organization:
            if depth is None or depth != state.account_depth + 1:
                return None
            return "organization"
        if not state.department:
            organization_depth = state.role_depths.get("organization") if state.role_depths else None
            if organization_depth is None or depth is None or depth != organization_depth + 1:
                return None
            return "department"
        if not state.supplier_rvp:
            department_depth = state.role_depths.get("department") if state.role_depths else None
            if department_depth is None or depth is None or depth != department_depth + 1:
                return None
            return "supplier"
        if depth is not None and state.role_depths:
            known = state.role_depths.get("supplier")
            if known is not None and depth == known:
                return "supplier"
            organization_depth = state.role_depths.get("organization")
            if organization_depth is not None and depth == organization_depth:
                return "organization"
            department_depth = state.role_depths.get("department")
            if department_depth is not None and depth == department_depth:
                return "department"
            return None
        # Without depth metadata an unlabelled row cannot safely distinguish a
        # supplier from a new organization or department boundary.
        return None


def parse_grouped_osv(
    source: Any,
    period_end: date | datetime | str | None = None,
    *,
    sheet_name: str | None = None,
) -> ParseResult:
    """Parse a grouped 1C ОСВ workbook or workbook-like XLSX source."""

    return GroupedOsvParser(period_end, sheet_name=sheet_name).parse(source)


parse_osv = parse_grouped_osv
parse_grouped_osv_xlsx = parse_grouped_osv


__all__ = [
    "EndingBalanceColumns",
    "GroupedOsvParser",
    "ParseResult",
    "ParserDiagnostic",
    "ParserDiagnosticCode",
    "parse_grouped_osv",
    "parse_grouped_osv_xlsx",
    "parse_osv",
]
