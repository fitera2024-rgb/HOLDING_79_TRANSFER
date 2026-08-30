from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from holding79_transfer import (
    CONTROL_SHEET_NAMES,
    OUTPUT_HEADERS,
    OUTPUT_SHEET_NAME,
    IntegrationRunConfig,
    IntegrationRunError,
    TransferConfig,
    build_synthetic_osv_workbook,
    run_integration,
    run_synthetic_integration,
)


def read_jsonl(path: Path) -> list[dict[str, object]]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines()]


def test_synthetic_run_publishes_all_artifacts_and_controls(tmp_path: Path):
    result = run_synthetic_integration(tmp_path / "run")

    assert result.is_success
    assert len(result.normalized_balances) == 4
    assert len(result.posting_rows) == 8
    assert len(result.diagnostics) == 1
    assert {balance.source_account.value for balance in result.normalized_balances} == {"79.2", "79.3"}
    assert {balance.ending_side.value for balance in result.normalized_balances} == {"DEBIT", "CREDIT"}

    expected_paths = {
        "input_manifest.json",
        "normalized_balances.jsonl",
        "posting_rows.jsonl",
        "summary.json",
        "run_control.xlsx",
        "export_manifest.json",
        "export/2024-12-31__АТ.xlsx",
        "export/2024-12-31__ГК.xlsx",
    }
    actual_paths = {
        path.relative_to(result.output_dir).as_posix()
        for path in result.output_dir.rglob("*")
        if path.is_file()
    }
    assert actual_paths == expected_paths

    summary = json.loads((result.output_dir / "summary.json").read_text(encoding="utf-8"))
    assert summary["status"] == "SUCCESS"
    assert summary["counts"] == {
        "actionable_source_rows": 4,
        "blocked_source_rows": 1,
        "export_rows": 8,
        "export_workbooks": 2,
        "no_action_source_rows": 0,
        "normalized_balances": 4,
        "parser_diagnostics": 1,
        "posting_rows": 8,
        "source_rows": 5,
    }
    assert summary["totals"] == {
        "difference": "0",
        "gk": "337089.60",
        "source_org": "337089.60",
    }

    control_workbook = load_workbook(result.output_dir / "run_control.xlsx", read_only=True)
    try:
        assert control_workbook.sheetnames == list(CONTROL_SHEET_NAMES)
        blocked_rows = list(control_workbook["Блокировки"].iter_rows(values_only=True))
        assert blocked_rows[1][0] == "BLOCKED_SOURCE_ROW!R9"
        assert blocked_rows[1][4] == "MISSING_SUPPLIER_RVP"
        assert blocked_rows[1][7:9] == (0, 0)
        effect_rows = list(control_workbook["Контроль_до_после"].iter_rows(values_only=True))
        assert len(effect_rows) == 5
        assert all(row[8:10] == ("0", "0") for row in effect_rows[1:])
    finally:
        control_workbook.close()

    for path in result.output_dir.glob("export/*.xlsx"):
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            assert workbook.sheetnames == [OUTPUT_SHEET_NAME]
            worksheet = workbook[OUTPUT_SHEET_NAME]
            assert tuple(cell.value for cell in worksheet[1]) == OUTPUT_HEADERS
            assert worksheet.max_column == 27
            assert worksheet.max_row == 5
        finally:
            workbook.close()


def test_blocked_parser_row_has_no_financial_output_and_is_retained_in_manifests(tmp_path: Path):
    result = run_synthetic_integration(tmp_path / "run")

    normalized = read_jsonl(result.output_dir / "normalized_balances.jsonl")
    postings = read_jsonl(result.output_dir / "posting_rows.jsonl")
    input_manifest = json.loads(
        (result.output_dir / "input_manifest.json").read_text(encoding="utf-8")
    )
    export_manifest = json.loads(
        (result.output_dir / "export_manifest.json").read_text(encoding="utf-8")
    )

    assert len(normalized) == 4
    assert all(record["status"] == "ACTIONABLE" for record in normalized)
    assert all(record["source_excel_row_ref"] != "BLOCKED_SOURCE_ROW!R9" for record in postings)
    assert input_manifest["parser_diagnostics"][0]["source_excel_row_ref"] == (
        "BLOCKED_SOURCE_ROW!R9"
    )
    assert all(
        "BLOCKED_SOURCE_ROW!R9" not in str(record)
        for record in export_manifest["workbooks"]
    )


def test_same_normalized_input_is_byte_stable_for_financial_artifacts(tmp_path: Path):
    first = run_synthetic_integration(tmp_path / "first")
    second = run_synthetic_integration(tmp_path / "second")

    assert first.run_id == second.run_id
    assert (first.output_dir / "normalized_balances.jsonl").read_bytes() == (
        second.output_dir / "normalized_balances.jsonl"
    ).read_bytes()
    assert (first.output_dir / "posting_rows.jsonl").read_bytes() == (
        second.output_dir / "posting_rows.jsonl"
    ).read_bytes()
    assert [row.financial_record_id for row in first.posting_rows] == [
        row.financial_record_id for row in second.posting_rows
    ]


def test_zero_balance_run_publishes_auditable_empty_export(tmp_path: Path):
    workbook = build_synthetic_osv_workbook()
    try:
        workbook.remove(workbook["BLOCKED_SOURCE_ROW"])
        for worksheet in workbook.worksheets:
            worksheet["H9"] = "0"
            worksheet["I9"] = "0"
        result = run_integration(workbook, tmp_path / "run", period_end="2024-12-31")
    finally:
        workbook.close()

    assert result.is_success
    assert result.posting_rows == ()
    assert result.exported_workbooks == ()
    assert all(balance.status.value == "NO_ACTION" for balance in result.normalized_balances)
    assert (result.output_dir / "export").is_dir()
    assert list((result.output_dir / "export").iterdir()) == []

    summary = json.loads((result.output_dir / "summary.json").read_text(encoding="utf-8"))
    assert summary["counts"] == {
        "actionable_source_rows": 0,
        "blocked_source_rows": 0,
        "export_rows": 0,
        "export_workbooks": 0,
        "no_action_source_rows": 4,
        "normalized_balances": 4,
        "parser_diagnostics": 0,
        "posting_rows": 0,
        "source_rows": 4,
    }
    export_manifest = json.loads(
        (result.output_dir / "export_manifest.json").read_text(encoding="utf-8")
    )
    assert export_manifest["financial_row_count"] == 0
    assert export_manifest["workbooks"] == []
    control_workbook = load_workbook(result.output_dir / "run_control.xlsx", read_only=True)
    try:
        assert control_workbook.sheetnames == list(CONTROL_SHEET_NAMES)
    finally:
        control_workbook.close()


def test_financial_record_id_collision_across_source_sheets_fails_closed(tmp_path: Path):
    workbook = build_synthetic_osv_workbook()
    duplicate = workbook.copy_worksheet(workbook["DEBIT_79_2_AT"])
    duplicate.title = "DUPLICATE_DEBIT_79_2"
    try:
        with pytest.raises(IntegrationRunError, match="financial_record_id"):
            run_integration(workbook, tmp_path / "run", period_end="2024-12-31")
    finally:
        workbook.close()
    assert not (tmp_path / "run").exists()


def test_corrupt_summary_json_is_rejected_before_publication(tmp_path: Path, monkeypatch):
    import holding79_transfer.integration as integration_module

    original_write_json = integration_module._write_json

    def corrupt_summary(path: Path, value: object) -> None:
        original_write_json(path, value)
        if path.name == "summary.json":
            path.write_text("{not valid json", encoding="utf-8")

    monkeypatch.setattr(integration_module, "_write_json", corrupt_summary)
    with pytest.raises(IntegrationRunError, match="summary.json"):
        run_synthetic_integration(tmp_path / "run")
    assert not (tmp_path / "run").exists()


def test_on_disk_synthetic_input_uses_the_same_pipeline(tmp_path: Path):
    input_path = tmp_path / "synthetic-input.xlsx"
    workbook = build_synthetic_osv_workbook()
    try:
        workbook.save(input_path)
    finally:
        workbook.close()

    result = run_integration(input_path, tmp_path / "run")

    assert result.is_success
    assert len(result.posting_rows) == 8
    assert result.diagnostics[0].code.value == "MISSING_SUPPLIER_RVP"


def test_run_fails_closed_without_publishing_partial_outputs(tmp_path: Path):
    output_dir = tmp_path / "run"
    output_dir.mkdir()

    with pytest.raises(IntegrationRunError, match="already exists"):
        run_synthetic_integration(output_dir)
    assert list(output_dir.iterdir()) == []


def test_unapproved_rules_fail_closed_without_publishing_partial_outputs(tmp_path: Path):
    workbook = build_synthetic_osv_workbook()
    try:
        with pytest.raises(IntegrationRunError, match="mandatory control failed"):
            run_integration(
                workbook,
                tmp_path / "run",
                config=IntegrationRunConfig(
                    period_end="2024-12-31",
                    transfer_config=TransferConfig(rules_version="H79_TRANSFER_V2"),
                ),
            )
    finally:
        workbook.close()
    assert not (tmp_path / "run").exists()


def test_bytes_input_is_supported_without_path_trace_in_artifacts(tmp_path: Path):
    workbook = build_synthetic_osv_workbook()
    stream = BytesIO()
    try:
        workbook.save(stream)
    finally:
        workbook.close()

    result = run_integration(stream.getvalue(), tmp_path / "run", input_name="synthetic.xlsx")
    manifest = json.loads(
        (result.output_dir / "input_manifest.json").read_text(encoding="utf-8")
    )
    assert manifest["input_name"] == "synthetic.xlsx"
    assert str(tmp_path) not in (result.output_dir / "input_manifest.json").read_text(
        encoding="utf-8"
    )
