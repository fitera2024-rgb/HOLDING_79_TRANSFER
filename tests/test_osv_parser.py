from __future__ import annotations

import zipfile
from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment

from holding79_transfer import (
    BalanceStatus,
    GroupedOsvParser,
    ParserDiagnostic,
    ParserDiagnosticCode,
    TransferEngine,
    parse_grouped_osv,
)
from holding79_transfer.parser import _detect_ending_columns, _find_grouping_column


def make_workbook(
    rows: list[tuple[str, object, object, int | None]],
    *,
    header_row: int = 4,
    sheet_name: str = "ОСВ",
) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.cell(1, 1).value = "Синтетическая ОСВ"
    worksheet.cell(2, 1).value = "Период: 01.12.2024 - 31.12.2024"
    worksheet.merge_cells(
        start_row=header_row,
        start_column=8,
        end_row=header_row,
        end_column=9,
    )
    worksheet.cell(header_row, 8).value = "Сальдо на конец периода"
    worksheet.cell(header_row + 1, 8).value = "Дебет"
    worksheet.cell(header_row + 1, 9).value = "Кредит"

    for offset, (grouping, debit, credit, indent) in enumerate(rows, header_row + 2):
        cell = worksheet.cell(offset, 1)
        cell.value = grouping
        if indent is not None:
            cell.alignment = Alignment(indent=indent)
        worksheet.cell(offset, 8).value = debit
        worksheet.cell(offset, 9).value = credit
    return workbook


def parse(rows: list[tuple[str, object, object, int | None]], **kwargs):
    return parse_grouped_osv(make_workbook(rows, **kwargs), period_end=date(2024, 12, 31))


def workbook_bytes(workbook: Workbook) -> bytes:
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def leaf(
    account: str,
    organization: str,
    department: str,
    supplier: str,
    debit: object = 0,
    credit: object = 0,
) -> list[tuple[str, object, object, int | None]]:
    return [
        (account, None, None, 0),
        (f"Организация: {organization}", None, None, 1),
        (f"ЦФО: {department}", None, None, 2),
        (f"Поставщик РВП: {supplier}", debit, credit, 3),
    ]


def test_supported_accounts_and_both_ending_sides_are_normalized():
    rows = leaf("79.2", "АТ", "Б_АТ Коммерческий отдел", "Дебетовый", "84272.40")
    rows += [("Поставщик РВП: Кредитовый", 0, "84272.40", 3)]
    rows += leaf("79.3", "БТ", "Б_БТ Финансы", "Дебетовый", "1.25")
    rows += [("Поставщик РВП: Кредитовый", 0, "2.50", 3)]

    result = parse(rows)

    assert result.status is BalanceStatus.ACTIONABLE
    assert len(result.balances) == 4
    assert [balance.source_account.value for balance in result.balances] == [
        "79.2",
        "79.2",
        "79.3",
        "79.3",
    ]
    assert result.balances[0].ending_debit == Decimal("84272.40")
    assert result.balances[1].ending_credit == Decimal("84272.40")
    assert result.balances[2].ending_debit == Decimal("1.25")
    assert result.balances[3].ending_credit == Decimal("2.50")
    assert all(isinstance(balance.ending_debit, Decimal) for balance in result.balances)


def test_zero_balance_is_no_action_and_keeps_canonical_identity():
    result = parse(leaf("79.2", "АТ", "ЦФО", "Нулевой"))

    assert result.status is BalanceStatus.NO_ACTION
    assert len(result.balances) == 1
    assert result.balances[0].status is BalanceStatus.NO_ACTION


def test_technical_ov_fv_rows_and_identical_presentation_rows_do_not_duplicate_leaf():
    rows = leaf("79.2", "АТ", "ЦФО", "Производитель", "10.00")
    rows += [
        ("ОВ", "10.00", 0, 4),
        ("ФВ", "10.00", 0, 4),
        ("Итого по поставщику", "10.00", 0, 3),
        ("Поставщик РВП: Производитель", "10.00", 0, 3),
    ]

    result = parse(rows)

    assert result.status is BalanceStatus.ACTIONABLE
    assert len(result.balances) == 1
    assert result.balances[0].supplier_rvp == "Производитель"


def test_presentation_dedup_does_not_cross_an_intervening_financial_supplier_leaf():
    rows = leaf("79.2", "АТ", "ЦФО", "Supplier A", "10.00")
    rows += [
        ("Supplier B", "20.00", 0, 3),
        ("ОВ", "20.00", 0, 4),
        ("ФВ", "20.00", 0, 4),
        ("Итого по поставщику", "20.00", 0, 3),
        ("Поставщик РВП: Supplier A", "10.00", 0, 3),
    ]

    result = parse(rows)

    assert result.status is BalanceStatus.ACTIONABLE
    assert [(balance.supplier_rvp, balance.ending_debit) for balance in result.balances] == [
        ("Supplier A", Decimal("10.00")),
        ("Supplier B", Decimal("20.00")),
        ("Supplier A", Decimal("10.00")),
    ]
    assert [balance.source_excel_row_ref for balance in result.balances] == [
        "ОСВ!R9",
        "ОСВ!R10",
        "ОСВ!R14",
    ]


def test_indentation_recovers_supplier_department_and_organization_boundaries():
    rows = [
        ("79.2", None, None, 0),
        ("АТ", None, None, 1),
        ("Коммерческий отдел", None, None, 2),
        ("Поставщик-1", "1.00", 0, 3),
        ("Поставщик-2", 0, "2.00", 3),
        ("Финансовый отдел", None, None, 2),
        ("Поставщик-3", "3.00", 0, 3),
        ("БТ", None, None, 1),
        ("Коммерческий отдел", None, None, 2),
        ("Поставщик-4", 0, "4.00", 3),
    ]

    result = parse(rows)

    assert result.status is BalanceStatus.ACTIONABLE
    assert [(b.organization, b.department, b.supplier_rvp) for b in result.balances] == [
        ("АТ", "Коммерческий отдел", "Поставщик-1"),
        ("АТ", "Коммерческий отдел", "Поставщик-2"),
        ("АТ", "Финансовый отдел", "Поставщик-3"),
        ("БТ", "Коммерческий отдел", "Поставщик-4"),
    ]


def test_account_boundary_resets_stale_context_and_missing_context_blocks():
    rows = leaf("79.2", "АТ", "ЦФО-1", "Поставщик-1", "1.00")
    rows += [
        ("79.3", None, None, 0),
        ("ЦФО: ЦФО-2", None, None, 2),
        ("Поставщик РВП: Поставщик-2", "2.00", 0, 3),
    ]

    result = parse(rows)

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert result.diagnostics[0].code is ParserDiagnosticCode.MISSING_ORGANIZATION


@pytest.mark.parametrize("debit,credit", [("100.00", 0), (0, "100.00")])
def test_financial_payload_with_blank_grouping_before_any_account_blocks(debit, credit):
    rows = [("", debit, credit, None)]
    rows += leaf("79.2", "АТ", "ЦФО", "После строки", "1.00")

    result = parse(rows)

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert result.diagnostics[0].code is ParserDiagnosticCode.MISSING_ACCOUNT_CONTEXT
    assert result.diagnostics[0].excel_row == 6


def test_financial_payload_after_unsupported_account_has_explicit_context_diagnostic():
    rows = [
        ("80.1", None, None, 0),
        ("", "100.00", 0, None),
    ]
    rows += leaf("79.2", "АТ", "ЦФО", "После строки", "1.00")

    result = parse(rows)

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert any(
        diagnostic.code is ParserDiagnosticCode.UNSUPPORTED_ACCOUNT
        for diagnostic in result.diagnostics
    )
    assert any(
        diagnostic.code is ParserDiagnosticCode.MISSING_ACCOUNT_CONTEXT
        and diagnostic.excel_row == 7
        for diagnostic in result.diagnostics
    )


def test_financial_payload_after_hierarchy_reset_with_incomplete_identity_blocks():
    rows = leaf("79.2", "АТ", "ЦФО", "До сброса", "1.00")
    rows += [
        ("Организация: Новая АТ", None, None, 1),
        ("", "100.00", 0, None),
    ]

    result = parse(rows)

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert any(
        diagnostic.code is ParserDiagnosticCode.AMBIGUOUS_HIERARCHY
        and diagnostic.excel_row == 11
        for diagnostic in result.diagnostics
    )


def test_bare_organization_boundary_invalidates_stale_context():
    rows = leaf("79.2", "A", "D", "S", "1.00")
    rows += [
        ("Организация", None, None, 1),
        ("Поставщик РВП: T", "2.00", 0, 3),
    ]

    result = parse(rows)

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert result.diagnostics[-1].code is ParserDiagnosticCode.MISSING_ORGANIZATION


def test_bare_department_boundary_exports_blank_without_stale_context():
    rows = leaf("79.2", "A", "D", "S", "1.00")
    rows += [
        ("Подразделение", None, None, 2),
        ("Поставщик РВП: T", "2.00", 0, 3),
    ]

    result = parse(rows)

    assert result.status is BalanceStatus.ACTIONABLE
    assert [(b.department, b.supplier_rvp) for b in result.balances] == [
        ("D", "S"),
        ("", "T"),
    ]


def test_new_department_cannot_repair_an_incomplete_organization_boundary():
    rows = leaf("79.2", "A", "D", "S", "1.00")
    rows += [
        ("Организация", None, None, 1),
        ("ЦФО: E", None, None, 2),
        ("Поставщик РВП: T", "2.00", 0, 3),
    ]

    result = parse(rows)

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert any(
        diagnostic.code is ParserDiagnosticCode.MISSING_ORGANIZATION
        for diagnostic in result.diagnostics
    )


def test_supplier_payload_after_incomplete_department_boundary_keeps_supplier():
    rows = leaf("79.2", "A", "D", "S", "1.00")
    rows += [
        ("ЦФО", None, None, 2),
        ("Поставщик РВП: T", "2.00", 0, 3),
    ]

    result = parse(rows)

    assert result.status is BalanceStatus.ACTIONABLE
    assert result.balances[-1].department == ""
    assert result.balances[-1].supplier_rvp == "T"


def test_complete_hierarchy_after_incomplete_boundary_restores_only_new_context():
    rows = [
        ("79.2", None, None, 0),
        ("Организация: A", None, None, 1),
        ("ЦФО: D", None, None, 2),
        ("Организация", None, None, 1),
        ("Организация: B", None, None, 1),
        ("ЦФО: E", None, None, 2),
        ("Поставщик РВП: T", "2.00", 0, 3),
    ]

    result = parse(rows)

    assert result.status is BalanceStatus.ACTIONABLE
    assert [(b.organization, b.department, b.supplier_rvp) for b in result.balances] == [
        ("B", "E", "T")
    ]


def test_non_hierarchy_presentation_text_does_not_reset_valid_context():
    workbook = make_workbook(leaf("79.2", "A", "D", "S", "1.00"))
    worksheet = workbook.active
    worksheet.cell(10, 1).value = None
    worksheet.cell(10, 2).value = "Сформировано автоматически"

    result = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.ACTIONABLE
    assert [(b.organization, b.department, b.supplier_rvp) for b in result.balances] == [
        ("A", "D", "S")
    ]


@pytest.mark.parametrize("source", [b"not xlsx", b"PK\x03\x04", b""])
def test_malformed_xlsx_bytes_return_invalid_source_blocked(source):
    result = parse_grouped_osv(source, period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert result.diagnostics[0].code is ParserDiagnosticCode.INVALID_SOURCE
    assert result.message == "source workbook is not a valid XLSX file"


def test_non_integer_worksheet_sheet_id_returns_invalid_source_blocked():
    source = workbook_bytes(make_workbook(leaf("79.2", "A", "D", "S", "1.00")))
    malformed_source = BytesIO()

    with (
        zipfile.ZipFile(BytesIO(source)) as workbook_archive,
        zipfile.ZipFile(malformed_source, "w") as malformed_archive,
    ):
        for info in workbook_archive.infolist():
            content = workbook_archive.read(info.filename)
            if info.filename == "xl/workbook.xml":
                assert b'sheetId="1"' in content
                content = content.replace(
                    b'sheetId="1"', b'sheetId="not-an-integer"', 1
                )
            malformed_archive.writestr(info, content)

    result = parse_grouped_osv(malformed_source.getvalue(), period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert result.diagnostics[0].code is ParserDiagnosticCode.INVALID_SOURCE
    assert result.message == "source workbook is not a valid XLSX file"


def test_valid_xlsx_bytes_still_parse_normally():
    source = workbook_bytes(make_workbook(leaf("79.2", "A", "D", "S", "1.00")))

    result = parse_grouped_osv(source, period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.ACTIONABLE
    assert [(b.organization, b.department, b.supplier_rvp) for b in result.balances] == [
        ("A", "D", "S")
    ]


def test_internal_type_error_is_not_converted_to_invalid_source(monkeypatch):
    workbook = make_workbook(leaf("79.2", "A", "D", "S", "1.00"))

    def raise_internal_type_error(_parser, _worksheet):
        raise TypeError("internal parser bug")

    monkeypatch.setattr(GroupedOsvParser, "_parse_worksheet", raise_internal_type_error)

    with pytest.raises(TypeError, match="internal parser bug"):
        GroupedOsvParser(period_end=date(2024, 12, 31)).parse(workbook)


def test_blank_non_financial_layout_row_can_be_skipped():
    rows = [("", 0, 0, None)]
    rows += leaf("79.2", "АТ", "ЦФО", "Поставщик", "1.00")

    result = parse(rows)

    assert result.status is BalanceStatus.ACTIONABLE
    assert len(result.balances) == 1


@pytest.mark.parametrize(
    "rows,code",
    [
        (
            [
                ("79.2", None, None, 0),
                ("АТ", None, None, 1),
                ("Итого", "100.00", 0, 2),
            ],
            ParserDiagnosticCode.MISSING_DEPARTMENT,
        ),
        (
            [
                ("79.2", None, None, 0),
                ("АТ", None, None, 1),
                ("ЦФО: ЦФО", None, None, 2),
                ("ОВ", "100.00", 0, 4),
            ],
            ParserDiagnosticCode.MISSING_SUPPLIER_RVP,
        ),
    ],
)
def test_financial_payload_on_incomplete_non_leaf_context_blocks(rows, code):
    result = parse(rows)

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert any(diagnostic.code is code for diagnostic in result.diagnostics)


@pytest.mark.parametrize("account", ["79.2", "79.3"])
def test_explicit_hierarchy_aggregates_are_not_supplier_financial_leaves(account):
    result = parse(
        [
            (account, "400.00", 0, 0),
            ("Организация: АТ", "300.00", 0, 1),
            ("ЦФО: ЦФО", "200.00", 0, 2),
            ("Поставщик РВП: Поставщик", "100.00", 0, 3),
        ]
    )

    assert result.status is BalanceStatus.ACTIONABLE
    assert result.diagnostics == ()
    assert len(result.balances) == 1
    assert result.balances[0].source_account.value == account
    assert result.balances[0].ending_debit == Decimal("100.00")


@pytest.mark.parametrize("account", ["79.2", "79.3"])
def test_supported_account_token_at_supplier_depth_with_payload_blocks_without_reset(account):
    rows = leaf(account, "ORG_A", "DEPARTMENT_A", "SUPPLIER_A", "1.00")
    rows += [
        (account, "2.00", 0, 3),
        ("Поставщик РВП: SUPPLIER_B", "3.00", 0, 3),
    ]

    result = parse(rows)

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert len(result.diagnostics) == 1
    diagnostic = result.diagnostics[0]
    assert diagnostic.code is ParserDiagnosticCode.AMBIGUOUS_HIERARCHY
    assert diagnostic.excel_row == 10
    assert "established account depth" in diagnostic.message


@pytest.mark.parametrize("depth", [1, 2])
def test_supported_account_token_at_other_non_account_depth_with_payload_blocks(depth):
    rows = leaf("79.2", "ORG_A", "DEPARTMENT_A", "SUPPLIER_A", "1.00")
    rows.append(("79.2", "2.00", 0, depth))

    result = parse(rows)

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert len(result.diagnostics) == 1
    assert result.diagnostics[0].code is ParserDiagnosticCode.AMBIGUOUS_HIERARCHY
    assert result.diagnostics[0].excel_row == 10


def test_unknown_financial_payload_after_complete_hierarchy_remains_fail_closed():
    rows = leaf("79.2", "АТ", "ЦФО", "Поставщик", "1.00")
    rows.append(("Неизвестная финансовая строка", "2.00", 0, 5))

    result = parse(rows)

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert result.diagnostics[0].code is ParserDiagnosticCode.AMBIGUOUS_HIERARCHY


def test_unlabelled_department_without_organization_blocks_instead_of_promoting_depth():
    rows = [
        ("79.2", None, None, 0),
        ("Department-looking row", None, None, 2),
        ("Supplier-1", "1.00", 0, 3),
        ("Supplier-2", "2.00", 0, 3),
    ]

    result = parse(rows)

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert any(
        diagnostic.code is ParserDiagnosticCode.AMBIGUOUS_HIERARCHY
        for diagnostic in result.diagnostics
    )


def test_new_unlabelled_organization_without_department_exports_blank_not_neighbor():
    rows = [
        ("79.2", None, None, 0),
        ("Organization-A", None, None, 1),
        ("Department-A", None, None, 2),
        ("Supplier-A", "1.00", 0, 3),
        ("Organization-B", None, None, 1),
        ("Supplier-B", "2.00", 0, 3),
    ]

    result = parse(rows)

    assert result.status is BalanceStatus.ACTIONABLE
    assert [(b.organization, b.department, b.supplier_rvp) for b in result.balances] == [
        ("Organization-A", "Department-A", "Supplier-A"),
        ("Organization-B", "", "Supplier-B"),
    ]


def test_missing_depth_metadata_blocks_unlabelled_supplier_boundary():
    rows = leaf("79.2", "АТ", "ЦФО", "Поставщик-1", "1.00")
    rows += [("Поставщик-2", "2.00", 0, None)]

    result = parse(rows)

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert result.diagnostics[-1].code is ParserDiagnosticCode.AMBIGUOUS_HIERARCHY


def test_wrong_hierarchy_depth_sequence_blocks_value_bearing_row():
    rows = [
        ("79.2", None, None, 0),
        ("Organization-A", None, None, 1),
        ("Department-A", None, None, 2),
        ("Supplier-A", "1.00", 0, 3),
        ("Wrong-depth-row", "2.00", 0, 4),
    ]

    result = parse(rows)

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert result.diagnostics[-1].code is ParserDiagnosticCode.AMBIGUOUS_HIERARCHY


def test_blank_supplier_is_exportable_but_ambiguous_identity_remains_blocked():
    blank = parse(leaf("79.2", "АТ", "ЦФО", "", "1.00"))
    assert blank.status is BalanceStatus.ACTIONABLE
    assert blank.diagnostics == ()
    assert blank.balances[0].supplier_rvp == ""

    ambiguous_rows = leaf("79.2", "АТ", "ЦФО", "Производитель", "1.00")
    workbook = make_workbook(ambiguous_rows)
    worksheet = workbook.active
    worksheet.cell(worksheet.max_row - 2, 1).value = "Организация"
    worksheet.cell(worksheet.max_row - 2, 2).value = "АТ"
    worksheet.cell(worksheet.max_row - 2, 3).value = "БТ"
    ambiguous = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))
    assert ambiguous.status is BalanceStatus.BLOCKED
    assert ambiguous.diagnostics[0].code is ParserDiagnosticCode.AMBIGUOUS_ORGANIZATION


@pytest.mark.parametrize(
    "supported,unsupported",
    [
        ("79.2", "80.1"),
        ("79.2", "79.4"),
        ("79.3", "62.1"),
        ("79.3", "179.2"),
        ("79.2", "79.2x"),
    ],
)
def test_unsupported_account_is_not_accepted_or_inherited(supported, unsupported):
    rows = leaf(supported, "АТ", "ЦФО", "До границы", "1.00")
    rows += [
        (unsupported, None, None, 0),
        ("Организация: АТ-unsupported", None, None, 1),
        ("ЦФО: ЦФО-unsupported", None, None, 2),
        ("Поставщик РВП: После границы", "2.00", 0, 3),
    ]
    result = parse(rows)

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert any(
        diagnostic.code is ParserDiagnosticCode.UNSUPPORTED_ACCOUNT
        for diagnostic in result.diagnostics
    )


def test_semantic_header_detection_is_independent_of_header_row_number():
    result = parse(
        leaf("79.3", "АТ", "ЦФО", "Поставщик", "12.30"),
        header_row=12,
    )

    assert result.status is BalanceStatus.ACTIONABLE
    assert result.balances[0].ending_debit == Decimal("12.30")


def test_account_label_and_exact_account_value_can_share_grouping_columns():
    workbook = make_workbook(leaf("79.2", "АТ", "ЦФО", "Поставщик", "12.30"))
    worksheet = workbook.active
    worksheet.cell(6, 1).value = "Счет"
    worksheet.cell(6, 2).value = "79.2"

    result = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.ACTIONABLE
    assert result.balances[0].source_account.value == "79.2"
    assert result.balances[0].organization == "АТ"


def make_structured_measure_workbook(*, grouping_column: int = 1) -> Workbook:
    """Build an actual-shape grouped OSV with several numeric sections."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "ОСВ"
    worksheet.merge_cells("E8:F8")
    worksheet["E8"] = "Сальдо на начало периода"
    worksheet["E9"] = "Дебет"
    worksheet["F9"] = "Кредит"
    worksheet.merge_cells("G8:H8")
    worksheet["G8"] = "Обороты за период"
    worksheet["G9"] = "Дебет"
    worksheet["H9"] = "Кредит"
    worksheet.merge_cells("I8:J8")
    worksheet["I8"] = "Обороты за период 2"
    worksheet["I9"] = "Дебет"
    worksheet["J9"] = "Кредит"
    worksheet.merge_cells("K8:L8")
    worksheet["K8"] = "Сальдо на конец периода"
    worksheet["K9"] = "Дебет"
    worksheet["L9"] = "Кредит"

    rows = [
        (79.2, 1.20, 79.30, 62.10, 179.20, 0, 0, 0),
        ("Организация: Synthetic Org", 2.25, 0, 0, 0, 0, 0, 0),
        ("ЦФО: Synthetic Department", 0, 3.50, 0, 0, 0, 0, 0),
        ("Поставщик РВП: Synthetic Supplier", 0, 0, 4.75, 0, 10.00, 0, 0),
    ]
    for row, values in enumerate(rows, 10):
        worksheet.cell(row, grouping_column).value = values[0]
        worksheet.cell(row, grouping_column).alignment = Alignment(
            indent=min(row - 10, 3)
        )
        for column, value in zip((5, 7, 9, 10, 11, 12, 6, 8), values[1:]):
            worksheet.cell(row, column).value = value
    return workbook


def make_merged_hierarchy_workbook(*, account: str = "79.2") -> Workbook:
    """Build merged hierarchy pairs with outline and visual indent on different scales."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Лист_1"
    worksheet.merge_cells("K8:L8")
    worksheet["K8"] = "Сальдо на конец периода"
    worksheet["K9"] = "Дебет"
    worksheet["L9"] = "Кредит"

    hierarchy = (
        (10, account, "400.00", 0, 0),
        (12, "ORG_A", "300.00", 2, 1),
        (14, "DEPARTMENT_A", "200.00", 4, 2),
        (16, "SUPPLIER_A", "100.00", 6, 3),
    )
    for row, identity, debit, indent, outline_level in hierarchy:
        worksheet.merge_cells(
            start_row=row,
            start_column=3,
            end_row=row + 1,
            end_column=5,
        )
        cell = worksheet.cell(row, 3)
        cell.value = identity
        cell.alignment = Alignment(indent=indent)
        worksheet.cell(row, 6).value = "ОВ"
        worksheet.cell(row + 1, 6).value = "ФВ"
        worksheet.cell(row, 11).value = debit
        worksheet.cell(row, 12).value = 0
        worksheet.cell(row + 1, 11).value = debit
        worksheet.cell(row + 1, 12).value = 0
        worksheet.row_dimensions[row].outlineLevel = outline_level
        worksheet.row_dimensions[row + 1].outlineLevel = outline_level
    return workbook


def append_root_grand_total_pair(
    workbook: Workbook,
    *,
    debit: object = "400.00",
    credit: object = 0,
    turnover_debit: object = "375.00",
    continuation_debit: object | None = None,
    continuation_credit: object | None = None,
    continuation_turnover_debit: object | None = None,
) -> None:
    """Append the real-source shape of a merged root OV/FV report total."""

    worksheet = workbook.active
    worksheet.merge_cells("G8:H8")
    worksheet["G8"] = "Сальдо на начало периода"
    worksheet["G9"] = "Дебет"
    worksheet["H9"] = "Кредит"
    worksheet.merge_cells("I8:J8")
    worksheet["I8"] = "Обороты за период"
    worksheet["I9"] = "Дебет"
    worksheet["J9"] = "Кредит"
    for aggregate_row in (10, 11):
        worksheet.cell(aggregate_row, 7).value = "25.00"
        worksheet.cell(aggregate_row, 8).value = 0
        worksheet.cell(aggregate_row, 9).value = "375.00"
        worksheet.cell(aggregate_row, 10).value = 0

    row = worksheet.max_row + 1
    worksheet.merge_cells(
        start_row=row,
        start_column=3,
        end_row=row + 1,
        end_column=5,
    )
    worksheet.cell(row, 3).value = "Итого"
    worksheet.cell(row, 3).alignment = Alignment(indent=0)
    worksheet.cell(row, 6).value = "ОВ"
    worksheet.cell(row + 1, 6).value = "ФВ"
    worksheet.cell(row, 7).value = "25.00"
    worksheet.cell(row, 8).value = 0
    worksheet.cell(row, 9).value = turnover_debit
    worksheet.cell(row, 10).value = 0
    worksheet.cell(row, 11).value = debit
    worksheet.cell(row, 12).value = credit
    worksheet.cell(row + 1, 7).value = "25.00"
    worksheet.cell(row + 1, 8).value = 0
    worksheet.cell(row + 1, 9).value = (
        turnover_debit
        if continuation_turnover_debit is None
        else continuation_turnover_debit
    )
    worksheet.cell(row + 1, 10).value = 0
    worksheet.cell(row + 1, 11).value = (
        debit if continuation_debit is None else continuation_debit
    )
    worksheet.cell(row + 1, 12).value = (
        credit if continuation_credit is None else continuation_credit
    )
    worksheet.row_dimensions[row].outlineLevel = 0
    worksheet.row_dimensions[row + 1].outlineLevel = 0


def test_real_r94_shape_exports_blank_department_and_supplier():
    workbook = make_merged_hierarchy_workbook()
    worksheet = workbook.active
    for row, identity, debit, indent, outline_level in (
        (18, "ORG_B", 0, 2, 1),
        (20, None, 0, 4, 2),
        (22, None, "100.00", 6, 3),
    ):
        worksheet.merge_cells(
            start_row=row,
            start_column=3,
            end_row=row + 1,
            end_column=5,
        )
        cell = worksheet.cell(row, 3)
        cell.value = identity
        cell.alignment = Alignment(indent=indent)
        worksheet.cell(row, 6).value = "ОВ"
        worksheet.cell(row + 1, 6).value = "ФВ"
        worksheet.cell(row, 11).value = debit
        worksheet.cell(row, 12).value = 0
        worksheet.cell(row + 1, 11).value = debit
        worksheet.cell(row + 1, 12).value = 0
        worksheet.row_dimensions[row].outlineLevel = outline_level
        worksheet.row_dimensions[row + 1].outlineLevel = outline_level

    result = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.ACTIONABLE
    assert result.diagnostics == ()
    assert len(result.balances) == 2
    balance = next(balance for balance in result.balances if balance.organization == "ORG_B")
    assert balance.department == ""
    assert balance.supplier_rvp == ""
    assert balance.ending_debit == Decimal("100.00")
    assert balance.source_excel_row_ref == "Лист_1!R22"


@pytest.mark.parametrize(
    ("department", "supplier"),
    (("", "SUPPLIER_A"), ("DEPARTMENT_A", ""), ("", "")),
)
def test_structural_supplier_leaf_preserves_known_and_blank_lower_analytics(
    department, supplier
):
    rows = [
        ("79.2", None, None, 0),
        ("ORG_A", None, None, 1),
        (department, None, None, 2),
        (supplier, "100.00", 0, 3),
    ]

    result = parse(rows)

    assert result.status is BalanceStatus.ACTIONABLE
    assert result.diagnostics == ()
    assert len(result.balances) == 1
    assert result.balances[0].organization == "ORG_A"
    assert result.balances[0].department == department
    assert result.balances[0].supplier_rvp == supplier


def test_blank_lower_analytics_never_inherit_neighbor_values():
    rows = [
        ("79.2", None, None, 0),
        ("ORG_A", None, None, 1),
        ("DEPARTMENT_A", None, None, 2),
        ("SUPPLIER_A", "10.00", 0, 3),
        ("", None, None, 2),
        ("", "20.00", 0, 3),
    ]

    result = parse(rows)

    assert result.status is BalanceStatus.ACTIONABLE
    assert [(b.department, b.supplier_rvp) for b in result.balances] == [
        ("DEPARTMENT_A", "SUPPLIER_A"),
        ("", ""),
    ]


def test_structural_leaf_without_organization_remains_blocked():
    result = parse(
        [
            ("79.2", None, None, 0),
            ("", None, None, 1),
            ("", None, None, 2),
            ("Поставщик РВП", "100.00", 0, 3),
        ]
    )

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert result.diagnostics[0].code is ParserDiagnosticCode.MISSING_ORGANIZATION


@pytest.mark.parametrize(
    ("role_label", "row_indent", "expected_code"),
    (
        ("ЦФО", 2, ParserDiagnosticCode.AMBIGUOUS_DEPARTMENT),
        ("Поставщик РВП", 3, ParserDiagnosticCode.AMBIGUOUS_SUPPLIER_RVP),
    ),
)
def test_multiple_lower_identity_candidates_remain_blocked(
    role_label, row_indent, expected_code
):
    rows = [
        ("79.2", None, None, 0),
        ("Организация: ORG_A", None, None, 1),
    ]
    if row_indent == 3:
        rows.append(("ЦФО: DEPARTMENT_A", None, None, 2))
    rows.append((role_label, "100.00", 0, row_indent))
    workbook = make_workbook(rows)
    worksheet = workbook.active
    worksheet.cell(worksheet.max_row, 2).value = "CANDIDATE_A"
    worksheet.cell(worksheet.max_row, 3).value = "CANDIDATE_B"

    result = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert result.diagnostics[-1].code is expected_code


@pytest.mark.parametrize("account", ["79.2", "79.3"])
def test_merged_outline_hierarchy_aggregates_are_context_not_financial_leaves(account):
    workbook = make_merged_hierarchy_workbook(account=account)

    result = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))

    assert result.diagnostics == ()
    assert result.status is BalanceStatus.ACTIONABLE
    assert len(result.balances) == 1
    balance = result.balances[0]
    assert balance.source_account.value == account
    assert balance.organization == "ORG_A"
    assert balance.department == "DEPARTMENT_A"
    assert balance.supplier_rvp == "SUPPLIER_A"
    assert balance.ending_debit == Decimal("100.00")
    assert balance.source_excel_row_ref == "Лист_1!R16"


@pytest.mark.parametrize("account", ["79.2", "79.3"])
def test_root_grand_total_ov_fv_pair_is_not_a_financial_leaf(account):
    workbook = make_merged_hierarchy_workbook(account=account)
    append_root_grand_total_pair(workbook)

    result = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.ACTIONABLE
    assert result.diagnostics == ()
    assert len(result.balances) == 1
    assert result.balances[0].supplier_rvp == "SUPPLIER_A"
    assert result.balances[0].source_excel_row_ref == "Лист_1!R16"
    batch = TransferEngine().generate_batch(result.balances)
    assert batch.status is BalanceStatus.ACTIONABLE
    assert len(batch.rows) == 2
    assert {row.source_excel_row_ref for row in batch.rows} == {"Лист_1!R16"}


@pytest.mark.parametrize(
    "mutate",
    [
        lambda workbook: append_root_grand_total_pair(workbook, debit="399.00"),
        lambda workbook: append_root_grand_total_pair(
            workbook,
            turnover_debit="374.00",
        ),
        lambda workbook: append_root_grand_total_pair(
            workbook,
            continuation_debit="401.00",
        ),
        lambda workbook: append_root_grand_total_pair(
            workbook,
            continuation_turnover_debit="376.00",
        ),
    ],
    ids=[
        "ending-not-account-aggregate",
        "turnover-not-account-aggregate",
        "conflicting-ending-ov-fv",
        "conflicting-turnover-ov-fv",
    ],
)
def test_unproven_root_total_pair_remains_fail_closed(mutate):
    workbook = make_merged_hierarchy_workbook()
    mutate(workbook)

    result = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert any(
        diagnostic.code is ParserDiagnosticCode.AMBIGUOUS_HIERARCHY
        for diagnostic in result.diagnostics
    )


@pytest.mark.parametrize("grouping", ["Unproven root payload", ""])
def test_ordinary_root_financial_payload_is_not_silently_skipped(grouping):
    workbook = make_merged_hierarchy_workbook()
    worksheet = workbook.active
    row = worksheet.max_row + 1
    worksheet.cell(row, 3).value = grouping
    worksheet.cell(row, 3).alignment = Alignment(indent=0)
    worksheet.cell(row, 11).value = "400.00"
    worksheet.cell(row, 12).value = 0
    worksheet.row_dimensions[row].outlineLevel = 0

    result = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert any(
        diagnostic.code is ParserDiagnosticCode.AMBIGUOUS_HIERARCHY
        and diagnostic.excel_row == row
        for diagnostic in result.diagnostics
    )


def test_merged_technical_continuation_with_conflicting_payload_remains_fail_closed():
    workbook = make_merged_hierarchy_workbook()
    workbook.active.cell(17, 11).value = "101.00"

    result = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert any(diagnostic.excel_row == 17 for diagnostic in result.diagnostics)


def test_structural_grouping_ignores_decimal_measure_candidates():
    workbook = make_structured_measure_workbook()
    layout = _detect_ending_columns(workbook.active)
    assert not isinstance(layout, ParserDiagnostic)
    assert _find_grouping_column(workbook.active, layout) == 1

    result = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.ACTIONABLE
    assert len(result.balances) == 1
    assert result.balances[0].source_account.value == "79.2"
    assert result.balances[0].organization == "Synthetic Org"
    assert result.balances[0].department == "Synthetic Department"
    assert result.balances[0].supplier_rvp == "Synthetic Supplier"
    assert result.balances[0].ending_debit == Decimal("10.00")


def test_shifted_grouping_column_is_detected_from_hierarchy_structure():
    workbook = make_structured_measure_workbook(grouping_column=3)
    layout = _detect_ending_columns(workbook.active)
    assert _find_grouping_column(workbook.active, layout) == 3

    result = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.ACTIONABLE
    assert result.balances[0].supplier_rvp == "Synthetic Supplier"


def test_outline_only_hierarchy_supports_grouping_column_detection():
    workbook = make_workbook(
        [
            ("79.2", None, None, None),
            ("Organization-A", None, None, None),
            ("Department-A", None, None, None),
            ("Supplier-A", "1.00", 0, None),
        ]
    )
    worksheet = workbook.active
    for row, outline_level in ((7, 1), (8, 2), (9, 3)):
        worksheet.row_dimensions[row].outlineLevel = outline_level

    result = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.ACTIONABLE
    assert result.diagnostics == ()
    assert len(result.balances) == 1
    assert result.balances[0].organization == "Organization-A"
    assert result.balances[0].department == "Department-A"
    assert result.balances[0].supplier_rvp == "Supplier-A"


def test_outline_hierarchy_excludes_numeric_measure_account_lookalikes():
    workbook = make_structured_measure_workbook()
    worksheet = workbook.active
    worksheet["A11"] = "Organization-A"
    worksheet["A12"] = "Department-A"
    worksheet["A13"] = "Supplier-A"
    for row, outline_level in ((10, 0), (11, 1), (12, 2), (13, 3)):
        worksheet.cell(row, 1).alignment = Alignment(indent=0)
        worksheet.row_dimensions[row].outlineLevel = outline_level

    layout = _detect_ending_columns(worksheet)
    assert not isinstance(layout, ParserDiagnostic)
    assert _find_grouping_column(worksheet, layout) == 1

    result = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.ACTIONABLE
    assert len(result.balances) == 1
    assert result.balances[0].source_account.value == "79.2"
    assert result.balances[0].organization == "Organization-A"
    assert result.balances[0].department == "Department-A"
    assert result.balances[0].supplier_rvp == "Supplier-A"


def test_numeric_measure_value_equal_to_supported_account_is_not_an_account_boundary():
    workbook = make_structured_measure_workbook()
    worksheet = workbook.active
    worksheet["E10"] = 79.2
    worksheet["G10"] = "1.20"
    worksheet["I10"] = "79.30"
    worksheet["J10"] = "62.10"

    layout = _detect_ending_columns(worksheet)
    assert not isinstance(layout, ParserDiagnostic)
    assert _find_grouping_column(worksheet, layout) == 1
    result = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.ACTIONABLE
    assert len(result.balances) == 1


def test_two_structurally_plausible_grouping_columns_remain_ambiguous():
    workbook = make_structured_measure_workbook()
    worksheet = workbook.active
    for row in range(10, 14):
        worksheet.cell(row, 3).value = worksheet.cell(row, 1).value
        worksheet.cell(row, 3).alignment = Alignment(indent=min(row - 10, 3))

    layout = _detect_ending_columns(worksheet)
    grouping = _find_grouping_column(worksheet, layout)

    assert grouping.code is ParserDiagnosticCode.AMBIGUOUS_ACCOUNT_CONTEXT
    result = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))
    assert result.status is BalanceStatus.BLOCKED
    assert result.diagnostics[0].code is ParserDiagnosticCode.AMBIGUOUS_ACCOUNT_CONTEXT


def test_missing_or_ambiguous_ending_balance_headers_block():
    workbook = make_workbook(leaf("79.2", "АТ", "ЦФО", "Поставщик", "1.00"))
    workbook.active.cell(5, 9).value = None
    missing = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))
    assert missing.status is BalanceStatus.BLOCKED
    assert missing.diagnostics[0].code is ParserDiagnosticCode.MISSING_ENDING_BALANCE_HEADERS

    ambiguous_workbook = make_workbook(leaf("79.2", "АТ", "ЦФО", "Поставщик", "1.00"))
    ambiguous_workbook.active.merge_cells("K4:L4")
    ambiguous_workbook.active["K4"] = "Сальдо на конец периода"
    ambiguous_workbook.active["K5"] = "Дебет"
    ambiguous_workbook.active["L5"] = "Кредит"
    ambiguous = parse_grouped_osv(ambiguous_workbook, period_end=date(2024, 12, 31))
    assert ambiguous.status is BalanceStatus.BLOCKED
    assert ambiguous.diagnostics[0].code is ParserDiagnosticCode.AMBIGUOUS_ENDING_BALANCE_HEADERS


def test_unrelated_nearby_side_label_cannot_complete_missing_header():
    missing_credit_workbook = make_workbook(
        leaf("79.2", "АТ", "ЦФО", "Поставщик", "1.00")
    )
    missing_credit_workbook.active["I5"] = None
    missing_credit_workbook.active["J5"] = "Кредит"
    missing_credit = parse_grouped_osv(
        missing_credit_workbook,
        period_end=date(2024, 12, 31),
    )

    missing_debit_workbook = make_workbook(
        leaf("79.2", "АТ", "ЦФО", "Поставщик", "1.00")
    )
    missing_debit_workbook.active["H5"] = None
    missing_debit_workbook.active["J5"] = "Дебет"
    missing_debit = parse_grouped_osv(
        missing_debit_workbook,
        period_end=date(2024, 12, 31),
    )

    assert missing_credit.status is BalanceStatus.BLOCKED
    assert missing_credit.diagnostics[0].code is ParserDiagnosticCode.MISSING_ENDING_BALANCE_HEADERS
    assert missing_debit.status is BalanceStatus.BLOCKED
    assert missing_debit.diagnostics[0].code is ParserDiagnosticCode.MISSING_ENDING_BALANCE_HEADERS


@pytest.mark.parametrize(
    "first_side,unrelated_side,second_unrelated_side",
    [("Дебет", "Кредит", "Дебет"), ("Кредит", "Дебет", "Кредит")],
)
def test_unmerged_header_children_are_anchored_to_the_semantic_parent(
    first_side, unrelated_side, second_unrelated_side
):
    workbook = make_workbook(leaf("79.2", "АТ", "ЦФО", "Поставщик", "1.00"))
    worksheet = workbook.active
    worksheet.unmerge_cells("H4:I4")
    worksheet["H4"] = "Сальдо на конец периода"
    worksheet["H5"] = first_side
    worksheet["I5"] = None
    worksheet["J5"] = unrelated_side
    worksheet["K5"] = second_unrelated_side

    result = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.BLOCKED
    assert result.diagnostics[0].code is ParserDiagnosticCode.MISSING_ENDING_BALANCE_HEADERS


def test_unmerged_header_children_support_one_column_shift():
    workbook = make_workbook(leaf("79.2", "АТ", "ЦФО", "Поставщик", "1.00"))
    worksheet = workbook.active
    worksheet.unmerge_cells("H4:I4")
    worksheet["H4"] = "Сальдо на конец периода"
    worksheet["H5"] = None
    worksheet["I5"] = "Дебет"
    worksheet["J5"] = "Кредит"
    for row in range(6, worksheet.max_row + 1):
        worksheet.cell(row, 10).value = worksheet.cell(row, 9).value
        worksheet.cell(row, 9).value = worksheet.cell(row, 8).value
        worksheet.cell(row, 8).value = None

    result = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.ACTIONABLE
    assert len(result.balances) == 1
    assert result.balances[0].ending_debit == Decimal("1.00")


def test_duplicate_coherent_ending_balance_groups_sharing_columns_block():
    workbook = make_workbook(leaf("79.2", "АТ", "ЦФО", "Поставщик", "1.00"))
    worksheet = workbook.active
    worksheet.merge_cells("H12:I12")
    worksheet["H12"] = "Сальдо на конец периода"
    worksheet["H13"] = "Дебет"
    worksheet["I13"] = "Кредит"

    result = parse_grouped_osv(workbook, period_end=date(2024, 12, 31))

    assert result.status is BalanceStatus.BLOCKED
    assert result.diagnostics[0].code is ParserDiagnosticCode.AMBIGUOUS_ENDING_BALANCE_HEADERS


def test_source_row_reference_is_stable_and_path_independent():
    first = parse(leaf("79.2", "АТ", "ЦФО", "Поставщик", "1.00"))
    second = parse(leaf("79.2", "АТ", "ЦФО", "Поставщик", "1.00"))

    assert first.balances[0].source_excel_row_ref == second.balances[0].source_excel_row_ref
    assert first.balances[0].source_excel_row_ref == "ОСВ!R9"


def test_decimal_safe_service_and_negative_representations_are_normalized():
    result = parse(leaf("79.2", "АТ", "ЦФО", "Поставщик", "1 234,50"))
    assert result.status is BalanceStatus.ACTIONABLE
    assert result.balances[0].ending_debit == Decimal("1234.50")

    negative = parse(leaf("79.3", "АТ", "ЦФО", "Поставщик", "-7.25"))
    assert negative.status is BalanceStatus.BLOCKED
    assert negative.balances == ()
    assert negative.diagnostics[0].code is ParserDiagnosticCode.INVALID_ENDING_BALANCE

    both_sides = parse(leaf("79.3", "АТ", "ЦФО", "Поставщик", "7.25", "1.00"))
    assert both_sides.status is BalanceStatus.BLOCKED
    assert both_sides.diagnostics[0].code is ParserDiagnosticCode.INVALID_ENDING_BALANCE


@pytest.mark.parametrize("debit,credit", [("-100", 0), (0, "-100"), ("(100)", 0), (0, "(100)")])
def test_negative_ending_values_are_blocked_without_side_inference(debit, credit):
    result = parse(leaf("79.2", "АТ", "ЦФО", "Поставщик", debit, credit))

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert result.diagnostics[0].code is ParserDiagnosticCode.INVALID_ENDING_BALANCE


def test_equal_business_values_in_distinct_supplier_rows_are_preserved():
    rows = [
        ("79.2", None, None, 0),
        ("АТ", None, None, 1),
        ("Department A", None, None, 2),
        ("Поставщик", "1.00", 0, 3),
        ("Поставщик", "1.00", 0, 3),
    ]

    result = parse(rows)

    assert result.status is BalanceStatus.ACTIONABLE
    assert len(result.balances) == 2
    assert [balance.source_excel_row_ref for balance in result.balances] == [
        "ОСВ!R9",
        "ОСВ!R10",
    ]


def test_conflicting_presentation_duplicate_is_blocked():
    rows = leaf("79.2", "АТ", "ЦФО", "Производитель", "10.00")
    rows += [
        ("ОВ", "10.00", 0, 4),
        ("ФВ", "10.00", 0, 4),
        ("Итого по поставщику", "10.00", 0, 3),
        ("Поставщик РВП: Производитель", "11.00", 0, 3),
    ]

    result = parse(rows)

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert result.diagnostics[0].code is ParserDiagnosticCode.DUPLICATE_SOURCE_ROW


@pytest.mark.parametrize("supplier", ["ОВ", "ФВ", "ОВ-Company", "ФВ-Company"])
def test_supplier_names_similar_to_technical_rows_are_preserved(supplier):
    result = parse(
        [
            ("79.2", None, None, 0),
            ("АТ", None, None, 1),
            ("Department A", None, None, 2),
            (supplier, "1.00", 0, 3),
        ]
    )

    assert result.status is BalanceStatus.ACTIONABLE
    assert len(result.balances) == 1
    assert result.balances[0].supplier_rvp == supplier


def test_ambiguous_ov_fv_without_structural_depth_blocks():
    result = parse(
        [
            ("79.2", None, None, 0),
            ("АТ", None, None, 1),
            ("Department A", None, None, 2),
            ("ОВ", "1.00", 0, None),
        ]
    )

    assert result.status is BalanceStatus.BLOCKED
    assert result.balances == ()
    assert result.diagnostics[0].code is ParserDiagnosticCode.AMBIGUOUS_HIERARCHY
