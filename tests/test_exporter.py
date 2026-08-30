from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from holding79_transfer import (
    ExportBlockedError,
    FilenameCollisionError,
    NormalizedBalance,
    OutputAdapterConfig,
    PostingRow,
    RoundTripMismatchError,
    SourceAccount,
    deterministic_filename,
    export_posting_rows,
    generate_transfer,
    validate_workbook_round_trip,
)
from holding79_transfer.output import OUTPUT_HEADERS, OUTPUT_SHEET_NAME


def make_balance(
    *,
    organization: str = "АТ",
    period_end: date = date(2024, 12, 31),
    source_account: SourceAccount = SourceAccount.ACCOUNT_79_2,
    source_ref: str = "synthetic:1",
    amount: Decimal = Decimal("84272.40"),
) -> NormalizedBalance:
    return NormalizedBalance(
        period_end=period_end,
        source_excel_row_ref=source_ref,
        organization=organization,
        source_account=source_account,
        department=f"Б_{organization} Коммерческий отдел",
        supplier_rvp="Производитель",
        ending_debit=amount,
    )


def make_posting_row(
    *,
    organization: str,
    source_ref: str,
    period_end: date = date(2024, 12, 31),
) -> PostingRow:
    balance = make_balance(
        organization=organization,
        source_ref=source_ref,
        period_end=period_end,
    )
    return generate_transfer(balance).rows[0]


def read_rows(path: Path) -> tuple[list[str], list[tuple[object, ...]], list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        worksheet = workbook[OUTPUT_SHEET_NAME]
        values = list(worksheet.iter_rows(values_only=True))
        return list(values[0]), values[1:], workbook.sheetnames
    finally:
        workbook.close()


def test_export_has_exact_sheet_headers_and_no_helper_sheet(tmp_path: Path):
    row = generate_transfer(make_balance()).rows[0]

    result = export_posting_rows((row,), tmp_path)

    assert result.status.value == "ACTIONABLE"
    header, rows, sheetnames = read_rows(result.paths[0])
    assert sheetnames == [OUTPUT_SHEET_NAME]
    assert header == list(OUTPUT_HEADERS)
    assert len(header) == 27
    assert len(rows) == 1


def test_export_groups_by_document_organization_and_date(tmp_path: Path):
    first = generate_transfer(
        make_balance(organization="АТ", period_end=date(2024, 12, 31), source_ref="synthetic:a")
    ).rows
    second = generate_transfer(
        make_balance(organization="ББ", period_end=date(2024, 12, 31), source_ref="synthetic:b")
    ).rows
    third = generate_transfer(
        make_balance(organization="АТ", period_end=date(2025, 1, 31), source_ref="synthetic:c")
    ).rows

    result = export_posting_rows((*first, *second, *third), tmp_path)

    assert [workbook.document_organization for workbook in result.workbooks] == [
        "АТ",
        "ББ",
        "ГК",
        "АТ",
        "ГК",
    ]
    assert [workbook.document_date for workbook in result.workbooks] == [
        date(2024, 12, 31),
        date(2024, 12, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2025, 1, 31),
    ]
    assert {path.name for path in result.paths} == {
        "2024-12-31__ГК.xlsx",
        "2024-12-31__АТ.xlsx",
        "2024-12-31__ББ.xlsx",
        "2025-01-31__ГК.xlsx",
        "2025-01-31__АТ.xlsx",
    }

    _, december_gk_rows, _ = read_rows(tmp_path / "2024-12-31__ГК.xlsx")
    _, december_at_rows, _ = read_rows(tmp_path / "2024-12-31__АТ.xlsx")
    assert len(december_gk_rows) == 2
    assert len(december_at_rows) == 1
    assert all(row[0] == "79.1" for row in december_gk_rows)
    assert december_at_rows[0][0] == "79.1"


def test_export_row_order_and_filenames_are_deterministic(tmp_path: Path):
    rows = tuple(
        row
        for balance in (
            make_balance(source_ref="synthetic:2"),
            make_balance(source_ref="synthetic:1"),
        )
        for row in generate_transfer(balance).rows
    )
    config = OutputAdapterConfig(run_id="synthetic-run")
    first_dir = tmp_path / "first"
    second_dir = tmp_path / "second"

    first = export_posting_rows(rows, first_dir, config)
    second = export_posting_rows(tuple(reversed(rows)), second_dir, config)

    assert [path.name for path in first.paths] == [path.name for path in second.paths]
    for first_path, second_path in zip(first.paths, second.paths, strict=True):
        assert read_rows(first_path) == read_rows(second_path)


def test_export_maps_amount_rule_identity_source_trace_and_content(tmp_path: Path):
    row = generate_transfer(make_balance()).rows[0]
    config = OutputAdapterConfig(currency="RUB", operation_type="ADJUST", run_id="run-42")

    result = export_posting_rows((row,), tmp_path, config)
    header, values, _ = read_rows(result.paths[0])
    record = dict(zip(header, values[0], strict=True))

    assert record["ВидОперации"] == "ADJUST"
    assert Decimal(str(record["СуммаВВалютеУчета"])) == Decimal("84272.40")
    assert record["СчетДтИсточник"] == record["СчетКтИсточник"] == "79.2"
    assert record["ПравилоДт"] == record["ПравилоКт"] == "H79_DEBIT_TRANSFER_V1"
    assert record["ИдентификаторФинЗаписи"] == row.financial_record_id
    assert record["СубконтоДт1"] == "ГК"
    assert record["СубконтоКт1"] == "Производитель"
    assert record["Содержание"] == (
        "run_id=run-42; source_row_ref=synthetic:1; period_end=2024-12-31; "
        "source_organization=АТ; source_account=79.2; "
        "source_department=Б_АТ Коммерческий отдел; source_supplier_rvp=Производитель; "
        "document_organization=АТ; direction=DEBIT; contract_version=0.3-approved; "
        "rules_version=H79_TRANSFER_V1"
    )


def test_default_operation_type_is_repost_and_round_trip_passes(tmp_path: Path):
    row = generate_transfer(make_balance()).rows[0]

    result = export_posting_rows((row,), tmp_path)

    _, values, _ = read_rows(result.paths[0])
    assert values[0][4] == "REPOST"
    validate_workbook_round_trip(result.paths[0], (row,))


def test_round_trip_mismatch_fails_closed_and_publishes_no_file(tmp_path: Path, monkeypatch):
    row = generate_transfer(make_balance()).rows[0]
    import holding79_transfer.exporter as exporter_module

    original_write = exporter_module._write_workbook

    def corrupt_write(path, mapped_rows, amount_number_format):
        original_write(path, mapped_rows, amount_number_format)
        workbook = load_workbook(path)
        workbook.active["J2"] = "1.00"
        workbook.save(path)
        workbook.close()

    monkeypatch.setattr(exporter_module, "_write_workbook", corrupt_write)

    with pytest.raises(RoundTripMismatchError, match="row order/content mismatch|duplicate/unexpected"):
        export_posting_rows((row,), tmp_path)
    assert list(tmp_path.glob("*.xlsx")) == []


def test_duplicate_output_is_detected_by_round_trip_control(tmp_path: Path, monkeypatch):
    row = generate_transfer(make_balance()).rows[0]
    import holding79_transfer.exporter as exporter_module

    original_write = exporter_module._write_workbook

    def duplicate_write(path, mapped_rows, amount_number_format):
        original_write(path, mapped_rows, amount_number_format)
        workbook = load_workbook(path)
        worksheet = workbook.active
        worksheet.append([cell.value for cell in worksheet[2]])
        workbook.save(path)
        workbook.close()

    monkeypatch.setattr(exporter_module, "_write_workbook", duplicate_write)

    with pytest.raises(RoundTripMismatchError, match="financial rows|duplicate/unexpected"):
        export_posting_rows((row,), tmp_path)
    assert list(tmp_path.glob("*.xlsx")) == []


def test_filename_collision_fails_closed(tmp_path: Path):
    rows = (
        make_posting_row(organization="A/B", source_ref="synthetic:slash"),
        make_posting_row(organization="A\\B", source_ref="synthetic:backslash"),
    )

    with pytest.raises(FilenameCollisionError, match="filename collision"):
        export_posting_rows(rows, tmp_path)
    assert list(tmp_path.glob("*.xlsx")) == []
    assert deterministic_filename(date(2024, 12, 31), "A/B") == "2024-12-31__A_B.xlsx"


def test_incomplete_posting_is_blocked_before_any_file_is_published(tmp_path: Path):
    row = generate_transfer(make_balance()).rows[0].model_copy(
        update={"source_excel_row_ref": None}
    )

    with pytest.raises(ExportBlockedError):
        export_posting_rows((row,), tmp_path)
    assert list(tmp_path.glob("*.xlsx")) == []


def test_exported_values_have_expected_accounts_and_departments(tmp_path: Path):
    debit_row, credit_row = generate_transfer(
        make_balance(source_account=SourceAccount.ACCOUNT_79_3)
    ).rows

    result = export_posting_rows((debit_row, credit_row), tmp_path)

    _, values, _ = read_rows(result.paths[0])
    assert len(values) == 1
    # The source row is in the source organization's workbook; its exact
    # debit/credit mapping remains the accepted Issue #2 direction.
    _, source_values, _ = read_rows(tmp_path / "2024-12-31__АТ.xlsx")
    assert source_values[0][0:2] == ("79.1", "79.3")
    assert source_values[0][5:7] == (
        "Б_АТ Коммерческий отдел",
        "Б_АТ Коммерческий отдел",
    )
    assert source_values[0][21] == "ГК"
    assert source_values[0][24] == "Производитель"
